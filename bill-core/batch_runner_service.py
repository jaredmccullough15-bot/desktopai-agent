from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from pathlib import Path
from threading import RLock
from typing import Any
from uuid import uuid4

from ci_checks_decision import evaluate_paid_through_date

BATCH_RUNS_FILENAME = "batch_runs.json"

_BATCH_LOCK = RLock()


def _iso_now() -> str:
    return datetime.utcnow().isoformat() + "Z"


def _parse_iso_datetime(value: object) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None


def _sanitize_header(name: object, index: int) -> str:
    raw = str(name or "").strip().lower()
    normalized = "".join(c if c.isalnum() else "_" for c in raw).strip("_")
    return normalized or f"column_{index + 1}"


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_blank_row(values: list[str]) -> bool:
    return all(not str(value or "").strip() for value in values)


def _best_effort_decode(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def _rows_from_csv(raw_bytes: bytes) -> list[list[str]]:
    text = _best_effort_decode(raw_bytes)
    reader = csv.reader(io.StringIO(text))
    return [[_normalize_cell(cell) for cell in row] for row in reader]


def _rows_from_xlsx(raw_bytes: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except Exception as error:
        raise ValueError("xlsx parsing requires openpyxl") from error

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([_normalize_cell(cell) for cell in list(row)])
        return rows
    finally:
        wb.close()


def parse_spreadsheet_upload(filename: str, raw_bytes: bytes) -> tuple[list[str], list[dict[str, str]], dict[str, Any]]:
    lower_name = str(filename or "").lower()
    if lower_name.endswith(".csv"):
        matrix = _rows_from_csv(raw_bytes)
    elif lower_name.endswith(".xlsx"):
        matrix = _rows_from_xlsx(raw_bytes)
    else:
        raise ValueError("Only .csv and .xlsx files are supported")

    non_blank_rows = [row for row in matrix if not _is_blank_row(row)]
    if not non_blank_rows:
        raise ValueError("Spreadsheet does not contain any non-empty rows")

    header_row = non_blank_rows[0]
    headers = [_sanitize_header(name, index) for index, name in enumerate(header_row)]

    data_rows: list[dict[str, str]] = []
    for values in non_blank_rows[1:]:
        if len(values) < len(headers):
            values = values + [""] * (len(headers) - len(values))
        if len(values) > len(headers):
            values = values[: len(headers)]
        data_rows.append(dict(zip(headers, values)))

    meta = {
        "total_rows_in_file": len(matrix),
        "non_blank_rows": len(non_blank_rows),
        "header_row_index": 0,
        "data_row_count": len(data_rows),
    }
    return headers, data_rows, meta


def _normalize_column_mapping(mapping: dict[str, Any]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in (mapping or {}).items():
        left = str(key or "").strip().lower()
        right = str(value or "").strip().lower()
        if left and right:
            normalized[left] = right
    return normalized


def validate_mapping(headers: list[str], mapping: dict[str, Any], required_fields: list[str]) -> dict[str, Any]:
    normalized_mapping = _normalize_column_mapping(mapping)
    available = {str(h or "").strip().lower() for h in headers}

    missing: list[str] = []
    invalid: list[str] = []

    for field in required_fields:
        target = normalized_mapping.get(field)
        if not target:
            missing.append(field)
            continue
        if target not in available:
            invalid.append(field)

    valid = not missing and not invalid
    return {
        "valid": valid,
        "missing_required_fields": missing,
        "invalid_mapped_fields": invalid,
        "normalized_mapping": normalized_mapping,
        "available_headers": sorted(available),
    }


def _load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"runs": []}
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return {"runs": []}
    if not isinstance(payload, dict):
        return {"runs": []}
    runs = payload.get("runs")
    if not isinstance(runs, list):
        return {"runs": []}
    return {"runs": runs}


def _save_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


class BatchRunnerStore:
    def __init__(self, storage_dir: Path):
        self._path = storage_dir / BATCH_RUNS_FILENAME

    def list_runs(self) -> list[dict[str, Any]]:
        with _BATCH_LOCK:
            return list(_load_json(self._path).get("runs", []))

    def get_run(self, batch_id: str) -> dict[str, Any] | None:
        target = str(batch_id or "").strip()
        if not target:
            return None
        with _BATCH_LOCK:
            for row in _load_json(self._path).get("runs", []):
                if str(row.get("batch_id")) == target:
                    return row
        return None

    def upsert_run(self, row: dict[str, Any]) -> dict[str, Any]:
        target = str(row.get("batch_id") or "").strip()
        if not target:
            raise ValueError("batch_id is required")

        with _BATCH_LOCK:
            payload = _load_json(self._path)
            runs = payload["runs"]
            for index, existing in enumerate(runs):
                if str(existing.get("batch_id")) == target:
                    runs[index] = row
                    _save_json(self._path, payload)
                    return row
            runs.append(row)
            _save_json(self._path, payload)
        return row


def build_batch_rows(
    parsed_rows: list[dict[str, str]],
    mapping: dict[str, str],
    *,
    workflow_name: str,
    target_machine_uuid: str,
    tenant_id: str,
) -> list[dict[str, Any]]:
    normalized_mapping = _normalize_column_mapping(mapping)
    rows: list[dict[str, Any]] = []

    required_target_keys = ["member_name", "member_id", "paid_through_date"]

    for index, source_row in enumerate(parsed_rows):
        mapped: dict[str, str] = {}
        for target_key, source_header in normalized_mapping.items():
            mapped[target_key] = str(source_row.get(source_header, "")).strip()

        required_missing = [
            key
            for key in required_target_keys
            if not str(mapped.get(key) or "").strip()
        ]

        decision = evaluate_paid_through_date(mapped.get("paid_through_date"))
        initial_status = "ready" if not required_missing else "invalid"

        rows.append(
            {
                "row_id": str(uuid4()),
                "row_number": index + 2,
                "source": source_row,
                "mapped": mapped,
                "required_missing": required_missing,
                "status": initial_status,
                "payment_status": decision.get("payment_status"),
                "decision_reason": decision.get("decision_reason"),
                "paid_through_date": decision.get("paid_through_date"),
                "current_month_end_date": decision.get("current_month_end_date"),
                "child_task_id": None,
                "child_task_status": None,
                "error": None,
                "created_at": _iso_now(),
                "updated_at": _iso_now(),
                "workflow_name": workflow_name,
                "target_machine_uuid": target_machine_uuid,
                "tenant_id": tenant_id,
            }
        )

    return rows


def summarize_rows(rows: list[dict[str, Any]]) -> dict[str, int]:
    summary = {
        "total": len(rows),
        "ready": 0,
        "invalid": 0,
        "queued": 0,
        "assigned": 0,
        "in_progress": 0,
        "completed": 0,
        "failed": 0,
        "canceled": 0,
        "needs_review": 0,
    }
    for row in rows:
        status = str(row.get("status") or "").strip().lower()
        if status in summary:
            summary[status] += 1
        if str(row.get("payment_status") or "") == "needs_review":
            summary["needs_review"] += 1
    return summary


def row_filter_status(row: dict[str, Any]) -> str:
    status = str(row.get("status") or "").strip().lower()
    payment = str(row.get("payment_status") or "").strip().lower()
    keap_task_created = bool(row.get("keap_task_created"))

    if status in {"queued", "assigned", "running", "in_progress"}:
        return "running"
    if status in {"ready", "uploaded"}:
        return "pending"
    if status in {"invalid", "skipped"}:
        return "skipped"
    if status == "failed":
        return "failed"
    if payment == "needs_review":
        return "needs_review"
    if payment == "good":
        return "good_no_action_needed"
    if payment == "bad" and keap_task_created:
        return "bad_payment_task_created"
    if status in {"completed", "canceled", "cancelled"}:
        return "completed"
    return "pending"


def compute_dashboard_summary(
    rows: list[dict[str, Any]],
    *,
    batch_started_at: str | None = None,
    batch_completed_at: str | None = None,
) -> dict[str, Any]:
    base = summarize_rows(rows)
    total_rows = int(base.get("total", 0))

    running_rows = int(base.get("assigned", 0) + base.get("in_progress", 0))
    pending_rows = int(base.get("ready", 0) + base.get("queued", 0))
    completed_rows = int(base.get("completed", 0))
    failed_rows = int(base.get("failed", 0))
    canceled_rows = int(base.get("canceled", 0))
    skipped_rows = int(base.get("invalid", 0))

    needs_review_rows = 0
    good_no_action_needed_rows = 0
    bad_payment_task_created_rows = 0

    terminal_rows = completed_rows + failed_rows + canceled_rows + skipped_rows
    row_durations_seconds: list[float] = []

    for row in rows:
        payment_status = str(row.get("payment_status") or "").strip().lower()
        if payment_status == "needs_review":
            needs_review_rows += 1
        if payment_status == "good":
            good_no_action_needed_rows += 1
        if payment_status == "bad" and bool(row.get("keap_task_created")):
            bad_payment_task_created_rows += 1

        row_completed = _parse_iso_datetime(row.get("completed_at"))
        if row_completed is None:
            continue

        row_start = (
            _parse_iso_datetime(row.get("row_started_at"))
            or _parse_iso_datetime(row.get("created_at"))
            or _parse_iso_datetime(batch_started_at)
        )
        if row_start is None:
            continue

        duration = (row_completed - row_start).total_seconds()
        if duration > 0:
            row_durations_seconds.append(duration)

    progress_percent = 0
    if total_rows > 0:
        progress_percent = int(round((terminal_rows / total_rows) * 100))
        progress_percent = max(0, min(100, progress_percent))

    estimated_remaining_seconds: int | None = None
    if batch_completed_at:
        estimated_remaining_seconds = 0
    elif row_durations_seconds and total_rows > 0:
        avg_duration = sum(row_durations_seconds) / len(row_durations_seconds)
        remaining_rows = max(0, total_rows - terminal_rows)
        estimated_remaining_seconds = int(round(avg_duration * remaining_rows))

    return {
        **base,
        "total_rows": total_rows,
        "pending_rows": pending_rows,
        "running_rows": running_rows,
        "completed_rows": completed_rows,
        "failed_rows": failed_rows,
        "needs_review_rows": needs_review_rows,
        "skipped_rows": skipped_rows,
        "canceled_rows": canceled_rows,
        "good_no_action_needed_rows": good_no_action_needed_rows,
        "bad_payment_task_created_rows": bad_payment_task_created_rows,
        "progress_percent": progress_percent,
        "estimated_remaining_seconds": estimated_remaining_seconds,
    }


def derive_batch_status(batch: dict[str, Any]) -> str:
    if bool(batch.get("cancel_requested")):
        return "canceled"

    summary = dict(batch.get("summary") or {})
    total_rows = int(summary.get("total_rows") or summary.get("total") or 0)
    terminal_rows = int(summary.get("completed_rows") or 0) + int(summary.get("failed_rows") or 0) + int(summary.get("canceled_rows") or 0) + int(summary.get("skipped_rows") or 0)

    if total_rows > 0 and terminal_rows >= total_rows:
        failed_rows = int(summary.get("failed_rows") or 0)
        needs_review_rows = int(summary.get("needs_review_rows") or 0)
        if failed_rows >= total_rows:
            return "failed"
        if failed_rows > 0 or needs_review_rows > 0:
            return "completed_with_errors"
        return "completed"

    if batch.get("started_at"):
        return "running"
    return "pending"


def build_batch_run_record(
    *,
    batch_id: str,
    created_by_user_id: str | None,
    created_by_name: str | None,
    tenant_id: str,
    workflow_name: str,
    target_machine_uuid: str,
    filename: str,
    headers: list[str],
    mapping: dict[str, str],
    parser_meta: dict[str, Any],
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    now_iso = _iso_now()
    summary = compute_dashboard_summary(rows)
    return {
        "batch_id": batch_id,
        "tenant_id": tenant_id,
        "workflow_name": workflow_name,
        "target_machine_uuid": target_machine_uuid,
        "status": "pending",
        "filename": filename,
        "headers": headers,
        "mapping": mapping,
        "parser_meta": parser_meta,
        "rows": rows,
        "summary": summary,
        "created_at": now_iso,
        "updated_at": now_iso,
        "started_at": None,
        "completed_at": None,
        "created_by_user_id": created_by_user_id,
        "created_by_name": created_by_name,
        "cancel_requested": False,
    }


def to_csv_bytes(batch: dict[str, Any]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "row_number",
            "member_name",
            "member_id",
            "paid_through_date",
            "payment_status",
            "decision_reason",
            "status",
            "child_task_id",
            "child_task_status",
            "matched_client_name",
            "keap_task_created",
            "keap_task_id",
            "worker_name",
            "assigned_machine_uuid",
            "completed_at",
            "error",
        ]
    )
    for row in batch.get("rows", []):
        mapped = row.get("mapped") or {}
        writer.writerow(
            [
                row.get("row_number"),
                mapped.get("member_name", ""),
                mapped.get("member_id", ""),
                row.get("paid_through_date", ""),
                row.get("payment_status", ""),
                row.get("decision_reason", ""),
                row.get("status", ""),
                row.get("child_task_id") or "",
                row.get("child_task_status") or "",
                row.get("matched_client_name") or "",
                str(bool(row.get("keap_task_created"))),
                row.get("keap_task_id") or "",
                row.get("worker_name") or "",
                row.get("assigned_machine_uuid") or "",
                row.get("completed_at") or "",
                row.get("error") or "",
            ]
        )
    return output.getvalue().encode("utf-8")
