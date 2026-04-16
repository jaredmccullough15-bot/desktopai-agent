# actions.py
# Reliable “employee-like” execution:
# - Use pywinauto first (stable)
# - Fall back to pyautogui only if needed
# - Prefer clipboard paste for long text

import os
import webbrowser
import time
import random
import re
import json
import subprocess
import threading
from datetime import datetime
from typing import Optional
from modules.app_logger import append_agent_log

import pyautogui

pyautogui.FAILSAFE = True

try:
    from pywinauto import Application, Desktop
    from pywinauto.keyboard import send_keys
except Exception:
    Application = None
    Desktop = None
    send_keys = None

try:
    import pygetwindow as gw
except Exception:
    gw = None

try:
    from playwright.sync_api import sync_playwright
except Exception:
    sync_playwright = None

_selenium_driver = None
_selenium_driver_lock = threading.RLock()
_last_debug_chrome_restart_ts = 0.0


def _trace_action_exception(context: str, error: Exception) -> None:
    try:
        append_agent_log(f"{context}: {error}", category="Action")
    except Exception:
        pass

def _get_selenium_driver():
    """Attach to existing Chrome via remote debugging port (default 9222)."""
    global _selenium_driver
    global _last_debug_chrome_restart_ts
    try:
        def _restart_debug_chrome_once() -> bool:
            try:
                cooldown = 20.0
                now = time.time()
                if (now - float(_last_debug_chrome_restart_ts or 0.0)) < cooldown:
                    return False

                repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
                script_path = os.path.join(repo_root, "start-chrome-debug.ps1")
                if not os.path.isfile(script_path):
                    return False

                _last_debug_chrome_restart_ts = now
                append_agent_log("Auto-recovery: restarting Chrome debug session for Selenium attach", category="Action")
                completed = subprocess.run(
                    [
                        "powershell",
                        "-NoProfile",
                        "-ExecutionPolicy",
                        "Bypass",
                        "-File",
                        script_path,
                    ],
                    cwd=repo_root,
                    capture_output=True,
                    text=True,
                    timeout=90,
                )
                if completed.returncode != 0:
                    stderr = (completed.stderr or "").strip()
                    _trace_action_exception("Auto-recovery Chrome restart failed", RuntimeError(stderr or f"exit={completed.returncode}"))
                    return False

                time.sleep(2.0)
                return True
            except Exception as restart_err:
                _trace_action_exception("Auto-recovery Chrome restart exception", restart_err)
                return False

        def _sanity_touch(driver, attempts: int = 3) -> bool:

            def _probe_driver() -> bool:
                original_handle = None
                try:
                    try:
                        original_handle = driver.current_window_handle
                    except Exception:
                        original_handle = None

                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass

                    handles = []
                    try:
                        handles = list(driver.window_handles or [])
                    except Exception:
                        handles = []

                    if handles:
                        ordered_handles = list(handles)
                        if original_handle in ordered_handles:
                            ordered_handles = [original_handle] + [h for h in ordered_handles if h != original_handle]

                        healthy_handle = None
                        for handle in ordered_handles:
                            try:
                                driver.switch_to.window(handle)
                                _ = driver.current_url
                                _ = driver.title
                                try:
                                    driver.execute_script("return document.readyState")
                                except Exception:
                                    pass
                                healthy_handle = handle
                                break
                            except Exception:
                                continue

                        if not healthy_handle:
                            return False

                        if original_handle and (healthy_handle != original_handle):
                            try:
                                driver.switch_to.window(original_handle)
                            except Exception:
                                # Keep a working handle if original is gone.
                                pass
                        return True

                    _ = driver.current_url
                    _ = driver.title
                    try:
                        driver.execute_script("return document.readyState")
                    except Exception:
                        pass
                    return True
                except Exception:
                    return False

            for attempt in range(max(1, attempts)):
                try:
                    if _probe_driver():
                        return True
                    raise RuntimeError("driver probe failed")
                except Exception as touch_err:
                    _trace_action_exception(f"Selenium sanity touch failed (attempt {attempt + 1}/{attempts})", touch_err)
                    time.sleep(0.35)
            return False

        def _create_attached_driver():
            import os
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options

            port = os.getenv("CHROME_DEBUG_PORT", "9222") or "9222"
            debugger_address = f"127.0.0.1:{port}"

            options = Options()
            options.add_experimental_option("debuggerAddress", debugger_address)
            try:
                return webdriver.Chrome(options=options)
            except Exception as driver_err:
                try:
                    from selenium.webdriver.chrome.service import Service
                    from webdriver_manager.chrome import ChromeDriverManager
                    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
                except Exception as wm_err:
                    print(f"⚠️ Selenium attach failed (driver): {driver_err} | webdriver-manager fallback failed: {wm_err}")
                    return None

        with _selenium_driver_lock:
            # If we already have a live driver, return it
            if _selenium_driver is not None:
                if _sanity_touch(_selenium_driver, attempts=2):
                    return _selenium_driver

                _trace_action_exception("Existing Selenium session invalid; recreating", RuntimeError("sanity touch failed"))
                stale_driver = _selenium_driver
                _selenium_driver = None
                try:
                    stale_driver.quit()
                except Exception as quit_err:
                    _trace_action_exception("Failed to quit stale Selenium session", quit_err)

            # Run headful, attach to existing debug instance
            _selenium_driver = _create_attached_driver()
            if _selenium_driver is None:
                return None

            if not _sanity_touch(_selenium_driver, attempts=3):
                _trace_action_exception("Selenium attach sanity check failed after retries", RuntimeError("execution context unavailable"))
                failed_driver = _selenium_driver
                _selenium_driver = None
                try:
                    failed_driver.quit()
                except Exception as quit_err:
                    _trace_action_exception("Failed to quit failed Selenium attach session", quit_err)

                # One reattach attempt in case the tab/frame was transiently reloading.
                _selenium_driver = _create_attached_driver()
                if _selenium_driver is None:
                    return None
                if not _sanity_touch(_selenium_driver, attempts=2):
                    _trace_action_exception("Selenium reattach sanity check failed", RuntimeError("execution context unavailable"))
                    reattach_driver = _selenium_driver
                    _selenium_driver = None
                    try:
                        reattach_driver.quit()
                    except Exception as quit_err:
                        _trace_action_exception("Failed to quit reattach Selenium session", quit_err)

                    # Final recovery: restart debug Chrome once, then attach again.
                    if _restart_debug_chrome_once():
                        _selenium_driver = _create_attached_driver()
                        if _selenium_driver is not None and _sanity_touch(_selenium_driver, attempts=4):
                            return _selenium_driver

                    return None

            return _selenium_driver
    except Exception as e:
        print(f"⚠️ Selenium attach failed: {e}")
        return None

def click_next_view_button() -> dict:
    """
    Click the next "View" action in the HealthSherpa clients list.
    Uses multiple locator strategies and tracks index across clicks.
    Returns dict with success status and client name.
    """
    global _view_button_index
    from selenium.webdriver.common.by import By

    driver = _get_selenium_driver()
    if driver is None:
        print("⚠️ click_next_view_button: Selenium driver not available")
        return {"success": False, "client_name": "", "index": _view_button_index}

    try:
        # Ensure we start search near the top
        try:
            driver.execute_script("window.scrollTo(0,0);")
            time.sleep(0.1)
        except Exception as e:
            _trace_action_exception("Failed to scroll to top before view button search", e)

        xpath_candidates = [
            # Exact text on anchors or buttons
            "//a[normalize-space(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'))='view']",
            "//button[normalize-space(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'))='view']",
            # Contains text variants (e.g., 'View Application')
            "//a[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view')]",
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view')]",
            # aria-label variants
            "//*[@aria-label and contains(translate(@aria-label, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view')]",
            # role-based link/button labels
            "//*[@role='link' and contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view')]",
            "//*[@role='button' and contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view')]",
        ]

        view_buttons = []
        chosen_selector = None
        for xp in xpath_candidates:
            try:
                elems = driver.find_elements(By.XPATH, xp)
                # Filter displayed elements
                elems = [e for e in elems if _safe_is_displayed(e)]
                if elems:
                    view_buttons = elems
                    chosen_selector = xp
                    break
            except Exception:
                continue

        _append_agent_log(f"View button search selector={'none' if chosen_selector is None else 'xpath'} count={len(view_buttons)}")

        if not view_buttons:
            print("⚠️ No View buttons found")
            return {"success": False, "client_name": "", "index": _view_button_index}

        if _view_button_index >= len(view_buttons):
            print(f"⚠️ View button index {_view_button_index} out of range (only {len(view_buttons)} buttons)")
            return {"success": False, "client_name": "", "index": _view_button_index}

        # Sort by Y position to ensure consistent ordering
        try:
            view_buttons.sort(key=lambda el: el.location.get('y', 0))
        except Exception as e:
            _trace_action_exception("Failed to sort view buttons by vertical position", e)

        target_button = view_buttons[_view_button_index]

        # Try to get client name from ancestor row
        client_name = f"Client {_view_button_index + 1}"
        try:
            row = _find_ancestor_row(target_button)
            if row is not None:
                row_text = (row.text or '').strip()
                parts = [p.strip() for p in row_text.split('\n') if p.strip() and p.strip().lower() != 'view']
                if parts:
                    client_name = parts[0]
        except Exception as e:
            _trace_action_exception("Failed to derive client name from row", e)

        print(f"📋 Clicking View button #{_view_button_index + 1} for client: {client_name}")

        windows_before = driver.window_handles
        current_url_before = driver.current_url

        # Scroll into view and click (fallback to JS click if needed)
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_button)
            time.sleep(0.1)
            target_button.click()
        except Exception:
            driver.execute_script("arguments[0].click();", target_button)
        time.sleep(0.2)

        windows_after = driver.window_handles
        opened_new_window = False
        if len(windows_after) > len(windows_before):
            new_window = [w for w in windows_after if w not in windows_before][0]
            driver.switch_to.window(new_window)
            opened_new_window = True
            print(f"🔄 Switched to new window (total windows: {len(windows_after)})")
        else:
            print(f"📄 Navigated in same window from {current_url_before[:50]} to {driver.current_url[:50]}")

        _view_button_index += 1
        print(f"✅ Successfully clicked View button for: {client_name}")
        _append_agent_log(f"Clicked View button index={_view_button_index - 1} client='{client_name}' new_window={opened_new_window}")
        return {
            "success": True,
            "client_name": client_name,
            "index": _view_button_index - 1,
            "opened_new_window": opened_new_window,
        }

    except Exception as e:
        print(f"⚠️ click_next_view_button failed: {e}")
        _append_agent_log(f"click_next_view_button failed: {e}")
        return {"success": False, "client_name": "", "index": _view_button_index}

def _safe_is_displayed(el) -> bool:
    try:
        return bool(el.is_displayed())
    except Exception:
        return True

def _find_ancestor_row(el):
    from selenium.webdriver.common.by import By
    try:
        return el.find_element(By.XPATH, "./ancestor::tr")
    except Exception:
        # Try ARIA grid row structure
        try:
            return el.find_element(By.XPATH, "./ancestor::*[@role='row']")
        except Exception:
            return None
def selenium_get_input_value_by_label(label_text: str) -> Optional[str]:
        driver = _get_selenium_driver()
        if driver is None:
                return None
        label = (label_text or "").strip()
        if not label:
                return None

        script = """
        const labelText = (arguments[0] || "").trim().toLowerCase();
        if (!labelText) return null;

        const candidates = Array.from(document.querySelectorAll("label,div,span,td,th"));
        let best = null;
        for (const el of candidates) {
            const text = (el.textContent || "").trim().toLowerCase();
            if (!text) continue;
            if (text === labelText || text.includes(labelText)) {
                best = el;
                break;
            }
        }
        if (!best) return null;

        let input = null;
        const forId = best.getAttribute && best.getAttribute("for");
        if (forId) input = document.getElementById(forId);
        if (!input) input = best.querySelector && best.querySelector("input,textarea,select");

        if (!input) {
            let el = best.parentElement;
            for (let i = 0; i < 4 && el; i += 1) {
                const candidate = el.querySelector && el.querySelector("input,textarea,select");
                if (candidate) {
                    input = candidate;
                    break;
                }
                el = el.parentElement;
            }
        }

        if (!input) {
            let sibling = best.nextElementSibling;
            for (let i = 0; i < 3 && sibling; i += 1) {
                const candidate = sibling.querySelector && sibling.querySelector("input,textarea,select");
                if (candidate) {
                    input = candidate;
                    break;
                }
                sibling = sibling.nextElementSibling;
            }
        }

        if (!input || !("value" in input)) return null;
        return input.value || "";
        """
        try:
                value = driver.execute_script(script, label)
                if value is None:
                        return None
                return str(value)
        except Exception:
                return None


def selenium_get_current_url() -> Optional[str]:
    driver = _get_selenium_driver()
    if driver is None:
        return None
    try:
        url = driver.current_url
        if not url:
            return None
        return str(url)
    except Exception:
        return None


def selenium_get_active_input_info() -> Optional[dict]:
        driver = _get_selenium_driver()
        if driver is None:
                return None
        script = """
        try {
            const el = document.activeElement;
            if (!el) return null;
            const tag = (el.tagName || "").toLowerCase();
            if (!(["input","textarea","select"].includes(tag))) return null;

            const labelFromLabels = () => {
                if (el.labels && el.labels.length) {
                    const text = (el.labels[0].textContent || "").trim();
                    return text || null;
                }
                return null;
            };

            const labelFromAria = () => {
                const aria = (el.getAttribute("aria-label") || "").trim();
                if (aria) return aria;
                const labelledBy = el.getAttribute("aria-labelledby");
                if (labelledBy) {
                    const ref = document.getElementById(labelledBy);
                    if (ref) {
                        const text = (ref.textContent || "").trim();
                        return text || null;
                    }
                }
                return null;
            };

            const labelFromPlaceholder = () => {
                const ph = (el.getAttribute("placeholder") || "").trim();
                return ph || null;
            };

            const label = labelFromLabels() || labelFromAria() || labelFromPlaceholder();
            return {
                id: el.id || "",
                name: el.name || "",
                label: label || "",
                value: ("value" in el ? (el.value || "") : ""),
                tag: tag
            };
        } catch (e) {
            return null;
        }
        """
        try:
                return driver.execute_script(script)
        except Exception:
                return None


def selenium_set_input_value(label_text: str, value: str, name: str = "", element_id: str = "") -> bool:
        driver = _get_selenium_driver()
        if driver is None:
                return False
        label = (label_text or "").strip()
        name = (name or "").strip()
        element_id = (element_id or "").strip()

        script = """
        const labelText = (arguments[0] || "").trim().toLowerCase();
        const name = (arguments[1] || "").trim();
        const elementId = (arguments[2] || "").trim();
        const value = (arguments[3] || "");

        let input = null;
        if (elementId) {
            input = document.getElementById(elementId);
        }
        if (!input && name) {
            input = document.querySelector(`[name="${CSS.escape(name)}"]`);
        }

        if (!input && labelText) {
            const candidates = Array.from(document.querySelectorAll("label,div,span,td,th"));
            let best = null;
            for (const el of candidates) {
                const text = (el.textContent || "").trim().toLowerCase();
                if (!text) continue;
                if (text === labelText || text.includes(labelText)) {
                    best = el;
                    break;
                }
            }
            if (best) {
                const forId = best.getAttribute && best.getAttribute("for");
                if (forId) input = document.getElementById(forId);
                if (!input) input = best.querySelector && best.querySelector("input,textarea,select");
                if (!input) {
                    let el = best.parentElement;
                    for (let i = 0; i < 4 && el; i += 1) {
                        const candidate = el.querySelector && el.querySelector("input,textarea,select");
                        if (candidate) { input = candidate; break; }
                        el = el.parentElement;
                    }
                }
                if (!input) {
                    let sibling = best.nextElementSibling;
                    for (let i = 0; i < 3 && sibling; i += 1) {
                        const candidate = sibling.querySelector && sibling.querySelector("input,textarea,select");
                        if (candidate) { input = candidate; break; }
                        sibling = sibling.nextElementSibling;
                    }
                }
            }
        }

        if (!input || !("value" in input)) return false;
        input.focus();
        input.value = value;
        input.dispatchEvent(new Event("input", { bubbles: true }));
        input.dispatchEvent(new Event("change", { bubbles: true }));
        return true;
        """

        try:
                return bool(driver.execute_script(script, label, name, element_id, value))
        except Exception:
                return False


def _normalize_app(app_name: str) -> str:
    a = (app_name or "").strip().lower()
    aliases = {
        "notepad": "notepad.exe",
        "calculator": "calc.exe",
        "calc": "calc.exe",
        "paint": "mspaint.exe",
        "wordpad": "write.exe",
    }
    return aliases.get(a, (app_name or "").strip())


def _clipboard_paste(text: str) -> None:
    import pyperclip

    pyperclip.copy(text)
    time.sleep(0.05)
    if send_keys is not None:
        send_keys("^v")  # Ctrl+V
    else:
        pyautogui.hotkey("ctrl", "v")


def _focus_chrome_window() -> bool:
    # Try pygetwindow first
    if gw is not None:
        try:
            wins = [w for w in gw.getAllWindows() if w.title and "chrome" in w.title.lower()]
            if wins:
                wins[0].activate()
                time.sleep(0.2)
                return True
        except Exception:
            pass
    # Fallback to pywinauto
    if Desktop is not None:
        try:
            for w in Desktop(backend="uia").windows():
                title = (w.window_text() or "").lower()
                if "chrome" in title:
                    w.set_focus()
                    time.sleep(0.2)
                    return True
        except Exception:
            pass
    return False


def focus_chrome_window() -> bool:
    return _focus_chrome_window()


def focus_window_by_title(title_hint: str) -> bool:
    hint = (title_hint or "").strip().lower()
    if not hint:
        return False
    if "chrome" in hint:
        return _focus_chrome_window()

    if gw is not None:
        try:
            wins = [w for w in gw.getAllWindows() if w.title and hint in w.title.lower()]
            if wins:
                wins[0].activate()
                time.sleep(0.2)
                return True
        except Exception:
            pass

    if Desktop is not None:
        try:
            for w in Desktop(backend="uia").windows():
                title = (w.window_text() or "").lower()
                if hint in title:
                    w.set_focus()
                    time.sleep(0.2)
                    return True
        except Exception:
            pass

    return False


def cycle_browser_tab(direction: str = "next", count: int = 1) -> bool:
    """Cycle browser tabs in the active window."""
    count = max(1, int(count))
    for _ in range(count):
        if direction == "prev":
            pyautogui.hotkey("ctrl", "shift", "tab")
        else:
            pyautogui.hotkey("ctrl", "tab")
        time.sleep(0.1)
    return True


def _pyautogui_open_and_type(app_name: str, text_to_type: Optional[str] = None) -> bool:
    try:
        pyautogui.press("win")
        time.sleep(0.4)
        pyautogui.write(app_name, interval=0.04)
        time.sleep(0.3)
        pyautogui.press("enter")

        if text_to_type:
            time.sleep(1.2)
            # focus center-ish
            w, h = pyautogui.size()
            pyautogui.click(w // 2, h // 3)
            time.sleep(0.2)
            _clipboard_paste(text_to_type)
        return True
    except Exception as e:
        print(f"💥 pyautogui fallback failed: {e}")
        return False


# ========== ADAPTIVE ELEMENT INTERACTION ==========

def _smart_scroll_and_find(find_function, scroll_strategies: list = None, wait_time: float = 1.0):
    """
    Adaptive helper that tries multiple scroll strategies to find elements.
    Learns which strategies work and applies them automatically.
    
    Args:
        find_function: Function to call after each scroll attempt (returns element or None)
        scroll_strategies: List of scroll strategy names to try ["top", "bottom", "middle", "incremental"]
        wait_time: How long to wait after scrolling before searching
    
    Returns:
        Result from find_function or None
    """
    from .memory import should_apply_pattern, add_learning_pattern
    
    driver = _get_selenium_driver()
    if driver is None:
        return None


def _normalize_compare_value(value: str, mode: str = "text") -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if mode == "email":
        return raw.lower().strip()
    if mode == "phone":
        digits = "".join(ch for ch in raw if ch.isdigit())
        return digits[-10:] if len(digits) >= 10 else digits
    if mode == "date":
        return "".join(ch for ch in raw if ch.isdigit())
    lowered = raw.lower()
    lowered = re.sub(r"\s+", " ", lowered).strip()
    lowered = re.sub(r"[^a-z0-9 ]", "", lowered)
    return lowered


def _infer_compare_mode(field_name: str) -> str:
    name = (field_name or "").strip().lower()
    if any(k in name for k in ("email", "e-mail")):
        return "email"
    if any(k in name for k in ("phone", "mobile", "cell")):
        return "phone"
    if any(k in name for k in ("dob", "date of birth", "birth", "date")):
        return "date"
    return "text"


def _find_window_handle_by_tokens(driver, tokens: list[str]) -> Optional[str]:
    norm_tokens = [str(t).strip().lower() for t in (tokens or []) if str(t).strip()]
    if not norm_tokens:
        return None

    handles = []
    try:
        handles = list(driver.window_handles)
    except Exception:
        handles = []

    for handle in handles:
        try:
            driver.switch_to.window(handle)
            title = (driver.title or "").lower()
            url = (driver.current_url or "").lower()
            hay = f"{title} {url}"
            if any(token in hay for token in norm_tokens):
                return handle
        except Exception:
            continue
    return None


def _get_labeled_values_from_current_page(labels: list[str]) -> dict:
    values = {}
    for label in labels:
        key = str(label or "").strip()
        if not key:
            continue
        val = selenium_get_input_value_by_label(key)
        values[key] = "" if val is None else str(val)
    return values


def _get_input_value_from_current_page(field_id: str = "", field_name: str = "", field_label: str = "") -> str:
        driver = _get_selenium_driver()
        if driver is None:
                return ""

        field_id = str(field_id or "").strip()
        field_name = str(field_name or "").strip()
        field_label = str(field_label or "").strip()

        script = """
        const idValue = (arguments[0] || '').trim();
        const nameValue = (arguments[1] || '').trim();

        function readValue(el) {
            if (!el) return null;
            if ('value' in el) return el.value || '';
            return (el.textContent || '').trim();
        }

        if (idValue) {
            const byId = document.getElementById(idValue);
            const v = readValue(byId);
            if (v !== null) return v;
        }

        if (nameValue) {
            const byName = document.querySelector(`[name="${CSS.escape(nameValue)}"]`);
            const v = readValue(byName);
            if (v !== null) return v;
        }

        return null;
        """

        try:
                value = driver.execute_script(script, field_id, field_name)
                if value is not None:
                        return str(value)
        except Exception:
                pass

        if field_label:
                fallback = selenium_get_input_value_by_label(field_label)
                return "" if fallback is None else str(fallback)

        return ""


def verify_marketplace_profile_match(
    checks: list,
    marketplace_match_tokens: Optional[list] = None,
    profile_match_tokens: Optional[list] = None,
) -> dict:
    """
    Compare key fields between Healthcare.gov Marketplace tab and Infusionsoft/Keap profile tab.

        checks item format:
      {
        "marketplace_label": "First name",
        "profile_label": "First Name",
                "marketplace_id": "Contact0PostalCode",   # optional
                "profile_id": "PostalCode",               # optional
                "marketplace_name": "Contact0PostalCode", # optional
                "profile_name": "PostalCode",             # optional
        "mode": "text|email|phone|date",  # optional
        "required": true                    # optional, default true
      }
    """
    driver = _get_selenium_driver()
    if driver is None:
        return {
            "success": False,
            "error": "Selenium driver unavailable",
            "matches": [],
            "mismatches": [],
        }

    checks = checks or []
    if not checks:
        return {
            "success": False,
            "error": "No checks provided",
            "matches": [],
            "mismatches": [],
        }

    marketplace_tokens = marketplace_match_tokens or ["healthcare.gov", "marketplace"]
    profile_tokens = profile_match_tokens or ["infusionsoft", "keap"]

    original_handle = None
    try:
        original_handle = driver.current_window_handle
    except Exception:
        original_handle = None

    marketplace_handle = _find_window_handle_by_tokens(driver, marketplace_tokens)
    profile_handle = _find_window_handle_by_tokens(driver, profile_tokens)

    if marketplace_handle is None:
        if original_handle:
            try:
                driver.switch_to.window(original_handle)
            except Exception:
                pass
        return {
            "success": False,
            "error": f"Marketplace tab not found (tokens={marketplace_tokens})",
            "matches": [],
            "mismatches": [],
        }

    if profile_handle is None:
        if original_handle:
            try:
                driver.switch_to.window(original_handle)
            except Exception:
                pass
        return {
            "success": False,
            "error": f"Profile tab not found (tokens={profile_tokens})",
            "matches": [],
            "mismatches": [],
        }

    matches = []
    mismatches = []

    try:
        for check in checks:
            if not isinstance(check, dict):
                continue

            m_label = str(check.get("marketplace_label", "")).strip()
            p_label = str(check.get("profile_label", "")).strip()
            m_id = str(check.get("marketplace_id", "")).strip()
            p_id = str(check.get("profile_id", "")).strip()
            m_name = str(check.get("marketplace_name", "")).strip()
            p_name = str(check.get("profile_name", "")).strip()

            if not (m_label or m_id or m_name) or not (p_label or p_id or p_name):
                continue

            required = bool(check.get("required", True))
            mode = str(check.get("mode", "")).strip().lower() or _infer_compare_mode(
                f"{m_label or m_id or m_name} {p_label or p_id or p_name}"
            )

            driver.switch_to.window(marketplace_handle)
            m_raw = _get_input_value_from_current_page(field_id=m_id, field_name=m_name, field_label=m_label)

            driver.switch_to.window(profile_handle)
            p_raw = _get_input_value_from_current_page(field_id=p_id, field_name=p_name, field_label=p_label)

            out_m_label = m_label or m_id or m_name
            out_p_label = p_label or p_id or p_name

            if required and (not m_raw or not p_raw):
                mismatches.append({
                    "marketplace_label": out_m_label,
                    "profile_label": out_p_label,
                    "marketplace_value": m_raw,
                    "profile_value": p_raw,
                    "reason": "missing_required_value",
                })
                continue

            m_val = _normalize_compare_value(m_raw, mode)
            p_val = _normalize_compare_value(p_raw, mode)

            if m_val == p_val:
                matches.append({
                    "marketplace_label": out_m_label,
                    "profile_label": out_p_label,
                    "value": m_raw,
                    "mode": mode,
                })
            else:
                mismatches.append({
                    "marketplace_label": out_m_label,
                    "profile_label": out_p_label,
                    "marketplace_value": m_raw,
                    "profile_value": p_raw,
                    "marketplace_normalized": m_val,
                    "profile_normalized": p_val,
                    "reason": "value_mismatch",
                    "mode": mode,
                })
    finally:
        if original_handle:
            try:
                driver.switch_to.window(original_handle)
            except Exception:
                pass

    return {
        "success": len(mismatches) == 0,
        "matches": matches,
        "mismatches": mismatches,
        "checked": len(matches) + len(mismatches),
        "marketplace_handle": marketplace_handle,
        "profile_handle": profile_handle,
    }
    
    if scroll_strategies is None:
        # Check if we've learned a preferred strategy for this context
        learned_pattern = should_apply_pattern("element_interaction", ["element not found", "viewport", "scroll"])
        if learned_pattern and learned_pattern.get("solution"):
            solution = learned_pattern.get("solution", "")
            if "scroll_to_bottom" in solution:
                scroll_strategies = ["bottom", "top", "incremental"]
            elif "scroll_to_top" in solution:
                scroll_strategies = ["top", "bottom", "incremental"]
            else:
                scroll_strategies = ["bottom", "top", "incremental"]
        else:
            # Default: try bottom first (most common for pagination, buttons)
            scroll_strategies = ["bottom", "top", "incremental"]
    
    print(f"🔍 Trying smart scroll strategies: {scroll_strategies}")
    
    # First try without scrolling
    result = find_function()
    if result is not None:
        return result
    
    # Try each scroll strategy
    for strategy in scroll_strategies:
        try:
            if strategy == "bottom":
                print("⬇️ Adaptive scroll: Moving to bottom of page...")
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(wait_time)
                
            elif strategy == "top":
                print("⬆️ Adaptive scroll: Moving to top of page...")
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(wait_time)
                
            elif strategy == "middle":
                print("↕️ Adaptive scroll: Moving to middle of page...")
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 2);")
                time.sleep(wait_time)
                
            elif strategy == "incremental":
                print("📜 Adaptive scroll: Scanning page incrementally...")
                # Scroll in chunks, checking after each
                viewport_height = driver.execute_script("return window.innerHeight;")
                page_height = driver.execute_script("return document.body.scrollHeight;")
                
                current_scroll = 0
                scroll_step = viewport_height * 0.8  # Scroll 80% of viewport each time
                
                while current_scroll < page_height:
                    driver.execute_script(f"window.scrollTo(0, {current_scroll});")
                    time.sleep(wait_time * 0.5)  # Shorter wait for incremental
                    
                    result = find_function()
                    if result is not None:
                        # Learn this pattern!
                        add_learning_pattern(
                            pattern_type="element_interaction",
                            context="element not found in viewport, needed incremental scroll",
                            solution="scroll_incrementally_and_search"
                        )
                        print(f"✅ Found element at scroll position {current_scroll}px (incremental)")
                        return result
                    
                    current_scroll += scroll_step
                
                # After full page scan, return to top
                driver.execute_script("window.scrollTo(0, 0);")
                continue
            
            # Try finding after this scroll strategy
            result = find_function()
            if result is not None:
                # Learn this successful pattern!
                add_learning_pattern(
                    pattern_type="element_interaction",
                    context=f"element not found in viewport, needed scroll to {strategy}",
                    solution=f"scroll_to_{strategy}_before_search"
                )
                print(f"✅ Found element after scrolling to {strategy}")
                return result
                
        except Exception as e:
            print(f"⚠️ Scroll strategy '{strategy}' failed: {e}")
            continue
    
    print("❌ Element not found after trying all scroll strategies")
    return None


def _smart_wait_for_element(selector_function, max_wait: float = 10.0, check_interval: float = 0.5):
    """
    Adaptive waiting that learns typical page load times and element appearance patterns.
    
    Args:
        selector_function: Function to call repeatedly (returns element or None)
        max_wait: Maximum time to wait in seconds
        check_interval: How often to check
    
    Returns:
        Result from selector_function or None
    """
    from .memory import should_apply_pattern, add_learning_pattern
    
    # Check if we've learned about slow loading for this type
    learned_pattern = should_apply_pattern("element_interaction", ["slow load", "wait", "delayed"])
    if learned_pattern:
        # Increase max wait if we've learned this is slow
        max_wait = max(max_wait, 15.0)
        print(f"🧠 Applied learning: Extended wait time to {max_wait}s")
    
    start_time = time.time()
    attempts = 0
    
    while time.time() - start_time < max_wait:
        result = selector_function()
        if result is not None:
            wait_duration = time.time() - start_time
            
            # Learn if this took a long time
            if wait_duration > 3.0:
                add_learning_pattern(
                    pattern_type="element_interaction",
                    context="element took long time to appear, needs extended wait",
                    solution=f"wait_extended_{int(wait_duration)}s"
                )
                print(f"🧠 Learned: This element type needs ~{wait_duration:.1f}s wait time")
            
            return result
        
        attempts += 1
        time.sleep(check_interval)
    
    print(f"⏱️ Element not found after {max_wait}s (tried {attempts} times)")
    return None


def open_app_and_type(app_name: str, text_to_type: Optional[str] = None, focus_timeout: float = 6.0) -> bool:
    """
    Opens an app and optionally types/pastes text.
    Returns True if it likely succeeded.
    """
    app_name = _normalize_app(app_name)
    print(f"🎬 EXEC: OPEN {app_name!r}" + (f" THEN TYPE {text_to_type!r}" if text_to_type else ""))

    # Prefer pywinauto if available
    if Application is not None and Desktop is not None:
        try:
            Application(backend="uia").start(app_name)
            time.sleep(1.0)

            # focus best-effort active window
            end = time.time() + focus_timeout
            win = None
            while time.time() < end:
                try:
                    win = Desktop(backend="uia").get_active()
                    if win:
                        win.set_focus()
                        break
                except Exception:
                    pass
                time.sleep(0.1)

            if text_to_type and win is not None:
                # Try to focus an Edit control if it exists (Notepad etc.)
                try:
                    edit = win.child_window(control_type="Edit")
                    if edit.exists(timeout=1.0):
                        edit.set_focus()
                        time.sleep(0.05)
                except Exception:
                    pass

                # Paste is most reliable for arbitrary text
                time.sleep(0.2)
                _clipboard_paste(text_to_type)

            return True

        except Exception as e:
            print(f"⚠️ pywinauto path failed: {e} — falling back to pyautogui")

    # Fallback
    return _pyautogui_open_and_type(app_name, text_to_type)


def focus_app_and_type(app_name: str, text_to_type: Optional[str] = None, focus_timeout: float = 6.0) -> bool:
    """
    Focuses an existing app window and optionally types/pastes text.
    Falls back to open_app_and_type if not found.
    """
    app_name = _normalize_app(app_name)

    if Application is not None and Desktop is not None:
        try:
            end = time.time() + focus_timeout
            target = None
            while time.time() < end:
                for w in Desktop(backend="uia").windows():
                    title = (w.window_text() or "").lower()
                    if app_name.replace(".exe", "") in title:
                        target = w
                        break
                if target:
                    break
                time.sleep(0.2)

            if target:
                target.set_focus()
                time.sleep(0.2)
                if text_to_type:
                    _clipboard_paste(text_to_type)
                return True
        except Exception:
            pass

    return open_app_and_type(app_name, text_to_type, focus_timeout=focus_timeout)


def save_active_file(filename: str, directory: Optional[str] = None) -> bool:
    """
    Tries to save the currently active document using Ctrl+S and filename.
    """
    try:
        filename = (filename or "").strip()
        if not filename:
            return False
        if not os.path.splitext(filename)[1]:
            filename = filename + ".txt"

        if directory is None:
            directory = os.path.join(os.path.expanduser("~"), "Desktop")

        full_path = os.path.join(directory, filename)

        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)

        # Prefer UIA Save As dialog if available
        if Desktop is not None:
            try:
                dlg = Desktop(backend="uia").get_active()
                if dlg:
                    # Best-effort: focus filename field
                    edit = dlg.child_window(control_type="Edit")
                    if edit.exists(timeout=1.0):
                        edit.set_focus()
                        time.sleep(0.1)
            except Exception:
                pass

        _clipboard_paste(full_path)
        time.sleep(0.2)
        pyautogui.press("enter")
        return True
    except Exception:
        return False


def click_at(x: int, y: int, jitter: int = 2) -> None:
    w, h = pyautogui.size()
    x = max(0, min(w - 1, x + random.randint(-jitter, jitter)))
    y = max(0, min(h - 1, y + random.randint(-jitter, jitter)))
    pyautogui.moveTo(x, y, duration=random.uniform(0.12, 0.28))
    pyautogui.click()


def press_key(key_name: str) -> None:
    pyautogui.press(key_name)


def type_string(text: str) -> None:
    # keep for short strings; prefer paste for long
    pyautogui.write(text, interval=random.uniform(0.03, 0.08))


def fill_login_fields(username: str, password: str, submit: bool = False) -> bool:
    """Fill current page login fields by typing username, tab, password; optionally submit."""
    try:
        username = (username or "").strip()
        password = (password or "").strip()
        if not username or not password:
            return False
        time.sleep(0.5)
        # focus page by clicking near center before typing
        w, h = pyautogui.size()
        pyautogui.click(w // 2, h // 3)
        time.sleep(0.2)
        # try to move focus to first form field (common in browsers)
        pyautogui.hotkey("ctrl", "l")
        time.sleep(0.1)
        pyautogui.press("tab")
        time.sleep(0.2)
        _clipboard_paste(username)
        time.sleep(0.2)
        pyautogui.press("tab")
        time.sleep(0.2)
        _clipboard_paste(password)
        if submit:
            time.sleep(0.2)
            pyautogui.press("enter")
        return True
    except Exception:
        return False


def open_url(url: str) -> bool:
    try:
        url = (url or "").strip()
        if not url:
            return False
        if not (url.startswith("http://") or url.startswith("https://")):
            url = "https://" + url
        prefer_chrome = os.getenv("PREFER_CHROME", "1") == "1"
        if prefer_chrome:
            if _focus_chrome_window():
                pyautogui.hotkey("ctrl", "t")
                time.sleep(0.2)
                _clipboard_paste(url)
                pyautogui.press("enter")
                return True
            chrome_path = os.getenv("CHROME_PATH", "").strip()
            candidates = [
                chrome_path,
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            ]
            for path in candidates:
                if path and os.path.isfile(path):
                    os.startfile(path, arguments=url)
                    return True
        if not webbrowser.open(url, new=2):
            os.startfile(url)
        return True
    except Exception:
        return False


def open_url_and_click_result(url: str, match_text: Optional[str] = None, timeout_ms: int = 15000) -> bool:
    if sync_playwright is None:
        return False
    try:
        url = (url or "").strip()
        if not url:
            return False
        if not (url.startswith("http://") or url.startswith("https://")):
            url = "https://" + url

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)

            locator = None
            if "youtube.com" in url:
                page.wait_for_load_state("networkidle", timeout=timeout_ms)
                page.wait_for_selector("ytd-video-renderer a#video-title", timeout=timeout_ms)
                if match_text:
                    locator = page.locator("ytd-video-renderer a#video-title", has_text=match_text)
                    if locator.count() == 0:
                        locator = page.locator("ytd-video-renderer a#video-title")
                else:
                    locator = page.locator("ytd-video-renderer a#video-title")
            else:
                if match_text:
                    locator = page.locator("a", has_text=match_text)
                    if locator.count() == 0:
                        locator = page.locator("a")
                else:
                    locator = page.locator("a")

            if locator is None:
                browser.close()
                return False

            # Ensure result exists before clicking
            if locator.count() == 0:
                browser.close()
                return False

            target_href = locator.first.get_attribute("href")
            if not target_href:
                browser.close()
                return False

            if target_href.startswith("/") and "youtube.com" in url:
                target_href = "https://www.youtube.com" + target_href

            browser.close()
            return open_url(target_href)
    except Exception:
        return False


def open_url_and_fill_login(url: str, username: str, password: str, submit: bool = True, timeout_ms: int = 20000) -> bool:
    if sync_playwright is None:
        return False
    try:
        url = (url or "").strip()
        username = (username or "").strip()
        password = (password or "").strip()
        if not url or not username or not password:
            return False
        if not (url.startswith("http://") or url.startswith("https://")):
            url = "https://" + url

        page = _get_playwright_page()
        if page is None:
            return False

        page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
        page.wait_for_timeout(500)

        user_selectors = [
            "input[type='email']",
            "input[name*='email' i]",
            "input[id*='email' i]",
            "input[autocomplete*='username' i]",
            "input[name*='user' i]",
            "input[id*='user' i]",
            "input[type='text']",
        ]
        pass_selectors = [
            "input[type='password']",
        ]

        user_field = None
        for sel in user_selectors:
            loc = page.locator(sel)
            if loc.count() > 0:
                user_field = loc.first
                break

        pass_field = None
        for sel in pass_selectors:
            loc = page.locator(sel)
            if loc.count() > 0:
                pass_field = loc.first
                break

        if user_field is None or pass_field is None:
            # Retry once after a short wait in case fields load late
            page.wait_for_timeout(800)
            if user_field is None:
                for sel in user_selectors:
                    loc = page.locator(sel)
                    if loc.count() > 0:
                        user_field = loc.first
                        break
            if pass_field is None:
                for sel in pass_selectors:
                    loc = page.locator(sel)
                    if loc.count() > 0:
                        pass_field = loc.first
                        break
        if user_field is None or pass_field is None:
            return False

        user_field.click()
        user_field.fill("")
        user_field.type(username, delay=30)
        pass_field.click()
        pass_field.fill("")
        pass_field.type(password, delay=30)

        if submit:
            try:
                pass_field.press("Enter")
            except Exception:
                submit_btn = page.locator("button[type='submit'], input[type='submit']")
                if submit_btn.count() > 0:
                    submit_btn.first.click()

        return True
    except Exception:
        return False


def wait_for_page_load(timeout_sec: float = 30.0) -> bool:
    """
    Wait for the current browser page to finish loading using Selenium.
    Returns True if page loaded successfully, False on timeout or error.
    """
    driver = _get_selenium_driver()
    if driver is None:
        return False
    
    try:
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        
        # Wait for document ready state
        WebDriverWait(driver, timeout_sec).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        
        # Additional small wait for dynamic content
        time.sleep(0.5)
        return True
    except Exception as e:
        print(f"⚠️ wait_for_page_load failed: {e}")
        return False


def change_url_parameter(param_name: str, param_value: str) -> bool:
    """
    Change a URL parameter in the current page and navigate to the new URL.
    Uses JavaScript to modify the URL parameter dynamically.
    
    Args:
        param_name: The parameter name to change (e.g., "page")
        param_value: The new value (e.g., "2")
    
    Returns:
        True if successful, False otherwise
    """
    driver = _get_selenium_driver()
    if driver is None:
        return False
    
    try:
        # Get current URL for debugging
        current_url = driver.current_url
        print(f"📍 Current URL: {current_url[:100]}...")
        
        # JavaScript to change URL parameter and navigate
        script = f"""
        const url = new URL(window.location.href);
        console.log('Before change:', url.href);
        url.searchParams.set('{param_name}', '{param_value}');
        console.log('After change:', url.href);
        return url.toString();
        """
        new_url = driver.execute_script(script)
        print(f"🔄 Changing URL parameter: {param_name}={param_value}")
        print(f"📍 New URL: {new_url[:100]}...")
        
        # Navigate to the new URL
        driver.get(new_url)
        time.sleep(1.0)
        
        # Verify we navigated correctly
        final_url = driver.current_url
        if f"{param_name}={param_value}" in final_url:
            print(f"✅ Successfully navigated to {param_name}={param_value}")
            return True
        else:
            print(f"⚠️ URL doesn't contain expected parameter: {final_url[:100]}")
            return False
            
    except Exception as e:
        print(f"⚠️ change_url_parameter failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def click_pagination(
    page_number: int,
    allow_next_control: bool = True,
    allow_url_fallback: bool = True,
    prefer_next_control_only: bool = False,
) -> bool:
    from .memory import add_learning_pattern, should_apply_pattern
    from .pagination import click_pagination as click_pagination_engine

    driver = _get_selenium_driver()
    if driver is None:
        print("⚠️ Selenium driver not available")
        return False

    return click_pagination_engine(
        driver,
        page_number=page_number,
        allow_next_control=allow_next_control,
        allow_url_fallback=allow_url_fallback,
        prefer_next_control_only=prefer_next_control_only,
        should_apply_pattern=should_apply_pattern,
        add_learning_pattern=add_learning_pattern,
        append_log=_append_agent_log,
        change_url_parameter_func=change_url_parameter,
    )


def search_page_for_identifier(identifier: str, search_type: str = "text") -> Optional[dict]:
    """
    Search the current page for an identifier (text, element, etc.).
    
    Args:
        identifier: The text or selector to search for
        search_type: "text" (search in page text), "selector" (CSS selector), or "xpath"
    
    Returns:
        Dict with {found: bool, location: str, text: str} or None on error
    """
    driver = _get_selenium_driver()
    if driver is None:
        return None
    
    identifier = (identifier or "").strip()
    if not identifier:
        return {"found": False, "location": None, "text": None}
    
    try:
        if search_type == "text":
            # Search for text in page body
            script = """
            const searchText = arguments[0];
            const body = document.body;
            if (!body) return null;
            
            const bodyText = body.innerText || body.textContent || "";
            const found = bodyText.toLowerCase().includes(searchText.toLowerCase());
            
            if (found) {
                // Try to find the element containing the text
                const walker = document.createTreeWalker(
                    body,
                    NodeFilter.SHOW_TEXT,
                    null,
                    false
                );
                
                let node;
                while (node = walker.nextNode()) {
                    const text = node.textContent || "";
                    if (text.toLowerCase().includes(searchText.toLowerCase())) {
                        const element = node.parentElement;
                        if (element) {
                            return {
                                found: true,
                                location: element.tagName + (element.id ? '#' + element.id : '') + (element.className ? '.' + element.className.split(' ')[0] : ''),
                                text: text.trim().substring(0, 200)
                            };
                        }
                    }
                }
            }
            
            return {found: found, location: null, text: bodyText.substring(0, 200)};
            """
            result = driver.execute_script(script, identifier)
            return result if result else {"found": False, "location": None, "text": None}
            
        elif search_type == "selector":
            # Search by CSS selector
            try:
                elements = driver.find_elements("css selector", identifier)
                if elements:
                    elem = elements[0]
                    return {
                        "found": True,
                        "location": identifier,
                        "text": elem.text[:200] if elem.text else None
                    }
                return {"found": False, "location": None, "text": None}
            except Exception:
                return {"found": False, "location": None, "text": None}
                
        elif search_type == "xpath":
            # Search by XPath
            try:
                elements = driver.find_elements("xpath", identifier)
                if elements:
                    elem = elements[0]
                    return {
                        "found": True,
                        "location": identifier,
                        "text": elem.text[:200] if elem.text else None
                    }
                return {"found": False, "location": None, "text": None}
            except Exception:
                return {"found": False, "location": None, "text": None}
        
        return {"found": False, "location": None, "text": None}
        
    except Exception as e:
        print(f"⚠️ search_page_for_identifier failed: {e}")
        return None


def close_current_tab() -> bool:
    """
    Close the current browser tab using Selenium.
    Switches back to the window with the client list (main page).
    Returns True on success.
    """
    driver = _get_selenium_driver()
    if driver is None:
        print("⚠️ close_current_tab: Selenium driver not available")
        return False
    
    try:
        # Get all window handles before closing
        all_windows_before = driver.window_handles
        current_handle = driver.current_window_handle
        
        # If only one window, don't close it
        if len(all_windows_before) <= 1:
            print("⚠️ close_current_tab: Only one window open, not closing")
            return False
        
        print(f"🔄 Closing current window (total windows before: {len(all_windows_before)})")
        
        # Close the current window/tab
        driver.close()
        time.sleep(0.5)
        
        # Get fresh window handles after closing
        remaining_windows = driver.window_handles
        print(f"✓ Window closed (remaining windows: {len(remaining_windows)})")
        
        if not remaining_windows:
            print("⚠️ No windows remaining after close!")
            return False
        
        # Try to find the main window (client list page) by checking URLs
        main_window = None
        for handle in remaining_windows:
            driver.switch_to.window(handle)
            current_url = driver.current_url
            if _is_healthsherpa_clients_list_url(current_url):
                main_window = handle
                print(f"✓ Found main window with client list: {current_url[:100]}")
                break
        
        # If we couldn't identify the main window, just use the first one
        if main_window is None:
            main_window = remaining_windows[0]
            driver.switch_to.window(main_window)
            print(f"ℹ️ Switched to first remaining window: {driver.current_url[:100]}")
        
        time.sleep(0.3)
        print("✓ Tab closed successfully")
        return True
    except Exception as e:
        print(f"⚠️ close_current_tab failed: {e}")
        # Try to recover by switching to any available window
        try:
            handles = driver.window_handles
            if handles:
                driver.switch_to.window(handles[0])
        except:
            pass
        return False


def click_element_by_text(text: str, element_type: str = "button") -> bool:
    """
    Click an element on the current page by its text content using Selenium.
    Now with adaptive learning - automatically applies smart scrolling strategies!
    
    Args:
        text: The text to search for (e.g., "view", "Submit", "Next")
        element_type: Type hint - "button", "link", "any" (default: "button")
    
    Returns:
        True if element was found and clicked, False otherwise
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    
    driver = _get_selenium_driver()
    if driver is None:
        return False
    
    text = (text or "").strip()
    if not text:
        return False
    
    # Define the find and click function that will be used with smart scrolling
    def find_and_click_element():
        try:
            # Build XPath based on element type
            if element_type == 'button':
                xpath = f"//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]"
            elif element_type == 'link':
                xpath = f"//a[.//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]]"
            else:  # any
                # Search in multiple element types
                xpaths = [
                    f"//a[.//*[normalize-space(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'))='{text.lower()}']]",
                    f"//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]",
                    f"//a[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]",
                    f"//*[contains(@onclick, 'click')][contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{text.lower()}')]",
                ]
                
                # Try each XPath
                for xpath in xpaths:
                    try:
                        element = driver.find_element(By.XPATH, xpath)
                        if element and element.is_displayed():
                            # Save current window handles before clicking
                            windows_before = driver.window_handles
                            
                            # Scroll into view
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                            time.sleep(0.3)
                            
                            # Click
                            element.click()
                            print(f"✅ Clicked element with text: {text}")
                            time.sleep(0.5)
                            
                            # Check if a new window was opened
                            windows_after = driver.window_handles
                            if len(windows_after) > len(windows_before):
                                new_window = [w for w in windows_after if w not in windows_before][0]
                                driver.switch_to.window(new_window)
                                print(f"🔄 Switched to new window (total windows: {len(windows_after)})")
                            
                            return True
                    except:
                        continue
                
                return None  # Not found
            
            # For specific button or link types
            element = driver.find_element(By.XPATH, xpath)
            
            if element and element.is_displayed():
                # Save current window handles
                windows_before = driver.window_handles
                
                # Scroll into view
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                time.sleep(0.3)
                
                # Click
                element.click()
                print(f"✅ Clicked element with text: {text}")
                time.sleep(0.5)
                
                # Check if a new window was opened
                windows_after = driver.window_handles
                if len(windows_after) > len(windows_before):
                    new_window = [w for w in windows_after if w not in windows_before][0]
                    driver.switch_to.window(new_window)
                    print(f"🔄 Switched to new window (total windows: {len(windows_after)})")
                
                return True
            else:
                return None  # Found but not visible
                
        except Exception:
            return None  # Not found
    
    # First attempt: Try to find element in current viewport
    result = find_and_click_element()
    if result:
        return True
    
    # Second attempt: Use adaptive smart scrolling to find the element
    print(f"🧠 Element '{text}' not visible, applying adaptive scroll strategies...")
    result = _smart_scroll_and_find(find_and_click_element, scroll_strategies=None, wait_time=1.0)
    
    if result:
        return True
    
    print(f"⚠️ Could not find element with text: {text}")
    return False


# Global counter to track which View button to click next
_view_button_index = 0
 


def reset_view_button_counter():
    """Reset the View button counter to start from the beginning."""
    global _view_button_index
    _view_button_index = 0
    print("🔄 Reset View button counter to 0")
    return True


def wait_for_element_with_text(text: str, timeout_sec: float = 30.0) -> bool:
    """
    Wait for an element containing specific text to appear on the page.
    
    Args:
        text: The text to search for (e.g., "Sync Complete")
        timeout_sec: Maximum time to wait in seconds (default: 30)
    
    Returns:
        True if element appears within timeout, False otherwise
    """
    driver = _get_selenium_driver()
    if driver is None:
        print("⚠️ wait_for_element_with_text: Selenium driver not available")
        return False
    
    text = (text or "").strip()
    if not text:
        print("⚠️ wait_for_element_with_text: No text provided")
        return False
    
    end_time = time.time() + max(0.5, float(timeout_sec or 0.5))
    text_lower = text.lower()

    js_visible_text_check = """
    const needle = (arguments[0] || '').toLowerCase();
    if (!needle) return false;
    const walker = document.createTreeWalker(document.body || document.documentElement, NodeFilter.SHOW_ELEMENT);
    while (walker.nextNode()) {
      const el = walker.currentNode;
      if (!el) continue;
      const tag = (el.tagName || '').toLowerCase();
      if (['script','style','noscript'].includes(tag)) continue;
      const style = window.getComputedStyle(el);
      if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') continue;
      const rect = el.getBoundingClientRect();
      if (rect.width <= 0 || rect.height <= 0) continue;
      const txt = ((el.innerText || el.textContent || '') + '').toLowerCase();
      if (txt.includes(needle)) return true;
    }
    return false;
    """

    try:
        while time.time() < end_time:
            try:
                found = bool(driver.execute_script(js_visible_text_check, text_lower))
            except Exception:
                found = False
            if found:
                print(f"   ✓ Element found: '{text}' is now visible on page")
                return True
            time.sleep(0.25)
        print(f"   ✗ Element not found: '{text}' did not appear within timeout")
        return False
    except Exception as e:
        print(f"   ✗ Error while waiting: {e}")
        return False


def _is_text_visible_on_page(text: str) -> bool:
    driver = _get_selenium_driver()
    if driver is None:
        return False
    needle = (text or "").strip().lower()
    if not needle:
        return False
    script = """
    const needle = (arguments[0] || '').toLowerCase();
    if (!needle) return false;
    const walker = document.createTreeWalker(document.body || document.documentElement, NodeFilter.SHOW_ELEMENT);
    while (walker.nextNode()) {
      const el = walker.currentNode;
      if (!el) continue;
      const tag = (el.tagName || '').toLowerCase();
      if (['script','style','noscript'].includes(tag)) continue;
      const style = window.getComputedStyle(el);
      if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') continue;
      const rect = el.getBoundingClientRect();
      if (rect.width <= 0 || rect.height <= 0) continue;
      const txt = ((el.innerText || el.textContent || '') + '').toLowerCase();
      if (txt.includes(needle)) return true;
    }
    return false;
    """
    try:
        return bool(driver.execute_script(script, needle))
    except Exception:
        return False


def _wait_for_post_click_sync_text(wait_text: str, wait_timeout: float, pre_click_visible: bool = False) -> bool:
    """Wait for Sync text after click while avoiding stale pre-existing text."""
    driver = _get_selenium_driver()
    if driver is None:
        return False

    start_url = ""
    start_title = ""
    try:
        start_url = str(driver.current_url or "")
        start_title = str(driver.title or "")
    except Exception as e:
        _trace_action_exception("Failed to read initial page context before sync wait", e)

    context_deadline = time.time() + 8.0
    while time.time() < context_deadline:
        try:
            cur_url = str(driver.current_url or "")
            cur_title = str(driver.title or "")
        except Exception:
            cur_url = ""
            cur_title = ""

        if (cur_url and cur_url != start_url) or (cur_title and cur_title != start_title):
            break
        time.sleep(0.2)

    if pre_click_visible:
        clear_deadline = time.time() + 6.0
        while time.time() < clear_deadline:
            if not _is_text_visible_on_page(wait_text):
                break
            time.sleep(0.25)

    return wait_for_element_with_text(wait_text, timeout_sec=wait_timeout)


def human_like_mouse_movement(target_x: int = None, target_y: int = None, duration: float = None):
    """
    Move mouse in a human-like curved path with variable speed.
    
    Args:
        target_x: Target X coordinate (None for random movement)
        target_y: Target Y coordinate (None for random movement)
        duration: Movement duration in seconds (None for random)
    """
    try:
        import numpy as np
        
        # Get current position
        current_x, current_y = pyautogui.position()
        
        # If no target specified, move to a random location on screen
        if target_x is None or target_y is None:
            screen_width, screen_height = pyautogui.size()
            target_x = random.randint(100, screen_width - 100)
            target_y = random.randint(100, screen_height - 100)
        
        # Random duration between 0.3 and 1.5 seconds if not specified
        if duration is None:
            duration = random.uniform(0.3, 1.5)
        
        # Calculate distance
        distance = ((target_x - current_x)**2 + (target_y - current_y)**2)**0.5
        
        # Number of steps based on distance and duration
        steps = max(10, int(distance / 5))
        
        # Generate curved path using Bezier curve
        control_x = (current_x + target_x) / 2 + random.randint(-50, 50)
        control_y = (current_y + target_y) / 2 + random.randint(-50, 50)
        
        points = []
        for i in range(steps + 1):
            t = i / steps
            # Quadratic Bezier curve
            x = (1-t)**2 * current_x + 2*(1-t)*t * control_x + t**2 * target_x
            y = (1-t)**2 * current_y + 2*(1-t)*t * control_y + t**2 * target_y
            points.append((int(x), int(y)))
        
        # Move along the path with variable speed
        delay_per_step = duration / steps
        for x, y in points:
            pyautogui.moveTo(x, y)
            time.sleep(delay_per_step * random.uniform(0.8, 1.2))
        
    except Exception as e:
        # Fallback to direct movement if bezier fails
        if target_x and target_y:
            pyautogui.moveTo(target_x, target_y, duration=duration or 0.5)


def take_human_pause(min_seconds: float = 0.5, max_seconds: float = 3.0, reason: str = ""):
    """
    Take a random pause to simulate human behavior.
    
    Args:
        min_seconds: Minimum pause duration
        max_seconds: Maximum pause duration
        reason: Reason for pause (for logging)
    """
    # Fast mode: minimize pauses to keep the agent responsive
    pause_duration = max(0.05, min(random.uniform(min_seconds, max_seconds), 0.3))
    if reason:
        print(f"⏸️  {reason} (pausing {pause_duration:.1f}s)")
    else:
        print(f"⏸️  Taking short pause ({pause_duration:.1f}s)")
    time.sleep(pause_duration)


def occasional_mouse_movement():
    """Disabled: No-op to avoid slow mouse movement."""
    return
        
def _append_agent_log(message: str):
    """Append a timestamped line to data/agent.log for action-level tracing."""
    try:
        append_agent_log(str(message), category="Action")
    except Exception:
        pass


def _is_healthsherpa_clients_list_url(url: str) -> bool:
    value = str(url or "").strip().lower()
    if not value or "healthsherpa.com" not in value:
        return False
    return bool(re.search(r"/agents/[^/]+/clients(?:\?|$)", value))


def _derive_healthsherpa_clients_list_url(url: str) -> str:
    value = str(url or "").strip()
    if not value or "healthsherpa.com" not in value.lower():
        return ""
    if _is_healthsherpa_clients_list_url(value):
        return value

    m = re.search(r"(https?://[^/]+/agents/[^/]+/clients)/[^?]+(\?.*)?$", value, flags=re.IGNORECASE)
    if m:
        base = (m.group(1) or "").strip()
        query = (m.group(2) or "").strip()
        return f"{base}{query}" if base else ""
    return ""


def smart_process_all_clients(
    wait_text: str = "Sync Complete",
    wait_timeout: float = 20.0,
    max_clients: int = 10000,
    close_after_sync: bool = True,
    pagination_mode: str = "auto",
) -> dict:
    """
    Intelligently process all clients across all pages with smart pagination detection.
    
    This function:
    1. Clicks next View button to open client
    2. Waits for completion text (e.g., "Sync Complete")
    3. Closes the client tab
    4. Repeats until no more View buttons
    5. Automatically detects when to paginate to next page
    6. Stops when all pages are exhausted
    
    Args:
        wait_text: Text to wait for after opening client (default: "Sync Complete")
        wait_timeout: Seconds to wait for completion text (default: 20)
        max_clients: Safety limit to prevent infinite loops (default: 10000)
        close_after_sync: Whether to close tab after sync (default: True)
    
    Returns:
        dict with:
            - success: bool - Whether operation completed successfully
            - clients_processed: int - Number of clients processed
            - pages_processed: int - Number of pages navigated
            - error: str - Error message if any
    """
    from selenium.webdriver.common.by import By
    
    driver = _get_selenium_driver()
    if driver is None:
        return {
            "success": False,
            "clients_processed": 0,
            "pages_processed": 0,
            "error": "Selenium driver not available"
        }

    navigation_memory = None
    reflection_logger = None
    observed_states: list[str] = []
    selected_selector: dict = {}

    try:
        from .navigation_memory import NavigationMemoryStore
        from .reflection_logger import ReflectionLogger

        navigation_memory = NavigationMemoryStore()
        reflection_logger = ReflectionLogger()
    except Exception:
        navigation_memory = None
        reflection_logger = None

    def _derive_url_pattern(url: str) -> str:
        raw = str(url or "").split("?")[0]
        return re.sub(r"/\d+", "/{id}", raw)

    def _finalize_result(result: dict, failure_class: str = "") -> dict:
        try:
            current_url = str(driver.current_url or "")
        except Exception:
            current_url = ""

        site_name = "healthsherpa"
        url_pattern = _derive_url_pattern(current_url)
        goal = "smart_process_all_clients"

        if navigation_memory is not None:
            try:
                navigation_memory.upsert_site_profile(
                    site_name=site_name,
                    url_pattern=url_pattern,
                    goal=goal,
                    iframe_notes="Selenium attach workflow",
                    popup_behavior="May open client in new tab per row",
                    download_behavior="N/A for sync run",
                    meta={
                        "pagination_mode": pagination_mode,
                        "wait_text": wait_text,
                    },
                )
                navigation_memory.mark_site_outcome(
                    site_name=site_name,
                    url_pattern=url_pattern,
                    goal=goal,
                    success=bool(result.get("success")),
                )
                navigation_memory.add_task_history(
                    site_name=site_name,
                    url=current_url,
                    goal=goal,
                    status="success" if result.get("success") else "failed",
                    failure_class=failure_class,
                    details=result,
                )
            except Exception:
                pass

        if reflection_logger is not None:
            try:
                reflection_logger.build_reflection(
                    goal=goal,
                    observed_states=observed_states,
                    successful_selector=selected_selector,
                    failure_details={"failure_class": failure_class, "error": str(result.get("error", ""))} if failure_class else None,
                    memory_recommendation={
                        "site": site_name,
                        "url_pattern": url_pattern,
                        "selector": selected_selector,
                        "next_hint": "retain known-good pattern unless repeated replacement success",
                    },
                    status="success" if result.get("success") else "failed",
                )
            except Exception:
                pass

        return result
    
    print("🚀 Starting smart client processing with auto-pagination...")
    print(f"   Settings: wait_text='{wait_text}', timeout={wait_timeout}s, max={max_clients}")
    print(f"   Pagination mode: {pagination_mode}")
    
    clients_processed = 0

    detected_page = -1
    try:
        detected_page = int(_get_healthsherpa_clients_active_page_number() or -1)
    except Exception:
        detected_page = -1

    if detected_page <= 0:
        try:
            cur_url_for_page = str(driver.current_url or "")
            m = re.search(r"[?&]page=(\d+)", cur_url_for_page, re.IGNORECASE)
            if m:
                detected_page = int(m.group(1))
        except Exception:
            detected_page = -1

    if detected_page <= 0:
        detected_page = 1

    current_page = detected_page
    pages_processed = 1
    print(f"   Starting page detected: {current_page}")
    observed_states.append(f"start_page:{current_page}")
    # Remember the clients list URL so we can recover if site navigates away
    clients_list_url = None
    try:
        cur = driver.current_url or ""
        clients_list_url = _derive_healthsherpa_clients_list_url(cur)
    except Exception as e:
        _trace_action_exception("Failed to capture initial clients list URL", e)
    
    try:
        # Reset the view button counter to start from beginning
        reset_view_button_counter()
        
        while clients_processed < max_clients:
            # Try to click the next view button
            result = click_next_view_button()
            
            if result.get("success"):
                selected_selector = {
                    "action": "click_next_view_button",
                    "type": "resilient_multi_strategy",
                    "value": "role/text/button heuristics",
                }
                # Successfully clicked a view button - process this client
                client_name = result.get("client_name", f"Client {clients_processed + 1}")
                opened_new_window = result.get("opened_new_window", False)
                observed_states.append(f"client_opened:{client_name}")
                
                print(f"\n{'='*60}")
                print(f"📋 Processing client {clients_processed + 1}: {client_name}")
                print(f"{'='*60}")
                
                # Human-like pause after clicking (reaction time)
                take_human_pause(0.3, 0.8, "After clicking View button")
                
                # Wait for page to load
                print(f"⏳ Waiting for client page to load...")
                wait_for_page_load(timeout_sec=5)
                print(f"✅ Page loaded")
                
                # Wait for completion indicator
                print(f"⏳ Waiting for '{wait_text}' to appear (up to {wait_timeout}s)...")
                wait_start = time.time()
                found = wait_for_element_with_text(wait_text, timeout_sec=wait_timeout)
                wait_duration = time.time() - wait_start
                
                if found:
                    print(f"✅ '{wait_text}' found after {wait_duration:.1f}s")
                    print(f"⏸️  Giving it a moment to ensure completion...")
                    time.sleep(2.0)  # Give sync process a moment to fully complete
                    print(f"✅ Client processing complete")
                    observed_states.append(f"sync_complete:{client_name}")
                else:
                    print(f"❌ ERROR: '{wait_text}' not found within {wait_timeout}s")
                    print(f"⚠️  Skipping this client and moving to next...")
                    observed_states.append(f"sync_timeout:{client_name}")
                    # Still try to go back to continue with other clients
                
                # Return to client list
                if close_after_sync:
                    if opened_new_window:
                        # Close the new window and return to main list
                        print(f"🔄 Closing new window and returning to list...")
                        close_success = close_current_tab()
                        if close_success:
                            print(f"✅ Window closed successfully, back on client list")
                        else:
                            print(f"⚠️ Warning: Window close failed, attempting recovery...")
                            # Try to navigate back as fallback
                            try:
                                driver.back()
                                time.sleep(1.0)
                                print(f"✅ Navigated back to client list")
                            except Exception as e:
                                _trace_action_exception("Fallback driver.back failed after close_current_tab failure", e)
                    else:
                        # Same window navigation - go back to list
                        print(f"🔄 Navigating back to client list...")
                        try:
                            driver.back()
                            time.sleep(1.0)
                            wait_for_page_load(timeout_sec=3)
                            print(f"✅ Navigated back to client list")
                        except Exception as e:
                            print(f"⚠️ Navigation back failed: {e}")
                    # If we are not on the clients list anymore, recover by reopening it
                    try:
                        cur = driver.current_url or ""
                        if clients_list_url and (not _is_healthsherpa_clients_list_url(cur)):
                            print("🔁 Recovering: reopening clients list page...")
                            driver.get(clients_list_url)
                            wait_for_page_load(timeout_sec=10)
                            time.sleep(1.0)
                    except Exception as e:
                        _trace_action_exception("Failed to recover back to clients list URL", e)
                    
                    time.sleep(0.5)  # Brief pause to ensure page is ready
                
                clients_processed += 1
                print(f"✅ Client {clients_processed} complete - Ready for next client\n")
                
                # Keep brief pause between clients for stability
                take_human_pause(0.1, 0.3, "Brief break between clients")
                
            else:
                # No more view buttons on current page - try pagination
                print(f"\n{'='*60}")
                print(f"📄 No more clients on page {current_page}")
                print(f"🔍 Looking for next page button...")
                print(f"{'='*60}")
                
                next_page = current_page + 1
                observed_states.append(f"page_exhausted:{current_page}")
                
                # Try to click pagination to next page
                use_next_arrow_only = str(pagination_mode or "auto").strip().lower() in {
                    "next_arrow_only",
                    "next-arrow-only",
                    "arrow_only",
                    "arrow-only",
                    "next_only",
                }
                paginated = click_pagination(
                    page_number=next_page,
                    allow_next_control=True,
                    allow_url_fallback=False,
                    prefer_next_control_only=use_next_arrow_only,
                )

                if paginated:
                    observed_page = _get_healthsherpa_clients_active_page_number()
                    if observed_page > 0 and observed_page != next_page:
                        print(
                            f"⚠️ Pagination landed on unexpected page {observed_page} (expected {next_page}); requesting assist instead."
                        )
                        _append_agent_log(
                            f"Pagination mismatch after click: expected={next_page} observed={observed_page}; forcing assist"
                        )
                        paginated = False
                
                if paginated:
                    print(f"✅ Successfully navigated to page {next_page}")
                    pages_processed += 1
                    current_page = next_page
                    observed_states.append(f"page_advanced:{current_page}")
                    
                    # Human-like pause after pagination click
                    take_human_pause(0.5, 1.2, "After clicking pagination")
                    
                    # Wait for new page to load
                    wait_for_page_load(timeout_sec=10)
                    time.sleep(1.0)  # Extra pause for content to load
                    
                    # Reset the view button counter for new page
                    reset_view_button_counter()
                    print(f"🔄 Starting to process clients on page {current_page}...")
                    # Refresh clients_list_url if URL indicates clients list
                    try:
                        cur = driver.current_url or ""
                        refreshed_clients_url = _derive_healthsherpa_clients_list_url(cur)
                        if refreshed_clients_url:
                            clients_list_url = refreshed_clients_url
                    except Exception:
                        pass
                else:
                    print("🆘 Auto-pagination failed. Requesting human assist to go to next page...")
                    while not paginated:
                        assist = _human_assist_pagination_and_learn(current_page=current_page, expected_page=next_page)
                        if not assist.get("advanced"):
                            reason = str(assist.get("reason", ""))
                            if reason.endswith("canceled"):
                                result_payload = {
                                    "success": False,
                                    "clients_processed": clients_processed,
                                    "pages_processed": pages_processed,
                                    "error": "Pagination assist canceled by user"
                                }
                                return _finalize_result(result_payload, failure_class="modal_blocking")
                            print("⚠️ Page did not advance. Please click only NEXT, then OK.")
                            continue

                        observed_page = int(assist.get("new_page") or next_page)
                        current_page = observed_page if observed_page > 0 else next_page
                        pages_processed += 1
                        reset_view_button_counter()
                        print(f"✅ Human assist moved to page {current_page}. Learned this paginator control.")
                        paginated = True

                    # Refresh clients_list_url if URL indicates clients list
                    try:
                        cur = driver.current_url or ""
                        refreshed_clients_url = _derive_healthsherpa_clients_list_url(cur)
                        if refreshed_clients_url:
                            clients_list_url = refreshed_clients_url
                    except Exception:
                        pass
        
        # Hit max_clients safety limit
        print(f"\n⚠️ Reached maximum client limit ({max_clients})")
        result_payload = {
            "success": True,
            "clients_processed": clients_processed,
            "pages_processed": pages_processed,
            "error": f"Stopped at safety limit: {max_clients} clients"
        }
        return _finalize_result(result_payload)
        
    except Exception as e:
        error_msg = f"Error during smart client processing: {e}"
        print(f"\n❌ {error_msg}")
        result_payload = {
            "success": False,
            "clients_processed": clients_processed,
            "pages_processed": pages_processed,
            "error": error_msg
        }
        return _finalize_result(result_payload, failure_class="stale_selector")


def _normalize_match_token(text: str) -> str:
    value = str(text or "").strip().lower()
    if not value:
        return ""
    value = re.sub(r"[^a-z0-9]", "", value)
    return value


_STATE_ABBR_TO_NAME = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas", "CA": "California",
    "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware", "FL": "Florida", "GA": "Georgia",
    "HI": "Hawaii", "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi", "MO": "Missouri",
    "MT": "Montana", "NE": "Nebraska", "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey",
    "NM": "New Mexico", "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont",
    "VA": "Virginia", "WA": "Washington", "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming",
    "DC": "District of Columbia",
}


def _expand_state_value(value: str) -> tuple[str, str]:
    raw = str(value or "").strip()
    if not raw:
        return "", ""
    upper = raw.upper()
    if upper in _STATE_ABBR_TO_NAME:
        return upper, _STATE_ABBR_TO_NAME[upper]
    if len(raw) == 2:
        return upper, _STATE_ABBR_TO_NAME.get(upper, "")
    for abbr, full in _STATE_ABBR_TO_NAME.items():
        if raw.lower() == full.lower():
            return abbr, full
    return upper[:2], raw.title()


def _normalize_dob_value(value) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""

    known_formats = (
        "%m/%d/%Y", "%m-%d-%Y", "%m.%d.%Y",
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
        "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
    )
    for fmt in known_formats:
        try:
            parsed = datetime.strptime(raw, fmt)
            return parsed.strftime("%m/%d/%Y")
        except Exception:
            continue

    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) >= 8:
        core = digits[:8]
        if core[:4].isdigit() and int(core[:4]) >= 1900:
            yyyy = core[:4]
            mm = core[4:6]
            dd = core[6:8]
            return f"{mm}/{dd}/{yyyy}"
        mm = core[0:2]
        dd = core[2:4]
        yyyy = core[4:8]
        return f"{mm}/{dd}/{yyyy}"

    return raw


def _load_client_search_rows_from_excel(mapping_excel_path: str, mapping_sheet: str = "") -> list[dict]:
    path = (mapping_excel_path or "").strip()
    if not path:
        return []

    resolved = ""
    candidates = []
    if os.path.isabs(path):
        candidates.append(path)
    else:
        candidates.append(os.path.join(os.getcwd(), path))
        candidates.append(os.path.join(os.getcwd(), "data", "mappings", os.path.basename(path)))

    for candidate in candidates:
        if os.path.isfile(candidate):
            resolved = candidate
            break

    if not resolved:
        _append_agent_log(f"Client search Excel not found: {path} candidates={candidates}")
        return []

    try:
        from openpyxl import load_workbook
    except Exception as e:
        _append_agent_log(f"openpyxl unavailable for client search load: {e}")
        return []

    try:
        wb = load_workbook(resolved, data_only=True)
        if mapping_sheet and mapping_sheet in wb.sheetnames:
            ws = wb[mapping_sheet]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header = [str(c or "").strip().lower() for c in rows[0]]
        idx = {name: i for i, name in enumerate(header) if name}

        def _pick(row_vals, aliases: list[str]) -> str:
            for alias in aliases:
                i = idx.get(alias)
                if i is None or i >= len(row_vals):
                    continue
                val = row_vals[i]
                if val is None:
                    continue
                text = str(val).strip()
                if text:
                    return text
            return ""

        out = []
        for row_number, r in enumerate(rows[1:], start=2):
            first_name = _pick(r, ["first_name", "firstname", "first"])
            last_name = _pick(r, ["last_name", "lastname", "last"])
            ffm_id = _pick(r, ["ffm_id", "ffm id", "ffm", "marketplace_id", "exchange_id"])
            ssn = _pick(r, ["ssn", "social_security_number", "social_security", "social"])
            dob = _pick(r, ["dob", "date_of_birth", "date of birth", "birth_date", "birth date", "dateofbirth"])
            dob = _normalize_dob_value(dob)
            zip_code = _pick(r, ["zip", "postal", "postal_code", "zip_code"])
            email = _pick(r, ["email", "email_address"])
            phone = _pick(r, ["phone", "phone_number", "mobile"])
            coverage_state = _pick(r, ["coverage_state", "coverage state", "coverage_state_abbr", "coverage state abbr", "state", "state_abbr", "state abbr", "state abbreviation"])
            search_query = _pick(r, ["search_query", "query", "search"])

            ssn_digits = "".join(ch for ch in ssn if ch.isdigit())
            dob_digits = "".join(ch for ch in dob if ch.isdigit())
            state_abbr, state_full = _expand_state_value(coverage_state)

            if not search_query:
                # Preferred search sequence for this workflow: full SSN, DOB, state.
                query_parts = [p for p in [ssn_digits, dob, state_abbr] if p]
                if not query_parts:
                    query_parts = [p for p in [first_name, last_name, zip_code] if p]
                if not query_parts:
                    query_parts = [p for p in [first_name, last_name, dob, email, phone, state_abbr] if p]
                search_query = " ".join(query_parts).strip()

            token_candidates = [first_name, last_name, zip_code, email, state_full, state_abbr]
            phone_digits = "".join(ch for ch in phone if ch.isdigit())
            if phone_digits:
                token_candidates.append(phone_digits[-4:])
            if dob_digits:
                token_candidates.append(dob_digits)
            if ffm_id:
                token_candidates.append(ffm_id)
            # SSN may not be visible in result rows; include only last 4 for safer row matching.
            if ssn_digits and len(ssn_digits) >= 4:
                token_candidates.append(ssn_digits[-4:])

            tokens = []
            for token in token_candidates:
                normalized = _normalize_match_token(token)
                if normalized and normalized not in tokens:
                    tokens.append(normalized)

            if not search_query and not tokens:
                continue

            out.append({
                "source_row": row_number,
                "search_query": search_query,
                "match_tokens": tokens,
                "ffm_id": ffm_id,
                "ssn": ssn_digits,
                "dob": dob,
                "coverage_state_abbr": state_abbr,
                "coverage_state_full": state_full,
                "display_name": " ".join([p for p in [first_name, last_name] if p]).strip() or f"Row {row_number}",
            })

        _append_agent_log(f"Loaded {len(out)} client search rows from Excel: {resolved} (sheet={ws.title})")
        return out
    except Exception as e:
        _append_agent_log(f"Failed loading client search rows from Excel '{resolved}': {e}")
        return []


def _set_healthsherpa_clients_search_query(query: str) -> bool:
    driver = _get_selenium_driver()
    if driver is None:
        return False

    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys

    q = (query or "").strip()
    if not q:
        return False

    selectors = [
        "input[type='search']",
        "input[placeholder*='search' i]",
        "input[name*='search' i]",
        "input[id*='search' i]",
    ]

    for sel in selectors:
        try:
            elems = driver.find_elements(By.CSS_SELECTOR, sel)
            for el in elems:
                try:
                    if not el.is_displayed():
                        continue
                    el.click()
                    try:
                        el.clear()
                    except Exception:
                        pass
                    el.send_keys(q)
                    el.send_keys(Keys.ENTER)
                    _append_agent_log(f"HealthSherpa search query set: '{q}'")
                    return True
                except Exception:
                    continue
        except Exception:
            continue
    return False


def _set_healthsherpa_client_filters(ssn: str = "", dob: str = "", coverage_state_abbr: str = "", coverage_state_full: str = "") -> bool:
        driver = _get_selenium_driver()
        if driver is None:
                return False


        def _ensure_healthsherpa_marketplace_search_context(timeout_sec: float = 10.0) -> bool:
                driver = _get_selenium_driver()
                if driver is None:
                        return False

                script = """
                function normalize(s) {
                    return (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
                }

                function normalizeCompact(s) {
                    return (s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
                }

                function isVisible(el) {
                    if (!el) return false;
                    const style = window.getComputedStyle(el);
                    if (!style || style.display === 'none' || style.visibility === 'hidden') return false;
                    const rect = el.getBoundingClientRect();
                    return rect.width > 0 && rect.height > 0;
                }

                function isSearchMarketplacePage() {
                    const controls = Array.from(document.querySelectorAll('input, select, textarea')).filter(isVisible);
                    for (const el of controls) {
                        const hints = normalizeCompact(
                            (el.id || '') + ' ' +
                            (el.name || '') + ' ' +
                            (el.getAttribute('placeholder') || '') + ' ' +
                            (el.getAttribute('aria-label') || '')
                        );
                        if (hints.includes('socialsecurity') || hints.includes('ssn') || hints.includes('dateofbirth') || hints.includes('coveragestate')) {
                            return true;
                        }
                    }
                    return false;
                }

                function clickSearchMarketplace() {
                    const candidates = Array.from(document.querySelectorAll('a, button, [role="button"], input[type="button"], input[type="submit"]'));
                    for (const c of candidates) {
                        if (!isVisible(c)) continue;
                        const txt = normalize((c.innerText || c.textContent || c.value || ''));
                        if (txt.includes('search marketplace')) {
                            c.click();
                            return true;
                        }
                    }
                    return false;
                }

                function setCriteriaToSSN() {
                    // 1) Select dropdown with "Social Security Number"
                    const selects = Array.from(document.querySelectorAll('select')).filter(isVisible);
                    for (const sel of selects) {
                        const options = Array.from(sel.options || []);
                        for (const opt of options) {
                            const txt = normalize((opt.textContent || '') + ' ' + (opt.value || ''));
                            if (txt.includes('social security')) {
                                sel.value = opt.value;
                                sel.dispatchEvent(new Event('change', { bubbles: true }));
                                return true;
                            }
                        }
                    }

                    // 2) Radio/checkbox style criteria
                    const labels = Array.from(document.querySelectorAll('label, span, div')).filter(isVisible);
                    for (const l of labels) {
                        const txt = normalize(l.textContent || '');
                        if (!txt.includes('social security')) continue;

                        const forId = l.getAttribute && l.getAttribute('for');
                        if (forId) {
                            const input = document.getElementById(forId);
                            if (input && (input.type === 'radio' || input.type === 'checkbox')) {
                                input.click();
                                return true;
                            }
                        }

                        const inLabel = l.querySelector && l.querySelector('input[type="radio"], input[type="checkbox"], button, [role="radio"], [role="button"]');
                        if (inLabel) {
                            inLabel.click();
                            return true;
                        }
                    }

                    return false;
                }

                const alreadySearch = isSearchMarketplacePage();
                let navClicked = false;
                if (!alreadySearch) {
                    navClicked = clickSearchMarketplace();
                }
                const criteriaSet = setCriteriaToSSN();
                const onSearchNow = isSearchMarketplacePage();
                return { alreadySearch, navClicked, criteriaSet, onSearchNow };
                """

                end = time.time() + max(1.0, float(timeout_sec or 10.0))
                nav_clicked_once = False
                while time.time() < end:
                        try:
                                result = driver.execute_script(script)
                                if isinstance(result, dict):
                                        nav_clicked_once = nav_clicked_once or bool(result.get("navClicked", False))
                                        on_search = bool(result.get("onSearchNow", False))
                                        criteria_set = bool(result.get("criteriaSet", False))
                                        if on_search and (criteria_set or bool(result.get("alreadySearch", False))):
                                                _append_agent_log(
                                                        f"Marketplace search context ready (nav_clicked={nav_clicked_once}, criteria_set={criteria_set})"
                                                )
                                                return True
                                if nav_clicked_once:
                                        try:
                                                wait_for_page_load(timeout_sec=5)
                                        except Exception:
                                                pass
                                time.sleep(0.6)
                        except Exception as e:
                                _append_agent_log(f"Marketplace search context check failed: {e}")
                                time.sleep(0.5)

                _append_agent_log("Marketplace search context could not be confirmed")
                return False

        ssn_digits = "".join(ch for ch in str(ssn or "") if ch.isdigit())
        dob_value = _normalize_dob_value(dob)
        state_abbr = str(coverage_state_abbr or "").strip().upper()
        state_full = str(coverage_state_full or "").strip()
        learned_ssn_panel_hint_raw = ""

        try:
            from .memory import should_apply_pattern
            learned = should_apply_pattern(
                "healthsherpa_ssn_panel_assist",
                ["healthsherpa", "search by ssn", "ssn panel", "structured search fields"],
            )
            if isinstance(learned, dict):
                learned_ssn_panel_hint_raw = str(learned.get("solution", "") or "").strip()
        except Exception:
            learned_ssn_panel_hint_raw = ""

        if not ssn_digits and not dob_value and not state_abbr and not state_full:
                return False

        script = """
        const ssn = (arguments[0] || '').trim();
        const dob = (arguments[1] || '').trim();
        const stateAbbr = (arguments[2] || '').trim().toUpperCase();
        const stateFull = (arguments[3] || '').trim();
        const learnedSsnPanelHintRaw = (arguments[4] || '').trim();

        let learnedSsnPanelHint = {};
        try {
            if (learnedSsnPanelHintRaw) learnedSsnPanelHint = JSON.parse(learnedSsnPanelHintRaw);
        } catch (_) {
            learnedSsnPanelHint = {};
        }

        function normalize(s) {
            return (s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
        }

        function isVisible(el) {
            if (!el) return false;
            const style = window.getComputedStyle(el);
            if (!style || style.display === 'none' || style.visibility === 'hidden') return false;
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        }

        function fieldHints(el) {
            const attrs = [
                el.id || '',
                el.name || '',
                el.getAttribute('placeholder') || '',
                el.getAttribute('aria-label') || '',
                el.getAttribute('data-testid') || '',
            ].join(' ');
            let labelText = '';
            try {
                if (el.labels && el.labels.length) {
                    labelText = Array.from(el.labels).map(l => l.textContent || '').join(' ');
                } else if (el.id) {
                    const lbl = document.querySelector(`label[for="${CSS.escape(el.id)}"]`);
                    if (lbl) labelText = lbl.textContent || '';
                }
            } catch (_) {}
            return normalize(attrs + ' ' + labelText);
        }

        function findControlNearLabel(requiredTokens) {
            const tokens = Array.isArray(requiredTokens) ? requiredTokens.map(t => normalize(t)).filter(Boolean) : [];
            if (!tokens.length) return null;
            const labelLike = Array.from(document.querySelectorAll('label, span, div, p, strong, th, td')).filter(isVisible);

            for (const node of labelLike) {
                const txt = normalize(node.textContent || '');
                if (!txt) continue;
                if (!tokens.every(t => txt.includes(t))) continue;

                const forId = node.getAttribute && node.getAttribute('for');
                if (forId) {
                    const byFor = document.getElementById(forId);
                    if (byFor && isVisible(byFor)) return byFor;
                }

                const inNode = node.querySelector && node.querySelector('select, input, textarea, [role="combobox"], [aria-haspopup="listbox"]');
                if (inNode && isVisible(inNode)) return inNode;

                let parent = node.parentElement;
                for (let depth = 0; depth < 4 && parent; depth += 1) {
                    const found = parent.querySelector && parent.querySelector('select, input, textarea, [role="combobox"], [aria-haspopup="listbox"]');
                    if (found && isVisible(found)) return found;
                    parent = parent.parentElement;
                }

                let sib = node.nextElementSibling;
                for (let i = 0; i < 3 && sib; i += 1) {
                    const found = sib.querySelector && sib.querySelector('select, input, textarea, [role="combobox"], [aria-haspopup="listbox"]');
                    if (found && isVisible(found)) return found;
                    sib = sib.nextElementSibling;
                }
            }
            return null;
        }

        function findControlByHints(tokenList) {
            const wanted = Array.isArray(tokenList) ? tokenList.map(t => normalize(t)).filter(Boolean) : [];
            if (!wanted.length) return null;
            const controls = Array.from(document.querySelectorAll('input, select, textarea, [role="combobox"]')).filter(isVisible);
            for (const el of controls) {
                const hints = fieldHints(el);
                if (!hints) continue;
                if (wanted.some(t => hints.includes(t))) {
                    return el;
                }
            }
            return null;
        }

        function setInputValue(el, value) {
            if (!el || value == null) return false;
            try {
                el.focus();
                el.value = '';
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.value = value;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            } catch (_) {
                return false;
            }
        }

        function setSelectValue(el, desiredValues) {
            if (!el || !desiredValues || !desiredValues.length) return false;
            const opts = Array.from(el.options || []);
            const desiredNorm = desiredValues.map(v => normalize(v)).filter(Boolean);
            for (const opt of opts) {
                const optNorm = normalize((opt.value || '') + ' ' + (opt.textContent || ''));
                if (desiredNorm.some(d => d && (optNorm === d || optNorm.includes(d) || d.includes(optNorm)))) {
                    el.value = opt.value;
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                    return true;
                }
            }
            return false;
        }

        function parseDobParts(rawDob) {
            const raw = (rawDob || '').trim();
            if (!raw) return null;

            const pieces = raw.split(/[^0-9]/).filter(Boolean);
            let month = '';
            let day = '';
            let year = '';

            if (pieces.length >= 3) {
                if (pieces[0].length === 4) {
                    year = pieces[0];
                    month = pieces[1];
                    day = pieces[2];
                } else if (pieces[2].length === 4) {
                    month = pieces[0];
                    day = pieces[1];
                    year = pieces[2];
                }
            }

            if (!year) {
                const digits = raw.replace(/\\D/g, '');
                if (digits.length >= 8) {
                    if (digits.slice(0, 4) >= '1900') {
                        year = digits.slice(0, 4);
                        month = digits.slice(4, 6);
                        day = digits.slice(6, 8);
                    } else {
                        month = digits.slice(0, 2);
                        day = digits.slice(2, 4);
                        year = digits.slice(4, 8);
                    }
                }
            }

            if (!month || !day || !year) return null;
            const mm = String(parseInt(month, 10)).padStart(2, '0');
            const dd = String(parseInt(day, 10)).padStart(2, '0');
            const yyyy = String(parseInt(year, 10));
            if (!mm || !dd || !yyyy || yyyy.length < 4) return null;
            return {
                month: mm,
                day: dd,
                year: yyyy,
                mdy: `${mm}/${dd}/${yyyy}`,
                ymd: `${yyyy}-${mm}-${dd}`,
            };
        }

        function hasStructuredSearchFields() {
            const controls = Array.from(document.querySelectorAll('input, select, textarea, [role="combobox"]')).filter(isVisible);
            for (const el of controls) {
                const hints = fieldHints(el);
                if (!hints) continue;
                if (
                    hints.includes('socialsecurity') ||
                    hints.includes('socialsecuritynumber') ||
                    hints.includes('ssn') ||
                    hints.includes('dateofbirth') ||
                    hints.includes('birthdate') ||
                    hints.includes('dob') ||
                    hints.includes('coveragestate')
                ) {
                    return true;
                }
            }
            return false;
        }

        function clickOrSearchBySsn() {
            const candidates = Array.from(document.querySelectorAll('a, button, [role="button"], input[type="button"], input[type="submit"], span, div')).filter(isVisible);
            let best = null;
            let bestScore = Number.NEGATIVE_INFINITY;
            for (const el of candidates) {
                const txt = normalize((el.innerText || el.textContent || el.value || ''));
                const tag = ((el.tagName || '') + '').toLowerCase();
                const role = ((el.getAttribute && el.getAttribute('role')) || '').toLowerCase();
                const cls = ((el.className || '') + '').toLowerCase();
                let score = Number.NEGATIVE_INFINITY;

                if (txt.includes('orsearchbyssn') || txt.includes('searchbyssn')) {
                    score = 100;
                } else {
                    score = -100;
                }

                if (learnedSsnPanelHint && typeof learnedSsnPanelHint === 'object') {
                    const learnedClass = ((learnedSsnPanelHint.button_class || '') + '').toLowerCase().trim();
                    const learnedTag = ((learnedSsnPanelHint.button_tag || '') + '').toLowerCase().trim();
                    const learnedRole = ((learnedSsnPanelHint.button_role || '') + '').toLowerCase().trim();
                    if (learnedClass && cls && cls.indexOf(learnedClass) !== -1) score = Math.max(score, 80);
                    if (learnedTag && tag === learnedTag) score += 6;
                    if (learnedRole && role && role === learnedRole) score += 6;
                }

                if (score > bestScore) {
                    best = el;
                    bestScore = score;
                }
            }

            if (best && bestScore >= 60) {
                try {
                    best.scrollIntoView({ block: 'center', inline: 'center' });
                } catch (_) {}
                try {
                    best.click();
                    return true;
                } catch (_) {}
                try {
                    best.dispatchEvent(new MouseEvent('mousedown', { bubbles: true }));
                    best.dispatchEvent(new MouseEvent('mouseup', { bubbles: true }));
                    best.dispatchEvent(new MouseEvent('click', { bubbles: true }));
                    return true;
                } catch (_) {}
            }
            return false;
        }

        let ssnPanelOpened = false;
        if (!hasStructuredSearchFields()) {
            ssnPanelOpened = clickOrSearchBySsn();
        }

        let controls = Array.from(document.querySelectorAll('input, select, textarea')).filter(isVisible);
        if (!hasStructuredSearchFields()) {
            const clickedAgain = clickOrSearchBySsn();
            if (clickedAgain) ssnPanelOpened = true;
            controls = Array.from(document.querySelectorAll('input, select, textarea')).filter(isVisible);
        }
        const dobParts = parseDobParts(dob);

        let setCount = 0;
        let firstFilled = null;
        let consentChecked = false;
        let stateSelected = false;
        let ssnSet = false;
        let dobSetAny = false;

        // Step 1: SSN (explicit targeting)
        if (ssn) {
            const ssnControl =
                findControlNearLabel(['social', 'security']) ||
                findControlNearLabel(['ssn']) ||
                findControlByHints(['socialsecuritynumber', 'socialsecurity', 'ssn']);
            if (ssnControl && !isBlockedActionElement(ssnControl)) {
                const ssnOk = setInputValue(ssnControl, ssn);
                if (ssnOk) {
                    setCount += 1;
                    ssnSet = true;
                    if (!firstFilled) firstFilled = ssnControl;
                }
            }
        }

        // Step 2: DOB (explicit targeting)
        if (dob) {
            const dobControl =
                findControlNearLabel(['date', 'birth']) ||
                findControlNearLabel(['dob']) ||
                findControlByHints(['dateofbirth', 'birthdate', 'dob']);
            if (dobControl && !isBlockedActionElement(dobControl)) {
                let dobSet = false;
                const dobTag = (dobControl.tagName || '').toLowerCase();
                const dobType = (dobControl.getAttribute('type') || '').toLowerCase();
                if (dobParts) {
                    if (dobType === 'date') {
                        dobSet = setInputValue(dobControl, dobParts.ymd);
                    } else if (dobTag === 'select') {
                        dobSet = setSelectValue(dobControl, [dobParts.mdy, dobParts.ymd, dobParts.month, dobParts.day, dobParts.year]);
                    } else {
                        dobSet = setInputValue(dobControl, dobParts.mdy) || setInputValue(dobControl, dob);
                    }
                } else {
                    dobSet = setInputValue(dobControl, dob);
                }
                if (dobSet) {
                    setCount += 1;
                    dobSetAny = true;
                    if (!firstFilled) firstFilled = dobControl;
                }
            }
        }

        for (const el of controls) {
            const hints = fieldHints(el);
            const tag = (el.tagName || '').toLowerCase();

            if (ssn && (hints.includes('ssn') || hints.includes('socialsecurity'))) {
                if (setInputValue(el, ssn)) {
                    setCount += 1;
                    ssnSet = true;
                    if (!firstFilled) firstFilled = el;
                    continue;
                }
            }

            if (dob && (
                hints.includes('dateofbirth') ||
                hints.includes('birthdate') ||
                hints.includes('dob') ||
                hints.includes('birth')
            )) {
                let dobSet = false;
                if (dobParts) {
                    const isMonthHint = hints.includes('month') && (hints.includes('birth') || hints.includes('dob'));
                    const isDayHint = hints.includes('day') && (hints.includes('birth') || hints.includes('dob'));
                    const isYearHint = hints.includes('year') && (hints.includes('birth') || hints.includes('dob'));

                    if (isMonthHint) {
                        dobSet = ((tag === 'select')
                            ? setSelectValue(el, [dobParts.month, String(parseInt(dobParts.month, 10))])
                            : setInputValue(el, dobParts.month));
                    } else if (isDayHint) {
                        dobSet = ((tag === 'select')
                            ? setSelectValue(el, [dobParts.day, String(parseInt(dobParts.day, 10))])
                            : setInputValue(el, dobParts.day));
                    } else if (isYearHint) {
                        dobSet = ((tag === 'select')
                            ? setSelectValue(el, [dobParts.year])
                            : setInputValue(el, dobParts.year));
                    } else {
                        const inputType = (el.getAttribute('type') || '').toLowerCase();
                        if (inputType === 'date') {
                            dobSet = setInputValue(el, dobParts.ymd);
                        } else if (tag === 'select') {
                            dobSet = setSelectValue(el, [dobParts.mdy, dobParts.ymd, dobParts.month, dobParts.day, dobParts.year]);
                        } else {
                            dobSet = setInputValue(el, dobParts.mdy) || setInputValue(el, dob);
                        }
                    }
                } else {
                    dobSet = setInputValue(el, dob);
                }

                if (dobSet) {
                    setCount += 1;
                    dobSetAny = true;
                    if (!firstFilled) firstFilled = el;
                    continue;
                }
            }

        }

        function looksLikeCoverageStateHints(hints) {
            return (
                hints.includes('coveragestate') ||
                hints.includes('statecoverage') ||
                (hints.includes('coverage') && hints.includes('state')) ||
                hints.includes('marketplacestate')
            );
        }

        function looksLikeStateSelectByOptions(el) {
            if (!el || (el.tagName || '').toLowerCase() !== 'select') return false;
            try {
                const optionsText = Array.from(el.options || []).map(o => normalize((o.textContent || '') + ' ' + (o.value || '')));
                const hasStateWords = optionsText.some(t =>
                    t.includes('alabama') || t.includes('alaska') || t.includes('arizona') || t.includes('california') ||
                    t.includes('michigan') || t.includes('texas') || t.includes('newyork') || t.includes('florida')
                );
                return hasStateWords;
            } catch (_) {
                return false;
            }
        }

        function optionText(el) {
            return (el.innerText || el.textContent || el.value || '').trim();
        }

        function isBlockedActionElement(el) {
            const txt = normalize(optionText(el));
            if (!txt) return false;
            return txt === 'savelead' || txt.includes('savelead');
        }

        function tryClickStateOption(preferredText) {
            const desiredNorm = normalize(preferredText || '');
            if (!desiredNorm) return false;

            const listboxRoots = Array.from(document.querySelectorAll('[role="listbox"], [role="menu"], [aria-expanded="true"]')).filter(isVisible);
            const optionSelectors = [
                '[role="option"]',
                '[role="menuitem"]',
                '.select__option',
                '.Select-option',
                '.react-select__option',
                '[id*="option" i]',
            ];

            if (!listboxRoots.length) return false;

            const optionPools = [];
            for (const root of listboxRoots) {
                optionPools.push(root);
            }

            for (const root of optionPools) {
                for (const selector of optionSelectors) {
                    const opts = Array.from(root.querySelectorAll(selector)).filter(isVisible);
                    for (const opt of opts) {
                        if (isBlockedActionElement(opt)) continue;
                        const txtNorm = normalize(optionText(opt));
                        if (!txtNorm) continue;
                        if (txtNorm === desiredNorm || txtNorm.includes(desiredNorm) || desiredNorm.includes(txtNorm)) {
                            if (clickElement(opt)) return true;
                        }
                    }
                }
            }
            return false;
        }

        function setAutocompleteState(control, stateText) {
            if (!control || !stateText) return false;
            try {
                control.focus();
            } catch (_) {}
            if (!setInputValue(control, stateText)) return false;
            try {
                control.dispatchEvent(new KeyboardEvent('keydown', { key: 'ArrowDown', bubbles: true }));
                control.dispatchEvent(new KeyboardEvent('keyup', { key: 'ArrowDown', bubbles: true }));
            } catch (_) {}
            if (tryClickStateOption(stateText)) return true;
            try {
                control.dispatchEvent(new KeyboardEvent('keydown', { key: 'Enter', bubbles: true }));
                control.dispatchEvent(new KeyboardEvent('keyup', { key: 'Enter', bubbles: true }));
                return true;
            } catch (_) {
                return false;
            }
        }

        function dismissSaveLeadPopup() {
            let dismissed = false;

            const closeSelectors = [
                '[aria-label*="close" i]',
                'button[title*="close" i]',
                '[data-testid*="close" i]',
                '.modal button',
                '[role="dialog"] button',
            ];

            const candidates = [];
            for (const selector of closeSelectors) {
                for (const el of Array.from(document.querySelectorAll(selector)).filter(isVisible)) {
                    candidates.push(el);
                }
            }

            for (const el of candidates) {
                const txt = normalize((el.innerText || el.textContent || el.value || el.getAttribute('aria-label') || ''));
                if (
                    txt.includes('close') || txt.includes('cancel') || txt.includes('dismiss') ||
                    txt.includes('notnow') || txt.includes('x')
                ) {
                    if (clickElement(el)) {
                        dismissed = true;
                        break;
                    }
                }
            }

            if (!dismissed) {
                const dialogs = Array.from(document.querySelectorAll('[role="dialog"], .modal, .MuiDialog-root')).filter(isVisible);
                for (const dlg of dialogs) {
                    const dlgTxt = normalize(dlg.innerText || dlg.textContent || '');
                    if (dlgTxt.includes('savelead') || (dlgTxt.includes('save') && dlgTxt.includes('lead'))) {
                        const btns = Array.from(dlg.querySelectorAll('button, [role="button"], input[type="button"], input[type="submit"]')).filter(isVisible);
                        for (const btn of btns) {
                            const txt = normalize(btn.innerText || btn.textContent || btn.value || '');
                            if (txt.includes('cancel') || txt.includes('close') || txt.includes('notnow') || txt.includes('no')) {
                                if (clickElement(btn)) {
                                    dismissed = true;
                                    break;
                                }
                            }
                        }
                    }
                    if (dismissed) break;
                }
            }

            if (!dismissed) {
                try {
                    document.dispatchEvent(new KeyboardEvent('keydown', { key: 'Escape', bubbles: true }));
                    document.dispatchEvent(new KeyboardEvent('keyup', { key: 'Escape', bubbles: true }));
                    dismissed = true;
                } catch (_) {}
            }

            return dismissed;
        }

        // Explicit coverage-state selection pass before consent checkbox.
        if ((stateAbbr || stateFull) && !stateSelected) {
            const preferredState = stateFull || stateAbbr;

            // 0a) Observed pattern: autocomplete state control.
            const observedStateControls = Array.from(document.querySelectorAll(
                'input[name="coverage_state" i], input[name*="coverage" i][name*="state" i], input[id*="coverage" i][id*="state" i], input[placeholder*="coverage state" i], input[aria-label*="coverage state" i], input[aria-autocomplete="list" i], [role="combobox"] input'
            )).filter(isVisible);
            for (const ctrl of observedStateControls) {
                if (isBlockedActionElement(ctrl)) continue;
                if (setAutocompleteState(ctrl, stateFull || stateAbbr)) {
                    stateSelected = true;
                    setCount += 1;
                    if (!firstFilled) firstFilled = ctrl;
                    break;
                }
            }

            // 0) Strict lookup by explicit coverage state label/container.
            const labeledStateControl = !stateSelected ? (findControlNearLabel(['coverage', 'state']) || findControlNearLabel(['state'])) : null;
            if (labeledStateControl && !isBlockedActionElement(labeledStateControl)) {
                const labelTag = (labeledStateControl.tagName || '').toLowerCase();
                const labelType = (labeledStateControl.getAttribute('type') || '').toLowerCase();
                if (labelTag === 'select') {
                    if (setSelectValue(labeledStateControl, [stateFull, stateAbbr])) {
                        stateSelected = true;
                        setCount += 1;
                        if (!firstFilled) firstFilled = labeledStateControl;
                    }
                } else {
                    clickElement(labeledStateControl);
                    if (labelTag === 'input' || labelTag === 'textarea' || labelType === 'search' || labelType === 'text') {
                        setInputValue(labeledStateControl, preferredState);
                    }
                    if (tryClickStateOption(stateFull) || tryClickStateOption(stateAbbr)) {
                        stateSelected = true;
                        setCount += 1;
                        if (!firstFilled) firstFilled = labeledStateControl;
                    } else if (labelTag === 'input' || labelTag === 'textarea') {
                        if (setInputValue(labeledStateControl, preferredState)) {
                            stateSelected = true;
                            setCount += 1;
                            if (!firstFilled) firstFilled = labeledStateControl;
                        }
                    }
                }
            }

            // 1) Native select controls tied to coverage state.
            const stateSelects = controls.filter(el => {
                if ((el.tagName || '').toLowerCase() !== 'select') return false;
                const hints = fieldHints(el);
                return looksLikeCoverageStateHints(hints) || looksLikeStateSelectByOptions(el);
            });

            for (const sel of stateSelects) {
                if (setSelectValue(sel, [stateFull, stateAbbr])) {
                    stateSelected = true;
                    setCount += 1;
                    if (!firstFilled) firstFilled = sel;
                    break;
                }
            }

            // 2) Custom combobox/select widgets.
            if (!stateSelected) {
                const customCandidates = Array.from(document.querySelectorAll(
                    '[role="combobox"], input[list], input[autocomplete], input[id*="state" i], input[name*="state" i], [data-testid*="state" i], [aria-label*="state" i], [placeholder*="state" i]'
                )).filter(isVisible);

                for (const el of customCandidates) {
                    const hints = fieldHints(el);
                    if (isBlockedActionElement(el)) continue;
                    if (!looksLikeCoverageStateHints(hints)) continue;

                    const tag = (el.tagName || '').toLowerCase();
                    const inputType = (el.getAttribute('type') || '').toLowerCase();

                    const canOpenDropdown = (
                        tag === 'input' ||
                        el.getAttribute('role') === 'combobox' ||
                        (el.getAttribute('aria-haspopup') || '').toLowerCase() === 'listbox'
                    );
                    if (!canOpenDropdown) continue;

                    let opened = clickElement(el);

                    if (!opened && (tag === 'input' || tag === 'textarea' || inputType === 'search' || inputType === 'text')) {
                        opened = setInputValue(el, preferredState);
                    }

                    if (!opened) continue;

                    if (tryClickStateOption(stateFull) || tryClickStateOption(stateAbbr)) {
                        stateSelected = true;
                        setCount += 1;
                        if (!firstFilled) firstFilled = el;
                        break;
                    }

                    if (tag === 'input' || tag === 'textarea' || inputType === 'search' || inputType === 'text') {
                        if (setInputValue(el, preferredState)) {
                            stateSelected = true;
                            setCount += 1;
                            if (!firstFilled) firstFilled = el;
                            break;
                        }
                    }
                }
            }
        }

        // Ensure consent checkbox is checked before searching
        const attestation = document.querySelector('input[name="permissionAttestation"]');
        if (attestation && isVisible(attestation)) {
            try {
                if (attestation.checked !== true) {
                    attestation.click();
                }
                consentChecked = true;
            } catch (_) {}
        }

        const consentCandidates = Array.from(document.querySelectorAll('input[type="checkbox"], [role="checkbox"]')).filter(isVisible);
        for (const cb of consentCandidates) {
            if (consentChecked) break;
            let hints = '';
            try {
                hints = fieldHints(cb);
            } catch (_) {
                hints = normalize((cb.innerText || cb.textContent || cb.getAttribute('aria-label') || ''));
            }
            const looksLikeConsent = (
                hints.includes('consent') ||
                (hints.includes('received') && hints.includes('consumer')) ||
                (hints.includes('work') && hints.includes('behalf'))
            );
            if (!looksLikeConsent) continue;

            try {
                const isChecked = (cb.checked === true) || (cb.getAttribute('aria-checked') === 'true');
                if (!isChecked) {
                    cb.click();
                }
                consentChecked = true;
                break;
            } catch (_) {}
        }

        function clickElement(el) {
            if (!el || !isVisible(el)) return false;
            if (isBlockedActionElement(el)) return false;
            try {
                el.scrollIntoView({ block: 'center', inline: 'center' });
            } catch (_) {}
            try {
                el.click();
                return true;
            } catch (_) {}
            try {
                el.dispatchEvent(new MouseEvent('mousedown', { bubbles: true }));
                el.dispatchEvent(new MouseEvent('mouseup', { bubbles: true }));
                el.dispatchEvent(new MouseEvent('click', { bubbles: true }));
                return true;
            } catch (_) {}
            return false;
        }

        // Trigger search if possible (prioritize exact CTA: "Search the marketplace")
        let triggered = false;
        let triggeredBy = '';
        const buttons = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a, [role="button"]'));

        for (const btn of buttons) {
            const txt = normalize((btn.innerText || btn.textContent || btn.value || ''));
            if (txt === 'searchthemarketplace' || txt.includes('searchthemarketplace')) {
                if (clickElement(btn)) {
                    triggered = true;
                    triggeredBy = (btn.innerText || btn.textContent || btn.value || '').trim();
                    break;
                }
            }
        }

        if (!triggered) {
            for (const btn of buttons) {
                const txt = normalize((btn.innerText || btn.textContent || btn.value || ''));
                if (txt.includes('search') || txt.includes('find') || txt.includes('apply')) {
                    if (clickElement(btn)) {
                        triggered = true;
                        triggeredBy = (btn.innerText || btn.textContent || btn.value || '').trim();
                        break;
                    }
                }
            }
        }

        // HealthSherpa may show Save Lead modal after search. Dismiss it so flow can continue.
        let saveLeadDismissed = false;
        if (triggered) {
            try {
                saveLeadDismissed = dismissSaveLeadPopup() || saveLeadDismissed;
            } catch (_) {}
        }

        const structuredReady = hasStructuredSearchFields();
        return { setCount, triggered, triggeredBy, consentChecked, stateSelected, saveLeadDismissed, ssnPanelOpened, structuredReady, ssnSet, dobSet: dobSetAny };
        """

        human_assist_used = False
        try:
            for _attempt in range(2):
                result = driver.execute_script(script, ssn_digits, dob_value, state_abbr, state_full, learned_ssn_panel_hint_raw)
                if not isinstance(result, dict):
                    break

                set_count = int(result.get("setCount", 0) or 0)
                triggered = bool(result.get("triggered", False))
                triggered_by = str(result.get("triggeredBy", "") or "").strip()
                consent_checked = bool(result.get("consentChecked", False))
                state_selected = bool(result.get("stateSelected", False))
                save_lead_dismissed = bool(result.get("saveLeadDismissed", False))
                ssn_panel_opened = bool(result.get("ssnPanelOpened", False))
                structured_ready = bool(result.get("structuredReady", False))
                ssn_set = bool(result.get("ssnSet", False))
                dob_set = bool(result.get("dobSet", False))

                required_missing = (
                    (bool(ssn_digits) and not ssn_set)
                    or (bool(dob_value) and not dob_set)
                    or (bool(state_abbr or state_full) and not state_selected)
                )

                if set_count > 0 and not required_missing:
                    _append_agent_log(
                    f"Set HealthSherpa filters: ssn={bool(ssn_digits)} ssn_set={ssn_set} dob={bool(dob_value)} dob_set={dob_set} state={bool(state_abbr or state_full)} state_selected={state_selected} fields={set_count} consent_checked={consent_checked} triggered={triggered} triggered_by='{triggered_by}' save_lead_dismissed={save_lead_dismissed} ssn_panel_opened={ssn_panel_opened} structured_ready={structured_ready} human_assist_used={human_assist_used}"
                    )
                    return True

                if required_missing and not human_assist_used:
                    _append_agent_log(
                        f"Required search fields missing after auto-fill (ssn_set={ssn_set}, dob_set={dob_set}, state_selected={state_selected}); requesting human assist"
                    )
                    assist_result = _human_assist_ssn_panel_and_learn()
                    human_assist_used = bool(assist_result.get("clicked", False))
                    if human_assist_used:
                        try:
                            from .memory import should_apply_pattern
                            learned = should_apply_pattern(
                                "healthsherpa_ssn_panel_assist",
                                ["healthsherpa", "search by ssn", "ssn panel", "structured search fields"],
                            )
                            if isinstance(learned, dict):
                                learned_ssn_panel_hint_raw = str(learned.get("solution", "") or "").strip()
                        except Exception:
                            pass
                        continue

                if not structured_ready and not human_assist_used:
                    _append_agent_log("SSN panel not found automatically; requesting human assist")
                    assist_result = _human_assist_ssn_panel_and_learn()
                    human_assist_used = bool(assist_result.get("clicked", False))
                    if human_assist_used:
                        try:
                            from .memory import should_apply_pattern
                            learned = should_apply_pattern(
                                "healthsherpa_ssn_panel_assist",
                                ["healthsherpa", "search by ssn", "ssn panel", "structured search fields"],
                            )
                            if isinstance(learned, dict):
                                learned_ssn_panel_hint_raw = str(learned.get("solution", "") or "").strip()
                        except Exception:
                            pass
                        continue
                _append_agent_log(
                    f"Set HealthSherpa filters incomplete: ssn_set={ssn_set} dob_set={dob_set} state_selected={state_selected} fields={set_count} structured_ready={structured_ready} human_assist_used={human_assist_used}"
                )
                break
        except Exception as e:
                _append_agent_log(f"Failed setting HealthSherpa filter fields: {e}")

        return False


def _ensure_healthsherpa_marketplace_search_context(timeout_sec: float = 10.0) -> bool:
        driver = _get_selenium_driver()
        if driver is None:
                return False

        script = """
        function normalize(s) {
            return (s || '').toLowerCase().replace(/\\s+/g, ' ').trim();
        }

        function normalizeCompact(s) {
            return (s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
        }

        function isVisible(el) {
            if (!el) return false;
            const style = window.getComputedStyle(el);
            if (!style || style.display === 'none' || style.visibility === 'hidden') return false;
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        }

        function isSearchMarketplacePage() {
            const controls = Array.from(document.querySelectorAll('input, select, textarea')).filter(isVisible);
            for (const el of controls) {
                const hints = normalizeCompact(
                    (el.id || '') + ' ' +
                    (el.name || '') + ' ' +
                    (el.getAttribute('placeholder') || '') + ' ' +
                    (el.getAttribute('aria-label') || '')
                );
                if (hints.includes('socialsecurity') || hints.includes('ssn') || hints.includes('dateofbirth') || hints.includes('coveragestate')) {
                    return true;
                }
            }
            return false;
        }

        function clickSearchMarketplace() {
            const exactSearchMarketplace = document.querySelector('a.ffm-links[data-public="true"][href^="/ffm_redirect"]');
            if (exactSearchMarketplace && isVisible(exactSearchMarketplace)) {
                const txt = normalize((exactSearchMarketplace.innerText || exactSearchMarketplace.textContent || exactSearchMarketplace.value || ''));
                if (txt.includes('search marketplace')) {
                    exactSearchMarketplace.click();
                    return true;
                }
            }

            const candidates = Array.from(document.querySelectorAll('a, button, [role="button"], input[type="button"], input[type="submit"]'));
            for (const c of candidates) {
                if (!isVisible(c)) continue;
                const txt = normalize((c.innerText || c.textContent || c.value || ''));
                if (txt.includes('search marketplace')) {
                    c.click();
                    return true;
                }
            }
            return false;
        }

        function clickDashboard() {
            const exactDashboard = document.querySelector('a[data-public="true"][href^="/agents/"]');
            if (exactDashboard && isVisible(exactDashboard)) {
                const txt = normalize((exactDashboard.innerText || exactDashboard.textContent || exactDashboard.value || ''));
                if (txt === 'dashboard' || txt.includes(' dashboard')) {
                    exactDashboard.click();
                    return true;
                }
            }

            const candidates = Array.from(document.querySelectorAll('a, button, [role="button"], input[type="button"], input[type="submit"]'));
            for (const c of candidates) {
                if (!isVisible(c)) continue;
                const txt = normalize((c.innerText || c.textContent || c.value || ''));
                if (txt === 'dashboard' || txt.includes(' dashboard')) {
                    c.click();
                    return true;
                }
            }
            return false;
        }

        function clickOrSearchBySSN() {
            const exactSsnLink = document.querySelector('a[data-public="true"][href="#short"]');
            if (exactSsnLink && isVisible(exactSsnLink)) {
                const txt = normalize((exactSsnLink.innerText || exactSsnLink.textContent || exactSsnLink.value || ''));
                if (txt.includes('or search by ssn') || txt.includes('search by ssn')) {
                    try {
                        exactSsnLink.click();
                        return true;
                    } catch (_) {}
                }
            }

            const candidates = Array.from(document.querySelectorAll('a, button, [role="button"], span, div')).filter(isVisible);
            for (const c of candidates) {
                const txt = normalize((c.innerText || c.textContent || c.value || ''));
                if (txt.includes('or search by ssn') || txt.includes('search by ssn')) {
                    try {
                        c.click();
                        return true;
                    } catch (_) {}
                }
            }
            return false;
        }

        function setCriteriaToSSN() {
            const selects = Array.from(document.querySelectorAll('select')).filter(isVisible);
            for (const sel of selects) {
                const options = Array.from(sel.options || []);
                for (const opt of options) {
                    const txt = normalize((opt.textContent || '') + ' ' + (opt.value || ''));
                    if (txt.includes('social security')) {
                        sel.value = opt.value;
                        sel.dispatchEvent(new Event('change', { bubbles: true }));
                        return true;
                    }
                }
            }

            const labels = Array.from(document.querySelectorAll('label, span, div')).filter(isVisible);
            for (const l of labels) {
                const txt = normalize(l.textContent || '');
                if (!txt.includes('social security')) continue;

                const forId = l.getAttribute && l.getAttribute('for');
                if (forId) {
                    const input = document.getElementById(forId);
                    if (input && (input.type === 'radio' || input.type === 'checkbox')) {
                        input.click();
                        return true;
                    }
                }

                const inLabel = l.querySelector && l.querySelector('input[type="radio"], input[type="checkbox"], button, [role="radio"], [role="button"]');
                if (inLabel) {
                    inLabel.click();
                    return true;
                }
            }

            return false;
        }

        const alreadySearch = isSearchMarketplacePage();
        let navClicked = false;
        let dashboardClicked = false;
        let ssnPanelClicked = false;
        if (!alreadySearch) {
            dashboardClicked = clickDashboard();
            navClicked = clickSearchMarketplace();
            ssnPanelClicked = clickOrSearchBySSN();
        } else {
            ssnPanelClicked = clickOrSearchBySSN();
        }
        const criteriaSet = setCriteriaToSSN();
        const onSearchNow = isSearchMarketplacePage();
        return { alreadySearch, navClicked, dashboardClicked, ssnPanelClicked, criteriaSet, onSearchNow };
        """

        end = time.time() + max(1.0, float(timeout_sec or 10.0))
        nav_clicked_once = False
        while time.time() < end:
                try:
                        result = driver.execute_script(script)
                        if isinstance(result, dict):
                                nav_clicked_once = nav_clicked_once or bool(result.get("navClicked", False))
                                dashboard_clicked = bool(result.get("dashboardClicked", False))
                                ssn_panel_clicked = bool(result.get("ssnPanelClicked", False))
                                on_search = bool(result.get("onSearchNow", False))
                                criteria_set = bool(result.get("criteriaSet", False))
                                if on_search and (criteria_set or bool(result.get("alreadySearch", False))):
                                        _append_agent_log(
                                                f"Marketplace search context ready (dashboard_clicked={dashboard_clicked}, nav_clicked={nav_clicked_once}, ssn_panel_clicked={ssn_panel_clicked}, criteria_set={criteria_set})"
                                        )
                                        return True
                        if nav_clicked_once:
                                try:
                                        wait_for_page_load(timeout_sec=5)
                                except Exception:
                                        pass
                        time.sleep(0.6)
                except Exception as e:
                        _append_agent_log(f"Marketplace search context check failed: {e}")
                        time.sleep(0.5)

        _append_agent_log("Marketplace search context could not be confirmed")
        return False


def _wait_for_overlay_clear(timeout_sec: float = 6.0) -> bool:
    driver = _get_selenium_driver()
    if driver is None:
        return False

    script = r"""
    function isVisible(el) {
        if (!el) return false;
        const style = window.getComputedStyle(el);
        if (!style || style.display === 'none' || style.visibility === 'hidden') return false;
        if (Number(style.opacity || '1') === 0) return false;
        const rect = el.getBoundingClientRect();
        return rect.width > 0 && rect.height > 0;
    }

    const selectors = [
        '.MuiBackdrop-root',
        '.MuiModal-backdrop',
        '.MuiCircularProgress-root',
        '.MuiLinearProgress-root',
        '.loading',
        '.loader',
        '.spinner',
        '[data-testid*="loading"]',
        '[aria-busy="true"]',
        '[role="progressbar"]'
    ];

    for (const sel of selectors) {
        const nodes = Array.from(document.querySelectorAll(sel));
        for (const node of nodes) {
            if (isVisible(node)) return false;
        }
    }

    return true;
    """

    end = time.time() + max(1.0, float(timeout_sec or 1.0))
    while time.time() < end:
        try:
            clear = bool(driver.execute_script(script))
            if clear:
                return True
        except Exception:
            return False
        time.sleep(0.2)
    return False


def _retry_click_add_client_for_tokens(
    match_tokens: list[str],
    add_button_text: str,
    expected_ffm_id: str,
    virtual_grid_mode: bool,
    max_attempts: int = 6,
) -> dict:
    attempts = max(1, int(max_attempts or 1))
    last_result = {"clicked": False, "reason": "no_attempts"}

    for attempt in range(1, attempts + 1):
        overlay_cleared = _wait_for_overlay_clear(timeout_sec=6.0)
        _append_agent_log(
            f"Search/Add click attempt {attempt}/{attempts} ffm_id='{expected_ffm_id}' overlay_cleared={overlay_cleared}"
        )

        result = _click_add_client_for_tokens(
            match_tokens,
            add_button_text=add_button_text,
            expected_ffm_id=expected_ffm_id,
            virtual_grid_mode=virtual_grid_mode,
        )
        if isinstance(result, dict):
            last_result = result
        else:
            last_result = {"clicked": False, "reason": "invalid_click_result"}

        if last_result.get("clicked"):
            last_result["clickAttempt"] = attempt
            return last_result

        if attempt < attempts:
            time.sleep(min(1.0, 0.15 * attempt))

    return last_result


def _retry_open_client_profile_by_expected_id(
        expected_ffm_id: str,
        add_button_text: str = "",
        search_url_before_click: str = "",
        max_attempts: int = 6,
) -> dict:
        driver = _get_selenium_driver()
        if driver is None:
                return {"opened": False, "reason": "driver_unavailable"}

        expected_raw = str(expected_ffm_id or "").strip()
        if not expected_raw:
                return {"opened": False, "reason": "missing_expected_ffm_id"}

        attempts = max(1, int(max_attempts or 1))
        target_text = str(add_button_text or "").strip().lower()

        script = r"""
        const expectedRaw = (arguments[0] || '').trim();
        const targetText = (arguments[1] || '').trim().toLowerCase();
        const expectedDigits = expectedRaw.replace(/\D/g, '');
        const phrases = [targetText, 'view in dash', 'view dash', 'view dashboard', 'view', 'add client', 'add to clients']
            .filter(Boolean);

        function isVisible(el) {
            if (!el) return false;
            const style = window.getComputedStyle(el);
            if (!style || style.display === 'none' || style.visibility === 'hidden') return false;
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        }

        function rowMatches(rowText) {
            const txt = String(rowText || '');
            if (!expectedDigits) return false;
            const digits = txt.replace(/\D/g, '');
            return !!digits && digits.indexOf(expectedDigits) !== -1;
        }

        function findBestActionInRow(row) {
            const clickables = Array.from(row.querySelectorAll('button, a, [role="button"], input[type="button"], input[type="submit"]')).filter(isVisible);
            let best = null;
            let bestScore = Number.NEGATIVE_INFINITY;
            for (const c of clickables) {
                const text = ((c.innerText || c.textContent || c.value || '') + '').trim().toLowerCase();
                const href = String(c.getAttribute('href') || c.href || '').trim();
                const tag = (c.tagName || '').toLowerCase();
                let score = 0;
                if (phrases.some(p => p && text.includes(p))) score += 30;
                if (text.includes('view')) score += 20;
                if (text.includes('dash')) score += 12;
                if (href && href.indexOf('/person_search') === -1) score += 14;
                if (tag === 'button') score += 8;
                if (tag === 'a') score += 6;
                if (score > bestScore) {
                    best = c;
                    bestScore = score;
                }
            }
            return best;
        }

        const rows = Array.from(document.querySelectorAll('.MuiDataGrid-row, [role="row"], tr, li, .table-row, .client-row')).filter(isVisible);
        for (const row of rows) {
            const rowText = (row.innerText || row.textContent || '').trim();
            if (!rowText || !rowMatches(rowText)) continue;
            const action = findBestActionInRow(row);
            if (!action) continue;
            try { action.scrollIntoView({ block: 'center', inline: 'center' }); } catch (_) {}
            const href = String(action.getAttribute('href') || action.href || '').trim();
            let clicked = false;
            try {
                ['pointerdown','mousedown','mouseup','click'].forEach(type => {
                    action.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window, button: 0, buttons: 1 }));
                });
                clicked = true;
            } catch (_) {}
            if (!clicked) {
                try {
                    action.click();
                    clicked = true;
                } catch (_) {}
            }
            return { matched: true, clicked, href, buttonText: ((action.innerText || action.textContent || action.value || '') + '').trim().toLowerCase() };
        }

        return { matched: false, clicked: false, href: '' };
        """

        for attempt in range(1, attempts + 1):
                _wait_for_overlay_clear(timeout_sec=6.0)
                try:
                        result = driver.execute_script(script, expected_raw, target_text)
                except Exception as e:
                        result = {"matched": False, "clicked": False, "href": "", "error": str(e)}

                matched = bool((result or {}).get("matched", False))
                clicked = bool((result or {}).get("clicked", False))
                href = str((result or {}).get("href", "") or "").strip()

                _append_agent_log(
                        f"Search/Add open retry attempt {attempt}/{attempts} ffm_id='{expected_raw}' matched={matched} clicked={clicked} href_present={bool(href)}"
                )

                moved = False
                move_deadline = time.time() + 2.2
                while clicked and time.time() < move_deadline:
                        try:
                                cur_now = str(driver.current_url or "")
                        except Exception:
                                cur_now = ""
                        if search_url_before_click and cur_now and (cur_now != search_url_before_click):
                                moved = True
                                break
                        time.sleep(0.15)

                if moved:
                        return {"opened": True, "reason": "relocate_click_navigated", "attempt": attempt}

                if href:
                        try:
                                driver.get(href)
                                return {"opened": True, "reason": "relocate_href_navigation", "attempt": attempt, "href": href}
                        except Exception:
                                pass

                if attempt < attempts:
                        time.sleep(min(0.8, 0.15 * attempt))

        return {"opened": False, "reason": "relocate_open_failed"}


def _click_add_client_for_tokens(
    match_tokens: list[str],
    add_button_text: str = "Add client",
    expected_ffm_id: str = "",
    virtual_grid_mode: bool = True,
) -> dict:
    driver = _get_selenium_driver()
    if driver is None:
        return {"clicked": False, "reason": "driver_unavailable"}

    target_text = (add_button_text or "Add client").strip().lower()
    normalized_tokens = [_normalize_match_token(t) for t in (match_tokens or []) if _normalize_match_token(t)]
    expected_ffm_id = str(expected_ffm_id or "").strip()

    if expected_ffm_id and virtual_grid_mode:
        _append_agent_log(
            f"Search/Add expected-id flow using mui_row_id_scan fallback ffm_id='{expected_ffm_id}'"
        )

    learned_hint_raw = ""
    try:
        from .memory import should_apply_pattern
        learned = should_apply_pattern(
            "healthsherpa_id_button_assist",
            ["healthsherpa", "expected id visible", "action button"],
        )
        if isinstance(learned, dict):
            learned_hint_raw = str(learned.get("solution", "") or "").strip()
    except Exception:
        learned_hint_raw = ""

    script = r"""
    const tokens = Array.isArray(arguments[0]) ? arguments[0] : [];
    const addText = (arguments[1] || 'add client').toLowerCase();
        const expectedFfmIdRaw = (arguments[2] || '').trim();
        const learnedHintRaw = (arguments[3] || '').trim();
    const buttonPhrases = [addText, 'add client', 'add to clients', 'view dash', 'view dashboard'];

        let learnedHint = {};
        try {
            if (learnedHintRaw) learnedHint = JSON.parse(learnedHintRaw);
        } catch (_) {
            learnedHint = {};
        }

    function norm(s) {
      return (s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
    }

        function hasAllTokens(text) {
            const n = norm(text);
            if (!tokens.length) return true;
            return tokens.every(t => n.includes(t));
    }

        function normId(s) {
            return (s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
        }

        function digitsOnly(s) {
            return (s || '').replace(/\\D/g, '');
        }

        function hasExpectedDigits(text, expectedRaw) {
            const expectedDigits = digitsOnly(expectedRaw);
            if (!expectedDigits) return false;
            const txtDigits = digitsOnly(text || '');
            if (!txtDigits) return false;
            return txtDigits.includes(expectedDigits);
        }

        function idMatchesExpected(idValue, expectedRaw) {
            const expectedNorm = normId(expectedRaw);
            if (!expectedNorm) return true;

            const idNorm = normId(idValue || '');
            if (idNorm === expectedNorm) return true;

            const expectedDigits = digitsOnly(expectedRaw);
            if (expectedDigits) {
                const idDigits = digitsOnly(idValue || '');
                if (idDigits && idDigits === expectedDigits) return true;
            }

            return false;
        }

        function extractIds(text) {
            const found = [];
            const src = String(text || '');
            const regex = new RegExp('\\\\bID\\\\b\\\\s*(?:[:#-])\\\\s*([A-Za-z0-9-]+(?:\\\\s+[A-Za-z0-9-]+){0,4})', 'gi');
            let m;
            while ((m = regex.exec(src)) !== null) {
                const raw = (m[1] || '').trim();
                const cleaned = normId(raw);
                if (cleaned && !found.includes(cleaned)) {
                    found.push(cleaned);
                }
            }
            return found;
        }

        function isVisible(el) {
            if (!el) return false;
            try {
                const style = window.getComputedStyle(el);
                if (!style || style.display === 'none' || style.visibility === 'hidden') return false;
                const rect = el.getBoundingClientRect();
                return rect.width > 0 && rect.height > 0;
            } catch (_) {
                return true;
            }
        }

        function findActionButton(container) {
            if (!container) return null;
            const clickables = Array.from(container.querySelectorAll('button, a, [role="button"], input[type="button"], input[type="submit"]')).filter(isVisible);
            let best = null;
            let bestScore = Number.NEGATIVE_INFINITY;
            for (const c of clickables) {
                const cText = ((c.innerText || c.textContent || c.value || '') + '').trim().toLowerCase();
                const cTag = ((c.tagName || '') + '').toLowerCase();
                const cRole = ((c.getAttribute && c.getAttribute('role')) || '').toLowerCase();

                if (c.disabled) continue;
                let rect;
                try {
                    rect = c.getBoundingClientRect();
                } catch (_) {
                    continue;
                }
                if (!rect || rect.width < 18 || rect.height < 14) continue;

                const phraseMatch = buttonPhrases.some(p => p && cText.includes(p));
                let score = 4;
                if (phraseMatch || cText === 'add') score += 6;
                if (cTag === 'button' || cRole === 'button') score += 8;
                if (cTag === 'input') score += 6;
                if (cTag === 'a' && !cRole && !phraseMatch) score -= 2;
                score += Math.min(8, Math.max(0, rect.left / 260));

                if (learnedHint && typeof learnedHint === 'object') {
                    const learnedClass = ((learnedHint.button_class || '') + '').toLowerCase().trim();
                    const learnedTag = ((learnedHint.button_tag || '') + '').toLowerCase().trim();
                    const learnedRole = ((learnedHint.button_role || '') + '').toLowerCase().trim();
                    const cls = ((c.className || '') + '').toLowerCase();
                    if (learnedClass && cls && cls.indexOf(learnedClass) !== -1) score += 5;
                    if (learnedTag && cTag === learnedTag) score += 3;
                    if (learnedRole && cRole && cRole === learnedRole) score += 3;
                }

                if (score > bestScore) {
                    best = c;
                    bestScore = score;
                }
            }
            return best;
        }

        function resolveActionUrl(btn, container) {
            function normalizeUrl(u) {
                const raw = (u || '').trim();
                if (!raw) return '';
                const low = raw.toLowerCase();
                if (low.startsWith('http://') || low.startsWith('https://')) return raw;
                if (raw.startsWith('/')) return window.location.origin + raw;
                if (raw.indexOf('/') > 0 && raw.indexOf('://') === -1) return window.location.origin + '/' + raw;
                return raw;
            }

            function looksLikeClientUrl(u) {
                const x = (u || '').toLowerCase();
                if (!x) return false;
                if (x.startsWith('javascript:')) return false;
                if (x.startsWith('http')) {
                    return x.includes('healthsherpa.com') && (
                        x.includes('/agents/') ||
                        x.includes('/clients') ||
                        x.includes('/applications') ||
                        x.includes('/marketplace')
                    );
                }
                return x.startsWith('/') || x.includes('/agents/') || x.includes('/applications');
            }

            function extractFirstUrlFromText(txt) {
                const src = String(txt || '');
                try {
                    const urlPattern = new RegExp("https?:\\\\/\\\\/[^\\\\s'\"\\\\)]+|\\\\/[A-Za-z0-9_\\\\-\\\\/?=&.%#]+", "i");
                    const m = src.match(urlPattern);
                    if (m && m[0]) return normalizeUrl(m[0]);
                } catch (_) {
                    return '';
                }
                return '';
            }

            function candidateUrl(el) {
                if (!el) return '';
                const attrNames = ['href', 'data-href', 'data-url', 'data-link', 'data-to', 'to', 'routerlink', 'ng-reflect-router-link'];
                for (const name of attrNames) {
                    try {
                        const v = normalizeUrl((el.getAttribute && el.getAttribute(name)) || '');
                        if (v && looksLikeClientUrl(v)) return v;
                    } catch (_) {}
                }

                const onClick = ((el.getAttribute && el.getAttribute('onclick')) || '') + '';
                const fromClick = extractFirstUrlFromText(onClick);
                if (fromClick && looksLikeClientUrl(fromClick)) return fromClick;

                try {
                    const allAttrs = Array.from(el.attributes || []);
                    for (const a of allAttrs) {
                        const val = normalizeUrl(a && a.value ? a.value : '');
                        if (val && looksLikeClientUrl(val)) return val;
                    }
                } catch (_) {}

                function scanObjForUrl(obj, depth, seen) {
                    if (!obj || depth <= 0) return '';
                    const t = typeof obj;
                    if (t === 'string') {
                        const s = normalizeUrl(obj);
                        if (s && looksLikeClientUrl(s)) return s;
                        const s2 = extractFirstUrlFromText(obj);
                        if (s2 && looksLikeClientUrl(s2)) return s2;
                        return '';
                    }
                    if (t !== 'object' && t !== 'function') return '';
                    if (seen.has(obj)) return '';
                    seen.add(obj);
                    let keys = [];
                    try {
                        keys = Object.keys(obj);
                    } catch (_) {
                        return '';
                    }
                    for (const k of keys) {
                        let v;
                        try {
                            v = obj[k];
                        } catch (_) {
                            continue;
                        }
                        const found = scanObjForUrl(v, depth - 1, seen);
                        if (found) return found;
                    }
                    return '';
                }

                try {
                    const ownKeys = Object.keys(el);
                    for (const k of ownKeys) {
                        if (!(k.startsWith('__reactProps$') || k.startsWith('__reactFiber$'))) continue;
                        const v = el[k];
                        const found = scanObjForUrl(v, 4, new Set());
                        if (found) return found;
                    }
                } catch (_) {}

                return '';
            }

            const direct = candidateUrl(btn);
            if (direct) return direct;

            const scope = container || (btn && btn.closest && btn.closest('.MuiDataGrid-row, tr, [role="row"], li, .table-row, .client-row')) || null;
            if (scope) {
                const links = Array.from(scope.querySelectorAll('a[href], [data-href], [data-url], [data-link], [data-to], [to], [routerlink], [onclick], button, [role="button"]')).filter(isVisible);
                for (const link of links) {
                    const url = candidateUrl(link);
                    if (url) return url;
                }

                const scopeText = (scope.innerText || scope.textContent || '') + '';
                const fromScope = extractFirstUrlFromText(scopeText);
                if (fromScope && looksLikeClientUrl(fromScope)) return fromScope;
            }

            return '';
        }

        function markOpenTarget(el) {
            if (!el || !isVisible(el)) return '';
            const id = 'jarvis-open-' + Date.now().toString(36) + '-' + Math.random().toString(36).slice(2, 8);
            try {
                el.setAttribute('data-jarvis-open-id', id);
                return id;
            } catch (_) {
                return '';
            }
        }

        function isRowScopedIdMatch(el, expectedRaw) {
            const expectedNorm = normId(expectedRaw);
            if (!expectedNorm) return true;
            if (!el) return false;

            const scope = el.closest('.MuiDataGrid-row, tr, [role="row"], li, .table-row, .client-row') || el;
            const scopeText = (scope.innerText || scope.textContent || '').trim();
            const ids = extractIds(scopeText);
            if (ids.length) {
                return ids.some(idVal => idMatchesExpected(idVal, expectedRaw));
            }

            const flat = normId(scopeText);
            return flat.includes('id' + expectedNorm);
        }

        function getMUIVirtualScrollers() {
            const roots = Array.from(document.querySelectorAll('div.MuiDataGrid-virtualScroller[role="presentation"], div.MuiDataGrid-virtualScroller')).filter(isVisible);
            return roots;
        }

        function getScrollerAncestor(el) {
            if (!el || !el.closest) return null;
            return el.closest('div.MuiDataGrid-virtualScroller[role="presentation"], div.MuiDataGrid-virtualScroller');
        }

        function findNearestActionButton(anchorEl) {
            if (!anchorEl) return null;
            let anchorRect;
            try {
                anchorRect = anchorEl.getBoundingClientRect();
            } catch (_) {
                return null;
            }

            const anchorY = (anchorRect.top + anchorRect.bottom) / 2;
            const anchorScroller = getScrollerAncestor(anchorEl);
            const candidates = Array.from(document.querySelectorAll('button, a, [role="button"], input[type="button"], input[type="submit"]')).filter(isVisible);
            let best = null;
            let bestDy = Number.POSITIVE_INFINITY;
            let bestDx = Number.POSITIVE_INFINITY;

            for (const c of candidates) {
                const cText = ((c.innerText || c.textContent || c.value || '') + '').trim().toLowerCase();
                const cTag = ((c.tagName || '') + '').toLowerCase();
                const cRole = ((c.getAttribute && c.getAttribute('role')) || '').toLowerCase();
                if (c.disabled) continue;

                if (anchorScroller) {
                    const cScroller = getScrollerAncestor(c);
                    if (cScroller !== anchorScroller) continue;
                }

                let r;
                try {
                    r = c.getBoundingClientRect();
                } catch (_) {
                    continue;
                }
                if (r.width < 18 || r.height < 14) continue;
                const cy = (r.top + r.bottom) / 2;
                const cx = (r.left + r.right) / 2;
                const ax = (anchorRect.left + anchorRect.right) / 2;
                const dy = Math.abs(cy - anchorY);
                const dx = Math.abs(cx - ax);

                const phraseMatch = buttonPhrases.some(p => p && cText.includes(p));
                let weightDy = dy;
                let weightDx = dx;
                if (cx >= ax) {
                    weightDx = Math.max(0, dx - 30);
                }
                if (cTag === 'button' || cRole === 'button' || cTag === 'input') {
                    weightDy = Math.max(0, weightDy - 10);
                }
                if (phraseMatch || cText === 'add') {
                    weightDy = Math.max(0, weightDy - 20);
                }

                if (learnedHint && typeof learnedHint === 'object') {
                    const learnedClass = ((learnedHint.button_class || '') + '').toLowerCase().trim();
                    const learnedTag = ((learnedHint.button_tag || '') + '').toLowerCase().trim();
                    const learnedRole = ((learnedHint.button_role || '') + '').toLowerCase().trim();
                    const cls = ((c.className || '') + '').toLowerCase();
                    if ((learnedClass && cls && cls.indexOf(learnedClass) !== -1)
                        || (learnedTag && cTag === learnedTag)
                        || (learnedRole && cRole && cRole === learnedRole)) {
                        weightDy = Math.max(0, dy - 40);
                    }
                }

                if (weightDy < bestDy || (weightDy === bestDy && weightDx < bestDx)) {
                    best = c;
                    bestDy = weightDy;
                    bestDx = weightDx;
                }
            }

            if (best && bestDy <= 90) {
                return best;
            }
            return null;
        }

        function rowMatchesExpectedId(rowText) {
            const expected = normId(expectedFfmIdRaw);
            if (!expected) return true;
            const ids = extractIds(rowText || '');
            if (ids.length) {
                return ids.some(idVal => idMatchesExpected(idVal, expectedFfmIdRaw));
            }
            if (hasExpectedDigits(rowText || '', expectedFfmIdRaw)) {
                return true;
            }
            const flat = normId(rowText || '');
            return flat.includes('id' + expected);
        }

    const expected = normId(expectedFfmIdRaw);
    const seenIds = [];

    function collectSeenIds(ids) {
      for (const idVal of (ids || [])) {
        if (idVal && !seenIds.includes(idVal)) {
          seenIds.push(idVal);
        }
      }
    }

    // Pass 1 (strict): scan page for explicit "ID:" text, match expected ffm_id, then click row action.
    if (expected) {
            // 1a) Target MUI DataGrid rows first when present.
            const scrollers = getMUIVirtualScrollers();
            for (const scroller of scrollers) {
                const originalTop = Number(scroller.scrollTop || 0);
                const maxTop = Math.max(0, Number(scroller.scrollHeight || 0) - Number(scroller.clientHeight || 0));
                const stepCount = Math.max(1, Math.min(24, Math.ceil((maxTop + 1) / Math.max(1, Number(scroller.clientHeight || 1)))));

                for (let step = 0; step <= stepCount; step += 1) {
                    const targetTop = (stepCount <= 0) ? 0 : Math.round((maxTop * step) / stepCount);
                    try {
                        scroller.scrollTop = targetTop;
                        scroller.dispatchEvent(new Event('scroll', { bubbles: true }));
                        void scroller.offsetHeight;
                    } catch (_) {}

                    const muiRows = Array.from(scroller.querySelectorAll('.MuiDataGrid-row, [role="row"]')).filter(isVisible);
                    for (const row of muiRows) {
                        const rowText = (row.innerText || row.textContent || '').trim();
                        if (!rowText) continue;
                        const ids = extractIds(rowText);
                        if (ids.length) collectSeenIds(ids);
                        const idMatched = ids.length && ids.some(idVal => idMatchesExpected(idVal, expectedFfmIdRaw));
                        const digitMatched = hasExpectedDigits(rowText, expectedFfmIdRaw);
                        if (!(idMatched || digitMatched)) continue;

                        const btn = findActionButton(row);
                        if (btn) {
                            const btnText = ((btn.innerText || btn.textContent || btn.value || '') + '').trim().toLowerCase();
                            const actionUrl = resolveActionUrl(btn, row);
                            if (actionUrl) {
                                return {
                                    clicked: true,
                                    rowText: rowText.slice(0, 240),
                                    buttonText: btnText,
                                    matchedBy: 'mui_row_id_scan_url',
                                    openedByUrl: true,
                                    actionUrl: actionUrl,
                                    seenIds: seenIds.slice(0, 20),
                                };
                            }
                            const openTargetId = markOpenTarget(btn);
                            if (!openTargetId) {
                                return {
                                    clicked: false,
                                    reason: 'new_tab_open_attempt_failed',
                                    rowText: rowText.slice(0, 240),
                                    buttonText: btnText,
                                    seenIds: seenIds.slice(0, 20),
                                };
                            }
                            return {
                                clicked: true,
                                rowText: rowText.slice(0, 240),
                                buttonText: btnText,
                                matchedBy: 'mui_row_id_scan_new_tab_target',
                                openTargetId: openTargetId,
                                seenIds: seenIds.slice(0, 20),
                            };
                        }
                    }
                }

                try {
                    scroller.scrollTop = originalTop;
                    scroller.dispatchEvent(new Event('scroll', { bubbles: true }));
                } catch (_) {}
            }

      const idNodes = Array.from(document.querySelectorAll('div, span, p, td, th, li, tr, [role="row"], [class*="result" i], [class*="client" i], [class*="card" i]')).filter(isVisible);
      for (const node of idNodes) {
        const nodeText = (node.innerText || node.textContent || '').trim();
        if (!nodeText || nodeText.toLowerCase().indexOf('id') === -1) continue;
        const ids = extractIds(nodeText);
        if (ids.length) collectSeenIds(ids);
        const idMatched = ids.length && ids.some(idVal => idMatchesExpected(idVal, expectedFfmIdRaw));
        const digitMatched = hasExpectedDigits(nodeText, expectedFfmIdRaw);
        if (!(idMatched || digitMatched)) continue;

                const container = node.closest('.MuiDataGrid-row, tr, [role="row"], li, .table-row, .client-row') || node.parentElement;
        const btn = findActionButton(container);
        if (btn) {
                    if (!isRowScopedIdMatch(btn, expectedFfmIdRaw)) {
                        continue;
                    }
          const actionUrl = resolveActionUrl(btn, container);
          const rowText = (container && (container.innerText || container.textContent || '').trim()) || nodeText;
          const btnText = ((btn.innerText || btn.textContent || btn.value || '') + '').trim().toLowerCase();
                    if (actionUrl) {
                        return {
                            clicked: true,
                            rowText: rowText.slice(0, 240),
                            buttonText: btnText,
                            matchedBy: 'id_scan_url',
                            openedByUrl: true,
                            actionUrl: actionUrl,
                            seenIds: seenIds.slice(0, 20),
                        };
                    }
          const openTargetId = markOpenTarget(btn);
                    if (!openTargetId) {
                        return {
                            clicked: false,
                            reason: 'new_tab_open_attempt_failed',
                            rowText: rowText.slice(0, 240),
                            buttonText: btnText,
                            seenIds: seenIds.slice(0, 20),
                        };
                    }
                    return {
                        clicked: true,
                        rowText: rowText.slice(0, 240),
                        buttonText: btnText,
                        matchedBy: 'id_scan_new_tab_target',
                        openTargetId: openTargetId,
                        seenIds: seenIds.slice(0, 20),
                    };
        }

                const nearBtn = findNearestActionButton(node);
                if (nearBtn) {
                    const nearText = ((nearBtn.innerText || nearBtn.textContent || nearBtn.value || '') + '').trim().toLowerCase();
                    const rowText = (container && (container.innerText || container.textContent || '').trim()) || nodeText;
                    const nearUrl = resolveActionUrl(nearBtn, container || node.parentElement);
                    if (nearUrl) {
                        return {
                            clicked: true,
                            rowText: rowText.slice(0, 240),
                            buttonText: nearText,
                            matchedBy: 'id_scan_proximity_url',
                            openedByUrl: true,
                            actionUrl: nearUrl,
                            seenIds: seenIds.slice(0, 20),
                        };
                    }
                    const openTargetId = markOpenTarget(nearBtn);
                    if (!openTargetId) {
                        return {
                            clicked: false,
                            reason: 'new_tab_open_attempt_failed',
                            rowText: rowText.slice(0, 240),
                            buttonText: nearText,
                            seenIds: seenIds.slice(0, 20),
                        };
                    }
                    return {
                        clicked: true,
                        rowText: rowText.slice(0, 240),
                        buttonText: nearText,
                        matchedBy: 'id_scan_proximity_new_tab_target',
                        openTargetId: openTargetId,
                        seenIds: seenIds.slice(0, 20),
                    };
                }
      }
    }

    const rowCandidates = Array.from(document.querySelectorAll('tr, [role="row"], .table-row, .client-row, li'));
        let fallbackMatch = null;
    for (const row of rowCandidates) {
      const rowText = (row.innerText || row.textContent || '').trim();
      if (!rowText) continue;
      if (!hasAllTokens(rowText)) continue;
            if (!rowMatchesExpectedId(rowText)) continue;

      const clickables = Array.from(row.querySelectorAll('button, a, [role="button"], input[type="button"], input[type="submit"]'));
      for (const c of clickables) {
        const cText = ((c.innerText || c.textContent || c.value || '') + '').trim().toLowerCase();
        if (!cText) continue;
                const phraseMatch = buttonPhrases.some(p => p && cText.includes(p));
                if (phraseMatch || cText === 'add') {
                    if (!tokens.length) {
                        if (fallbackMatch) {
                            return { clicked: false, reason: 'ambiguous_no_tokens_multiple_add_rows' };
                        }
                        fallbackMatch = { node: c, rowText: rowText.slice(0, 240), buttonText: cText, rowNode: row };
                    } else {
                        const actionUrl = resolveActionUrl(c, row);
                        if (actionUrl) {
                            return { clicked: true, rowText: rowText.slice(0, 240), buttonText: cText, matchedBy: expected ? 'row_id_tokens_url' : 'tokens_url', openedByUrl: true, actionUrl: actionUrl, seenIds: seenIds.slice(0, 20) };
                        }
                                                                                                const openTargetId = markOpenTarget(c);
                                                                                                if (!openTargetId) {
                                                        return { clicked: false, reason: 'new_tab_open_attempt_failed', rowText: rowText.slice(0, 240), buttonText: cText, seenIds: seenIds.slice(0, 20) };
                                                }
                                                                                                                                                                                                return { clicked: true, rowText: rowText.slice(0, 240), buttonText: cText, matchedBy: expected ? 'row_id_tokens_new_tab_target' : 'tokens_new_tab_target', openTargetId: openTargetId, seenIds: seenIds.slice(0, 20) };
                    }
        }
      }
    }

        if (fallbackMatch && fallbackMatch.node) {
            const actionUrl = resolveActionUrl(fallbackMatch.node, fallbackMatch.rowNode || null);
            if (actionUrl) {
                return { clicked: true, rowText: fallbackMatch.rowText, buttonText: fallbackMatch.buttonText, reason: 'single_add_fallback', matchedBy: expected ? 'row_id_fallback_url' : 'fallback_url', openedByUrl: true, actionUrl: actionUrl, seenIds: seenIds.slice(0, 20) };
            }
            const openTargetId = markOpenTarget(fallbackMatch.node);
            if (!openTargetId) {
                return { clicked: false, reason: 'new_tab_open_attempt_failed', rowText: fallbackMatch.rowText, buttonText: fallbackMatch.buttonText, seenIds: seenIds.slice(0, 20) };
            }
                        return { clicked: true, rowText: fallbackMatch.rowText, buttonText: fallbackMatch.buttonText, reason: 'single_add_fallback', matchedBy: expected ? 'row_id_fallback_new_tab_target' : 'fallback_new_tab_target', openTargetId: openTargetId, seenIds: seenIds.slice(0, 20) };
        }

    if (expectedFfmIdRaw) {
            return { clicked: false, reason: 'no_row_with_matching_ffm_id_and_add_button', seenIds: seenIds.slice(0, 30) };
    }
        return { clicked: false, reason: 'no_token_row_with_add_button', seenIds: seenIds.slice(0, 30) };
    """

    try:
        result = driver.execute_script(script, normalized_tokens, target_text, expected_ffm_id, learned_hint_raw)
        if isinstance(result, dict):
            return result
    except Exception as e:
        return {"clicked": False, "reason": f"script_error:{e}"}

    return {"clicked": False, "reason": "unknown"}


def _digits_only(text: str) -> str:
    return re.sub(r"\D+", "", str(text or ""))


def _seen_ids_match_expected(seen_ids: list, expected_ffm_id: str) -> bool:
    expected_digits = _digits_only(expected_ffm_id)
    if not expected_digits:
        return False
    for raw in (seen_ids or []):
        if _digits_only(raw) == expected_digits:
            return True
    return False


def _direct_click_client_action_by_expected_id(expected_ffm_id: str, add_button_text: str = "") -> dict:
    driver = _get_selenium_driver()
    if driver is None:
        return {"clicked": False, "reason": "driver_unavailable"}

    expected = str(expected_ffm_id or "").strip()
    if not expected:
        return {"clicked": False, "reason": "missing_expected_ffm_id"}

    learned_hint_raw = ""
    learned_hint = {}
    try:
        from .memory import should_apply_pattern
        learned = should_apply_pattern(
            "healthsherpa_id_button_assist",
            ["healthsherpa", "expected id visible", "action button"],
        )
        if isinstance(learned, dict):
            learned_hint_raw = str(learned.get("solution", "") or "").strip()
            if learned_hint_raw:
                _append_agent_log(
                    f"Search/Add learned id-button pattern loaded ffm_id='{expected}'"
                )
                try:
                    learned_hint = json.loads(learned_hint_raw)
                except Exception:
                    learned_hint = {}
    except Exception:
        learned_hint_raw = ""
        learned_hint = {}

    target_text = (add_button_text or "").strip().lower()
    script = r"""
    const expectedRaw = (arguments[0] || '').trim();
    const targetText = (arguments[1] || '').trim().toLowerCase();
    const learnedHintRaw = (arguments[2] || '').trim();
    const expectedNorm = (expectedRaw || '').toLowerCase().replace(/[^a-z0-9]/g, '');
    const phrases = [targetText, 'add client', 'add to clients', 'view in dash', 'view dash', 'view dashboard'].filter(Boolean);
    const idRegex = new RegExp('\\bID\\b\\s*(?:[:#-])\\s*([A-Za-z0-9-]+(?:\\s+[A-Za-z0-9-]+){0,4})', 'gi');

    let learnedHint = {};
    try {
        if (learnedHintRaw) learnedHint = JSON.parse(learnedHintRaw);
    } catch (_) {
        learnedHint = {};
    }

    function norm(s) {
        return (s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
    }

    function isVisible(el) {
        if (!el) return false;
        const style = window.getComputedStyle(el);
        if (!style || style.display === 'none' || style.visibility === 'hidden') return false;
        const rect = el.getBoundingClientRect();
        return rect.width > 0 && rect.height > 0;
    }

    function extractIds(text) {
        const found = [];
        const src = String(text || '');
        idRegex.lastIndex = 0;
        let m;
        while ((m = idRegex.exec(src)) !== null) {
            const cleaned = norm((m[1] || '').trim());
            if (cleaned && !found.includes(cleaned)) found.push(cleaned);
        }
        return found;
    }

    function rowMatchesExpected(rowText) {
        const ids = extractIds(rowText || '');
        if (ids.some(v => v === expectedNorm)) return true;
        const flat = norm(rowText || '');
        return expectedNorm && flat.includes('id' + expectedNorm);
    }

    function clickActionInRow(row) {
        if (!row) return null;
        const rowText = (row.innerText || row.textContent || '').trim();
        if (!rowText || !rowMatchesExpected(rowText)) return null;

        const clickables = Array.from(row.querySelectorAll('button, a, [role="button"], input[type="button"], input[type="submit"]')).filter(isVisible);
        if (!clickables.length) return null;

        let best = null;
        let bestScore = Number.NEGATIVE_INFINITY;
        for (const c of clickables) {
            const cText = ((c.innerText || c.textContent || c.value || '') + '').trim().toLowerCase();
            const href = String(c.getAttribute('href') || c.href || '').trim();
            const tag = (c.tagName || '').toLowerCase();
            const role = ((c.getAttribute && c.getAttribute('role')) || '').toLowerCase();
            let score = 0;
            if (phrases.some(p => p && cText.includes(p))) score += 30;
            if (cText.includes('view') || cText.includes('dash') || cText.includes('add')) score += 18;
            if (href && href.indexOf('/person_search') === -1) score += 14;
            if (tag === 'button' || tag === 'input') score += 10;
            if (tag === 'a' || role === 'button') score += 8;
            const learnedClass = ((learnedHint.button_class || '') + '').toLowerCase().trim();
            const learnedTag = ((learnedHint.button_tag || '') + '').toLowerCase().trim();
            const learnedRole = ((learnedHint.button_role || '') + '').toLowerCase().trim();
            const cls = ((c.className || '') + '').toLowerCase();
            if (learnedClass && cls && cls.indexOf(learnedClass) !== -1) score += 25;
            if (learnedTag && tag === learnedTag) score += 10;
            if (learnedRole && role === learnedRole) score += 8;
            try {
                const r = c.getBoundingClientRect();
                score += Math.max(0, Math.min(20, r.left / 60));
                score += Math.max(0, Math.min(10, r.width / 30));
            } catch (_) {}
            if (score > bestScore) {
                bestScore = score;
                best = c;
            }
        }

        if (!best) return null;
        const bestText = ((best.innerText || best.textContent || best.value || '') + '').trim().toLowerCase();
        const bestHref = String(best.getAttribute('href') || best.href || '').trim();

        function resolveActionUrlFromRow(primary, rowNode) {
            const rawPrimary = String(primary || '').trim();
            if (rawPrimary && rawPrimary.indexOf('/person_search') === -1) return rawPrimary;
            const scope = rowNode || row;
            if (!scope) return rawPrimary;
            const links = Array.from(scope.querySelectorAll('a[href]'));
            for (const link of links) {
                const href = String(link.getAttribute('href') || link.href || '').trim();
                if (!href) continue;
                const low = href.toLowerCase();
                if (low.includes('/clients/') || low.includes('/applications/') || low.includes('/agents/')) return href;
            }
            for (const link of links) {
                const href = String(link.getAttribute('href') || link.href || '').trim();
                if (href && href.indexOf('/person_search') === -1) return href;
            }
            return rawPrimary;
        }

        const actionUrl = resolveActionUrlFromRow(bestHref, row);
        const openTargetId = 'jarvis-open-' + Date.now().toString(36) + '-' + Math.random().toString(36).slice(2, 8);
        try { best.setAttribute('data-jarvis-open-id', openTargetId); } catch (_) {}
        try { best.scrollIntoView({ block: 'center', inline: 'center' }); } catch (_) {}
        const bestClass = ((best.className || '') + '').toLowerCase();
        return { clicked: true, rowText: rowText.slice(0, 240), buttonText: bestText, href: bestHref, actionUrl: actionUrl, score: bestScore, openTargetId: openTargetId, buttonClass: bestClass };
        return null;
    }

    function scanVisibleRows(root) {
        const scope = root || document;
        const rows = Array.from(scope.querySelectorAll('.MuiDataGrid-row, [role="row"], tr, li, .table-row, .client-row')).filter(isVisible);
        for (const row of rows) {
            const result = clickActionInRow(row);
            if (result) return result;
        }
        return null;
    }

    const scrollers = Array.from(document.querySelectorAll('div.MuiDataGrid-virtualScroller[role="presentation"], div.MuiDataGrid-virtualScroller')).filter(isVisible);
    for (const scroller of scrollers) {
        const originalTop = Number(scroller.scrollTop || 0);
        const maxTop = Math.max(0, Number(scroller.scrollHeight || 0) - Number(scroller.clientHeight || 0));
        const steps = Math.max(1, Math.min(28, Math.ceil((maxTop + 1) / Math.max(1, Number(scroller.clientHeight || 1)))));

        for (let step = 0; step <= steps; step += 1) {
            const targetTop = (steps <= 0) ? 0 : Math.round((maxTop * step) / steps);
            try {
                scroller.scrollTop = targetTop;
                scroller.dispatchEvent(new Event('scroll', { bubbles: true }));
                void scroller.offsetHeight;
            } catch (_) {}

            const hit = scanVisibleRows(scroller);
            if (hit) return hit;
        }

        try {
            scroller.scrollTop = originalTop;
            scroller.dispatchEvent(new Event('scroll', { bubbles: true }));
        } catch (_) {}
    }

    const fallback = scanVisibleRows(null);
    if (fallback) return fallback;
    return { clicked: false, reason: 'expected_id_row_action_not_found' };
    """

    try:
        result = driver.execute_script(script, expected, target_text, learned_hint_raw)
        if isinstance(result, dict):
            if result.get("clicked"):
                selected_class = str(result.get("buttonClass", "") or "")
                learned_class = str(learned_hint.get("button_class", "") or "")
                used_learned = bool(learned_class and selected_class and (learned_class.lower() in selected_class.lower()))
                _append_agent_log(
                    f"Search/Add direct click selection ffm_id='{expected}' used_learned_pattern={used_learned} score={result.get('score', '')} button='{result.get('buttonText', '')}'"
                )
            return result
        return {"clicked": False, "reason": "unexpected_script_result"}
    except Exception as e:
        return {"clicked": False, "reason": f"script_error:{e}"}


def _enable_human_click_capture() -> bool:
    driver = _get_selenium_driver()
    if driver is None:
        return False
    script = """
    try {
      function normId(s) {
        return (s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
      }
      function extractIds(text) {
        const found = [];
        const src = String(text || '');
        const regex = /\\bID\\b\\s*(?:[:#-])\\s*([A-Za-z0-9-]+(?:\\s+[A-Za-z0-9-]+){0,4})/gi;
        let m;
        while ((m = regex.exec(src)) !== null) {
          const raw = (m[1] || '').trim();
          const cleaned = normId(raw);
          if (cleaned && !found.includes(cleaned)) found.push(cleaned);
        }
        return found;
      }

      if (!window.__jarvisHumanAssistCaptureInstalled) {
        window.__jarvisHumanAssistCaptureInstalled = true;
        window.__jarvisHumanAssistLastClick = null;
        document.addEventListener('click', function(ev) {
          try {
            const t = ev.target;
            if (!t) return;
            const row = t.closest('.MuiDataGrid-row, tr, [role="row"], li, .table-row, .client-row');
            const rowText = (row && (row.innerText || row.textContent || '').trim()) || '';
            const allText = ((t.innerText || t.textContent || t.value || '') + ' ' + rowText).trim();
            const ids = extractIds(allText);
            const rect = t.getBoundingClientRect();
            window.__jarvisHumanAssistLastClick = {
              timestamp: Date.now(),
              tag: (t.tagName || '').toLowerCase(),
              role: (t.getAttribute && t.getAttribute('role')) || '',
              button_text: ((t.innerText || t.textContent || t.value || '') + '').trim().toLowerCase(),
              button_class: ((t.className || '') + '').toLowerCase(),
              aria_label: ((t.getAttribute && t.getAttribute('aria-label')) || '').toLowerCase(),
              row_text: rowText.slice(0, 240),
              ids: ids.slice(0, 12),
              x: Math.round((rect.left + rect.right) / 2),
              y: Math.round((rect.top + rect.bottom) / 2),
            };
          } catch (_) {}
        }, true);
      } else {
        window.__jarvisHumanAssistLastClick = null;
      }
      return true;
    } catch (_) {
      return false;
    }
    """
    try:
        return bool(driver.execute_script(script))
    except Exception:
        return False


def _read_human_click_capture() -> dict:
    driver = _get_selenium_driver()
    if driver is None:
        return {}
    try:
        result = driver.execute_script("return (window.__jarvisHumanAssistLastClick || {});")
        if isinstance(result, dict):
            return result
    except Exception:
        pass
    return {}


def _ask_user_for_human_assist(expected_ffm_id: str) -> bool:
    msg = (
        "I found the expected ID but could not safely find its button.\n\n"
        f"Please click the correct row action button for ID: {expected_ffm_id}\n"
        "(e.g., View Dash / Add to clients), then click OK."
    )
    try:
        from tkinter import messagebox
        return bool(messagebox.askokcancel("Human Assist Needed", msg))
    except Exception:
        try:
            print("\n[Human Assist Needed] " + msg)
            ans = input("Type 'ok' after you click the correct button (or anything else to cancel): ").strip().lower()
            return ans in {"ok", "y", "yes"}
        except Exception:
            return False


def _ask_user_for_ssn_panel_assist() -> bool:
    msg = (
        "I could not find the SSN/DOB/state search fields automatically.\n\n"
        "Please click whatever gets to that search area (for example 'Or search by SSN' or the SSN field),\n"
        "then click OK so I can learn it for next time."
    )
    try:
        from tkinter import messagebox
        return bool(messagebox.askokcancel("Human Assist Needed", msg))
    except Exception:
        try:
            print("\n[Human Assist Needed] " + msg)
            ans = input("Type 'ok' after you open the SSN search panel (or anything else to cancel): ").strip().lower()
            return ans in {"ok", "y", "yes"}
        except Exception:
            return False


def _ask_user_for_reset_navigation_assist() -> bool:
    msg = (
        "Before I reset for the next row, please navigate back to the Marketplace search page the way you prefer.\n\n"
        "When you are back on the search page, click OK so I continue."
    )
    try:
        from tkinter import messagebox
        return bool(messagebox.askokcancel("Human Assist: Return to Search", msg))
    except Exception:
        try:
            print("\n[Human Assist: Return to Search] " + msg)
            ans = input("Type 'ok' after you return to the search page (or anything else to cancel): ").strip().lower()
            return ans in {"ok", "y", "yes"}
        except Exception:
            return False


def _get_healthsherpa_clients_active_page_number() -> int:
        from .pagination import get_active_page_number
        return get_active_page_number(_get_selenium_driver())


def _ask_user_for_pagination_assist(current_page: int, expected_page: int) -> bool:
    from .pagination import ask_user_for_pagination_assist
    return ask_user_for_pagination_assist(current_page, expected_page)


def _human_assist_pagination_and_learn(current_page: int, expected_page: int) -> dict:
    from .memory import add_learning_pattern
    from .pagination import human_assist_pagination_and_learn

    return human_assist_pagination_and_learn(
        _get_selenium_driver(),
        current_page=current_page,
        expected_page=expected_page,
        enable_human_click_capture=_enable_human_click_capture,
        read_human_click_capture=_read_human_click_capture,
        append_log=_append_agent_log,
        add_learning_pattern=add_learning_pattern,
    )


def _wait_for_sync_or_timeout(wait_text: str = "Sync Complete", timeout_sec: float = 20.0) -> bool:
    end = time.time() + max(1.0, float(timeout_sec or 20.0))
    while time.time() < end:
        try:
            if _is_text_visible_on_page(wait_text):
                return True
        except Exception:
            pass
        time.sleep(0.4)
    return False


def _human_assist_ssn_panel_and_learn() -> dict:
    if not _enable_human_click_capture():
        _append_agent_log("SSN panel human assist capture init failed")
        return {"clicked": False, "reason": "ssn_panel_assist_capture_init_failed"}

    proceed = _ask_user_for_ssn_panel_assist()
    if not proceed:
        _append_agent_log("SSN panel human assist canceled")
        return {"clicked": False, "reason": "ssn_panel_assist_canceled"}

    time.sleep(0.3)
    click_info = _read_human_click_capture()
    if not click_info:
        _append_agent_log("SSN panel human assist completed but no click was captured")
        return {"clicked": False, "reason": "ssn_panel_assist_no_click_captured"}

    button_text = str((click_info or {}).get("button_text", "") or "").strip().lower()
    button_class = str((click_info or {}).get("button_class", "") or "").strip().lower()
    button_tag = str((click_info or {}).get("tag", "") or "").strip().lower()
    button_role = str((click_info or {}).get("role", "") or "").strip().lower()
    click_x = int((click_info or {}).get("x", 0) or 0)
    click_y = int((click_info or {}).get("y", 0) or 0)

    try:
        from .memory import add_learning_pattern
        solution = json.dumps(
            {
                "button_text": button_text,
                "button_class": button_class[:120],
                "button_tag": button_tag,
                "button_role": button_role,
                "click_x": click_x,
                "click_y": click_y,
            }
        )
        add_learning_pattern(
            pattern_type="healthsherpa_ssn_panel_assist",
            context="healthsherpa search by ssn panel not visible; use learned click to open structured search fields",
            solution=solution,
            success_count=1,
        )
    except Exception as e:
        _append_agent_log(f"SSN panel human assist learning save failed: {e}")

    _append_agent_log(
        f"SSN panel human assist captured click button='{button_text}' tag='{button_tag}' role='{button_role}'"
    )
    return {
        "clicked": True,
        "reason": "ssn_panel_human_assist",
        "buttonText": button_text,
    }


def _human_assist_click_and_learn(expected_ffm_id: str) -> dict:
    expected = str(expected_ffm_id or "").strip()
    if not expected:
        return {"clicked": False, "reason": "missing_expected_ffm_id"}

    if not _enable_human_click_capture():
        _append_agent_log("Human assist capture init failed")
        return {"clicked": False, "reason": "human_assist_capture_init_failed"}

    proceed = _ask_user_for_human_assist(expected)
    if not proceed:
        _append_agent_log(f"Human assist canceled for ffm_id='{expected}'")
        return {"clicked": False, "reason": "human_assist_canceled"}

    time.sleep(0.3)
    click_info = _read_human_click_capture()
    ids = click_info.get("ids", []) if isinstance(click_info, dict) else []
    button_text = str((click_info or {}).get("button_text", "") or "").strip().lower()
    button_class = str((click_info or {}).get("button_class", "") or "").strip().lower()
    button_tag = str((click_info or {}).get("tag", "") or "").strip().lower()
    button_role = str((click_info or {}).get("role", "") or "").strip().lower()
    click_x = int((click_info or {}).get("x", 0) or 0)
    click_y = int((click_info or {}).get("y", 0) or 0)

    if not click_info:
        _append_agent_log(f"Human assist completed but no click was captured for ffm_id='{expected}'")
        return {"clicked": False, "reason": "human_assist_no_click_captured"}

    if not _seen_ids_match_expected(ids, expected):
        _append_agent_log(
            f"Human assist captured click but IDs did not match expected ffm_id='{expected}' captured_ids={ids}"
        )

    try:
        from .memory import add_learning_pattern
        solution = json.dumps(
            {
                "button_class": button_class[:120],
                "button_tag": button_tag,
                "button_role": button_role,
                "click_x": click_x,
                "click_y": click_y,
            }
        )
        add_learning_pattern(
            pattern_type="healthsherpa_id_button_assist",
            context="healthsherpa expected id visible but action button not auto-bound",
            solution=solution,
            success_count=1,
        )
    except Exception as e:
        _append_agent_log(f"Human assist learning save failed: {e}")

    _append_agent_log(
        f"Human assist captured click for ffm_id='{expected}' button='{button_text}' tag='{button_tag}' role='{button_role}' ids={ids}"
    )
    return {
        "clicked": True,
        "reason": "human_assist",
        "buttonText": button_text,
        "matchedBy": "human_assist",
        "seenIds": ids,
    }


def smart_search_and_add_clients(
    mapping_excel_path: str,
    mapping_sheet: str = "",
    strict_match: bool = True,
    max_clients: int = 500,
    add_button_text: str = "Add client",
    search_wait_sec: float = 1.0,
    wait_text: str = "Sync Complete",
    wait_timeout: float = 20.0,
    close_after_sync: bool = True,
    search_context_url: str = "",
    virtual_grid_mode: bool = True,
) -> dict:
    driver = _get_selenium_driver()
    if driver is None:
        return {
            "success": False,
            "processed": 0,
            "added": 0,
            "failed": 0,
            "error": "Selenium driver not available",
            "failures": [],
        }

    rows = _load_client_search_rows_from_excel(mapping_excel_path, mapping_sheet)
    if not rows:
        return {
            "success": False,
            "processed": 0,
            "added": 0,
            "failed": 0,
            "error": "No client search rows loaded from Excel",
            "failures": [],
        }

    processed = 0
    added = 0
    failed = 0
    failures = []
    clients_list_url = ""
    preferred_search_url = str(search_context_url or "").strip()
    default_search_url = "https://www.healthsherpa.com/person_search?use_case=fetch&_agent_id=jared-chapdelaine-mccullough#short"
    if (not preferred_search_url) and default_search_url:
        preferred_search_url = default_search_url
    try:
        cur = (driver.current_url or "").strip()
        if "healthsherpa.com" in cur and "/person_search" in cur:
            preferred_search_url = cur
        if "healthsherpa.com" in cur and "/clients" in cur:
            clients_list_url = cur
    except Exception:
        clients_list_url = ""

    for row in rows[: max(1, int(max_clients))]:
        processed += 1
        row_id = row.get("source_row")
        display_name = row.get("display_name", f"Row {row_id}")
        query = str(row.get("search_query", "") or "").strip()
        tokens = row.get("match_tokens", []) or []
        ffm_id = str(row.get("ffm_id", "") or "").strip()
        ssn = str(row.get("ssn", "") or "")
        dob = str(row.get("dob", "") or "")
        coverage_state_abbr = str(row.get("coverage_state_abbr", "") or "")
        coverage_state_full = str(row.get("coverage_state_full", "") or "")

        print(f"\n🔎 Search/Add {processed}/{len(rows)}: {display_name} (row {row_id})")

        if not _ensure_healthsherpa_marketplace_search_context(timeout_sec=10.0):
            # Non-blocking: user may already be on the correct search page even if detector misses it.
            _append_agent_log(
                f"Search/Add warning row={row_id}: marketplace context not auto-confirmed; proceeding with current page"
            )

        if not query and not (ssn or dob or coverage_state_abbr or coverage_state_full):
            failed += 1
            reason = "missing_search_query"
            failures.append({"row": row_id, "name": display_name, "reason": reason})
            _append_agent_log(f"Search/Add failed row={row_id}: {reason}")
            if strict_match:
                return {
                    "success": False,
                    "processed": processed,
                    "added": added,
                    "failed": failed,
                    "error": f"Strict stop: {reason} on row {row_id}",
                    "failures": failures,
                }
            continue

        used_structured_filters = _set_healthsherpa_client_filters(
            ssn=ssn,
            dob=dob,
            coverage_state_abbr=coverage_state_abbr,
            coverage_state_full=coverage_state_full,
        )

        if not used_structured_filters and query and (not _set_healthsherpa_clients_search_query(query)):
            failed += 1
            reason = "search_input_not_found"
            failures.append({"row": row_id, "name": display_name, "reason": reason})
            _append_agent_log(f"Search/Add failed row={row_id}: {reason}")
            if strict_match:
                return {
                    "success": False,
                    "processed": processed,
                    "added": added,
                    "failed": failed,
                    "error": f"Strict stop: {reason} on row {row_id}",
                    "failures": failures,
                }
            continue

        if not used_structured_filters and not query:
            failed += 1
            reason = "filter_fields_not_found"
            failures.append({"row": row_id, "name": display_name, "reason": reason})
            _append_agent_log(f"Search/Add failed row={row_id}: {reason}")
            if strict_match:
                return {
                    "success": False,
                    "processed": processed,
                    "added": added,
                    "failed": failed,
                    "error": f"Strict stop: {reason} on row {row_id}",
                    "failures": failures,
                }
            continue

        time.sleep(max(10.0, float(search_wait_sec or 1.0)))

        windows_before = []
        current_handle_before = ""
        search_url_before_click = ""
        search_title_before_click = ""
        try:
            windows_before = list(driver.window_handles)
            current_handle_before = driver.current_window_handle
            search_url_before_click = str(driver.current_url or "")
            search_title_before_click = str(driver.title or "")
        except Exception:
            windows_before = []
            current_handle_before = ""
            search_url_before_click = ""
            search_title_before_click = ""

        if ffm_id:
            click_result = _retry_click_add_client_for_tokens(
                tokens,
                add_button_text=add_button_text,
                expected_ffm_id=ffm_id,
                virtual_grid_mode=virtual_grid_mode,
                max_attempts=6,
            )
        else:
            click_result = _click_add_client_for_tokens(
                tokens,
                add_button_text=add_button_text,
                expected_ffm_id=ffm_id,
                virtual_grid_mode=virtual_grid_mode,
            )
        if not click_result.get("clicked"):
            reason = str(click_result.get("reason", "") or "")
            seen_ids = click_result.get("seenIds", []) if isinstance(click_result, dict) else []
            if (
                ffm_id
                and reason == "no_row_with_matching_ffm_id_and_add_button"
                and _seen_ids_match_expected(seen_ids, ffm_id)
            ):
                _append_agent_log(
                    f"Search/Add human assist requested row={row_id} ffm_id='{ffm_id}' seen_ids={seen_ids}"
                )
                assist_result = _human_assist_click_and_learn(ffm_id)
                if assist_result.get("clicked"):
                    click_result = assist_result

        if click_result.get("clicked"):
            added += 1
            pre_click_sync_visible = _is_text_visible_on_page(wait_text)
            clicked_button_text = str(click_result.get("buttonText", "") or "").strip().lower()
            matched_by = str(click_result.get("matchedBy", "") or "").strip().lower()
            _append_agent_log(
                f"Search/Add success row={row_id} query='{query}' ffm_id='{ffm_id}' button='{click_result.get('buttonText', '')}' matched_by='{click_result.get('matchedBy', '')}'"
            )
            if click_result.get("clickAttempt"):
                _append_agent_log(
                    f"Search/Add click succeeded on attempt={click_result.get('clickAttempt')} row={row_id} ffm_id='{ffm_id}'"
                )
            opened_new_window = False
            opened_client_profile = False
            opened_by_url = bool(click_result.get("openedByUrl", False))
            open_target_id = str(click_result.get("openTargetId", "") or "").strip()
            action_url = str(click_result.get("actionUrl", "") or "").strip()
            target_element = None
            prefer_same_tab_open = True

            post_click_delay = 0.8
            if open_target_id and (matched_by == "direct_expected_id_click"):
                post_click_delay = 0.05
            time.sleep(post_click_delay)

            if opened_by_url and action_url:
                try:
                    driver.switch_to.new_window('tab')
                    driver.get(action_url)
                    opened_client_profile = True
                    _append_agent_log(f"Opened client in new tab via action URL row={row_id} url='{action_url[:180]}'")
                    time.sleep(0.4)
                except Exception as e:
                    _append_agent_log(f"Failed to open client in new tab via action URL row={row_id}: {e}")

            if (not opened_client_profile) and ffm_id:
                open_retry = _retry_open_client_profile_by_expected_id(
                    expected_ffm_id=ffm_id,
                    add_button_text=add_button_text,
                    search_url_before_click=search_url_before_click,
                    max_attempts=6,
                )
                if open_retry.get("opened"):
                    opened_client_profile = True
                    _append_agent_log(
                        f"Opened client profile via re-locate retry row={row_id} ffm_id='{ffm_id}' reason='{open_retry.get('reason', '')}' attempt={open_retry.get('attempt', '')}"
                    )

            if (not opened_by_url) and open_target_id and prefer_same_tab_open:
                try:
                    target_selector = f'[data-jarvis-open-id="{open_target_id}"]'
                    immediate_result = driver.execute_script(
                        """
                        const selector = arguments[0];
                        const el = document.querySelector(selector);
                        if (!el) return { found: false, clicked: false, href: '' };
                        try { el.scrollIntoView({ block: 'center', inline: 'center' }); } catch (_) {}
                        const href = String(el.getAttribute('href') || el.href || '').trim();
                        let clicked = false;
                        try {
                            ['pointerdown','mousedown','mouseup','click'].forEach(type => {
                                el.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window, button: 0, buttons: 1 }));
                            });
                            clicked = true;
                        } catch (_) {}
                        if (!clicked) {
                            try {
                                el.click();
                                clicked = true;
                            } catch (_) {}
                        }
                        return { found: true, clicked, href };
                        """,
                        target_selector,
                    )
                    immediate_clicked = bool((immediate_result or {}).get("clicked", False))
                    immediate_href = str((immediate_result or {}).get("href", "") or "").strip()

                    if immediate_clicked:
                        moved = False
                        move_deadline = time.time() + 2.2
                        while time.time() < move_deadline:
                            try:
                                cur_now = str(driver.current_url or "")
                            except Exception:
                                cur_now = ""
                            if search_url_before_click and cur_now and (cur_now != search_url_before_click):
                                moved = True
                                break
                            time.sleep(0.15)

                        if moved:
                            opened_client_profile = True
                            _append_agent_log(
                                f"Opened client profile via immediate same-tab click row={row_id} target_id='{open_target_id}'"
                            )
                        elif immediate_href:
                            try:
                                driver.get(immediate_href)
                                opened_client_profile = True
                                _append_agent_log(
                                    f"Opened client profile via immediate href navigation row={row_id} target_id='{open_target_id}'"
                                )
                                time.sleep(0.4)
                            except Exception as nav_e:
                                _append_agent_log(
                                    f"Immediate href navigation failed row={row_id} target_id='{open_target_id}': {nav_e}"
                                )
                except Exception as e:
                    _append_agent_log(f"Immediate same-tab open failed row={row_id} target_id='{open_target_id}': {e}")

            if (not opened_by_url) and open_target_id and (not opened_client_profile):
                _append_agent_log(
                    f"Skipping Selenium Ctrl+click path row={row_id} target_id='{open_target_id}' (disabled; using URL/same-tab strategy)"
                )

            try:
                windows_after = list(driver.window_handles)
                if len(windows_after) > len(windows_before):
                    new_window = [w for w in windows_after if w not in windows_before][0]
                    driver.switch_to.window(new_window)
                    opened_new_window = True
                    opened_client_profile = True
            except Exception:
                opened_new_window = False

            # Fallback: open in current tab when tab-opening is blocked.
            if (not opened_new_window) and open_target_id and (not opened_client_profile):
                try:
                    target_selector = f'[data-jarvis-open-id="{open_target_id}"]'
                    fallback_result = driver.execute_script(
                        """
                        const selector = arguments[0];
                        const el = document.querySelector(selector);
                        if (!el) return { found: false, clicked: false, href: '' };
                        try { el.scrollIntoView({ block: 'center', inline: 'center' }); } catch (_) {}
                        const href = String(el.getAttribute('href') || el.href || '').trim();
                        let clicked = false;
                        try {
                            ['pointerdown','mousedown','mouseup','click'].forEach(type => {
                                el.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window, button: 0, buttons: 1 }));
                            });
                            clicked = true;
                        } catch (_) {}
                        if (!clicked) {
                            try {
                                el.click();
                                clicked = true;
                            } catch (_) {}
                        }
                        return { found: true, clicked, href };
                        """,
                        target_selector,
                    )
                    clicked_fallback = bool((fallback_result or {}).get("clicked", False))
                    href_fallback = str((fallback_result or {}).get("href", "") or "").strip()
                    if (not clicked_fallback) and href_fallback:
                        driver.get(href_fallback)
                        clicked_fallback = True
                        _append_agent_log(
                            f"Tab open blocked; navigated directly to client href row={row_id} target_id='{open_target_id}'"
                        )
                    if not clicked_fallback:
                        raise RuntimeError("target click did not execute")
                    _append_agent_log(f"Tab open blocked; clicked client in same tab row={row_id} target_id='{open_target_id}'")
                    time.sleep(0.6)

                    # If click fired but page did not move, force navigate by href as recovery.
                    if href_fallback:
                        try:
                            cur_after_fallback = str(driver.current_url or "")
                        except Exception:
                            cur_after_fallback = ""
                        if (
                            search_url_before_click
                            and cur_after_fallback
                            and (cur_after_fallback == search_url_before_click)
                        ):
                            try:
                                driver.get(href_fallback)
                                _append_agent_log(
                                    f"Same-tab recovery via href row={row_id} target_id='{open_target_id}'"
                                )
                                time.sleep(0.5)
                            except Exception as nav_e:
                                _append_agent_log(
                                    f"Same-tab href recovery failed row={row_id} target_id='{open_target_id}': {nav_e}"
                                )
                except Exception as e:
                    _append_agent_log(f"Same-tab fallback click failed row={row_id} target_id='{open_target_id}': {e}")

            try:
                current_url_after_click = str(driver.current_url or "")
            except Exception:
                current_url_after_click = ""

            if (
                (not opened_client_profile)
                and search_url_before_click
                and current_url_after_click
                and (current_url_after_click != search_url_before_click)
                and ("healthsherpa.com" in current_url_after_click.lower())
            ):
                opened_client_profile = True

            # Fallback: if click navigated current tab, promote that client page into a new tab,
            # then restore search page in the original tab.
            if not opened_new_window:
                if (
                    search_url_before_click
                    and current_url_after_click
                    and (current_url_after_click != search_url_before_click)
                    and ("healthsherpa.com" in current_url_after_click.lower())
                ):
                    try:
                        original_handle = driver.current_window_handle
                        driver.switch_to.new_window('tab')
                        driver.get(current_url_after_click)
                        time.sleep(0.4)
                        handles_now = list(driver.window_handles)
                        new_handles = [h for h in handles_now if h not in windows_before]
                        if new_handles:
                            promoted_handle = new_handles[-1]
                            try:
                                if current_handle_before:
                                    driver.switch_to.window(current_handle_before)
                                    driver.get(search_url_before_click)
                                    wait_for_page_load(timeout_sec=6)
                            except Exception:
                                pass
                            driver.switch_to.window(promoted_handle)
                            opened_new_window = True
                            opened_client_profile = True
                            _append_agent_log(
                                f"Promoted same-tab client navigation to new tab row={row_id} url='{current_url_after_click[:180]}'"
                            )
                        else:
                            try:
                                if original_handle:
                                    driver.switch_to.window(original_handle)
                            except Exception:
                                pass
                    except Exception as e:
                        _append_agent_log(f"Failed to promote same-tab navigation to new tab row={row_id}: {e}")

            if (opened_by_url or bool(open_target_id)) and not opened_new_window:
                _append_agent_log(
                    f"Search/Add continuing with same-tab client flow row={row_id} ffm_id='{ffm_id}' (new-tab unavailable)"
                )

            if (opened_by_url or bool(open_target_id)) and (not opened_client_profile) and ffm_id:
                recovery = _direct_click_client_action_by_expected_id(ffm_id, add_button_text=add_button_text)
                if recovery.get("clicked"):
                    recovery_href = str(recovery.get("href", "") or "").strip()
                    moved = False
                    move_deadline = time.time() + 2.5
                    while time.time() < move_deadline:
                        try:
                            cur_now = str(driver.current_url or "")
                        except Exception:
                            cur_now = ""
                        if search_url_before_click and cur_now and (cur_now != search_url_before_click):
                            moved = True
                            break
                        time.sleep(0.15)

                    if (not moved) and recovery_href:
                        try:
                            driver.get(recovery_href)
                            moved = True
                        except Exception as nav_e:
                            _append_agent_log(f"Search/Add direct recovery href nav failed row={row_id} ffm_id='{ffm_id}': {nav_e}")

                    if moved:
                        opened_client_profile = True
                        _append_agent_log(
                            f"Search/Add direct recovery click opened profile row={row_id} ffm_id='{ffm_id}' button='{recovery.get('buttonText', '')}'"
                        )
                else:
                    _append_agent_log(
                        f"Search/Add direct recovery click did not open row={row_id} ffm_id='{ffm_id}' reason='{recovery.get('reason', '')}'"
                    )

            if (opened_by_url or bool(open_target_id)) and not opened_client_profile:
                _append_agent_log(
                    f"Search/Add profile open not yet confirmed row={row_id} ffm_id='{ffm_id}'; proceeding to sync wait"
                )

            try:
                wait_for_page_load(timeout_sec=8)
            except Exception:
                pass

            try:
                current_url_for_validation = str(driver.current_url or "")
            except Exception:
                current_url_for_validation = ""
            try:
                current_title_for_validation = str(driver.title or "")
            except Exception:
                current_title_for_validation = ""

            view_style_action = any(
                token in clicked_button_text
                for token in ["view in dash", "view dashboard", "view in dashboard", "view"]
            )
            title_changed_after_click = bool(
                search_title_before_click
                and current_title_for_validation
                and (current_title_for_validation != search_title_before_click)
            )
            navigated_to_client_context = bool(
                current_url_for_validation
                and search_url_before_click
                and (current_url_for_validation != search_url_before_click)
                and ("healthsherpa.com" in current_url_for_validation.lower())
                and ("/person_search" not in current_url_for_validation.lower())
            )

            if view_style_action and (not navigated_to_client_context) and action_url:
                try:
                    driver.get(action_url)
                    wait_for_page_load(timeout_sec=8)
                    time.sleep(0.4)
                    current_url_for_validation = str(driver.current_url or "")
                    current_title_for_validation = str(driver.title or "")
                    title_changed_after_click = bool(
                        search_title_before_click
                        and current_title_for_validation
                        and (current_title_for_validation != search_title_before_click)
                    )
                    navigated_to_client_context = bool(
                        current_url_for_validation
                        and search_url_before_click
                        and (current_url_for_validation != search_url_before_click)
                        and ("healthsherpa.com" in current_url_for_validation.lower())
                        and ("/person_search" not in current_url_for_validation.lower())
                    )
                    if navigated_to_client_context:
                        opened_client_profile = True
                    _append_agent_log(
                        f"Search/Add forced navigation via action_url row={row_id} ffm_id='{ffm_id}' success={navigated_to_client_context} url='{current_url_for_validation[:180]}'"
                    )
                except Exception as nav_e:
                    _append_agent_log(
                        f"Search/Add forced navigation via action_url failed row={row_id} ffm_id='{ffm_id}': {nav_e}"
                    )

            if view_style_action:
                _append_agent_log(
                    f"Search/Add view validation row={row_id} ffm_id='{ffm_id}' opened_client_profile={opened_client_profile} navigated_to_client_context={navigated_to_client_context} title_changed={title_changed_after_click} action_url_present={bool(action_url)} open_target_present={bool(open_target_id)} url='{current_url_for_validation[:180]}' title='{current_title_for_validation[:120]}'"
                )

            if (
                view_style_action
                and ffm_id
                and not (opened_client_profile or navigated_to_client_context or title_changed_after_click)
            ):
                _append_agent_log(
                    f"Search/Add human assist requested post-click row={row_id} ffm_id='{ffm_id}'"
                )
                assist_result = _human_assist_click_and_learn(ffm_id)
                if assist_result.get("clicked"):
                    try:
                        wait_for_page_load(timeout_sec=8)
                    except Exception:
                        pass
                    try:
                        current_url_for_validation = str(driver.current_url or "")
                    except Exception:
                        current_url_for_validation = ""
                    try:
                        current_title_for_validation = str(driver.title or "")
                    except Exception:
                        current_title_for_validation = ""

                    title_changed_after_click = bool(
                        search_title_before_click
                        and current_title_for_validation
                        and (current_title_for_validation != search_title_before_click)
                    )
                    navigated_to_client_context = bool(
                        current_url_for_validation
                        and search_url_before_click
                        and (current_url_for_validation != search_url_before_click)
                        and ("healthsherpa.com" in current_url_for_validation.lower())
                        and ("/person_search" not in current_url_for_validation.lower())
                    )
                    if navigated_to_client_context or title_changed_after_click:
                        opened_client_profile = True
                        _append_agent_log(
                            f"Search/Add post-click human assist confirmed navigation row={row_id} ffm_id='{ffm_id}' url='{current_url_for_validation[:180]}'"
                        )
                else:
                    _append_agent_log(
                        f"Search/Add post-click human assist canceled/failed row={row_id} ffm_id='{ffm_id}' reason='{assist_result.get('reason', '')}'"
                    )

            if view_style_action and (opened_client_profile or navigated_to_client_context or title_changed_after_click):
                found_sync = True
                _append_agent_log(
                    f"Search/Add view action validated by navigation row={row_id} ffm_id='{ffm_id}' url='{current_url_for_validation[:180]}'"
                )
            else:
                if view_style_action:
                    _append_agent_log(
                        f"Search/Add view action not confirmed by navigation row={row_id} ffm_id='{ffm_id}'; falling back to sync wait"
                    )
                found_sync = _wait_for_post_click_sync_text(
                    wait_text=wait_text,
                    wait_timeout=wait_timeout,
                    pre_click_visible=pre_click_sync_visible,
                )
            if not found_sync:
                reason = f"sync_text_not_found:{wait_text}"
                failed += 1
                failures.append({
                    "row": row_id,
                    "name": display_name,
                    "reason": reason,
                    "query": query,
                    "ffm_id": ffm_id,
                })
                _append_agent_log(
                    f"Search/Add sync wait failed row={row_id} ffm_id='{ffm_id}' wait_text='{wait_text}'"
                )
                if strict_match:
                    return {
                        "success": False,
                        "processed": processed,
                        "added": added,
                        "failed": failed,
                        "error": f"Strict stop: {reason} on row {row_id}",
                        "failures": failures,
                    }

            if close_after_sync:
                sync_or_timeout_seen = _wait_for_sync_or_timeout(wait_text=wait_text, timeout_sec=20.0)
                _append_agent_log(
                    f"Search/Add pre-reset hold complete row={row_id} sync_seen={sync_or_timeout_seen} timeout_sec=20"
                )

                if opened_new_window:
                    close_current_tab()
                else:
                    _append_agent_log("No new tab to close after sync; continuing with same-tab client flow")

                assisted_return_to_search = False
                if (not opened_new_window) and opened_client_profile:
                    try:
                        cur_before_reset = (driver.current_url or "").strip().lower()
                    except Exception:
                        cur_before_reset = ""

                    # Try automatic return first; only ask for human assist if still not in search context.
                    try:
                        if preferred_search_url:
                            driver.get(preferred_search_url)
                            wait_for_page_load(timeout_sec=8)
                            time.sleep(0.4)
                    except Exception as auto_nav_e:
                        _append_agent_log(f"Search/Add auto-reset pre-assist nav failed row={row_id}: {auto_nav_e}")

                    try:
                        cur_after_auto = (driver.current_url or "").strip().lower()
                    except Exception:
                        cur_after_auto = ""

                    if "/person_search" in cur_after_auto:
                        assisted_return_to_search = True
                        _append_agent_log(
                            f"Search/Add auto reset reached search context row={row_id}"
                        )
                    else:
                        _append_agent_log(f"Search/Add human assist requested before reset row={row_id}")
                        assist_ok = _ask_user_for_reset_navigation_assist()
                        if assist_ok:
                            try:
                                cur_after_assist = (driver.current_url or "").strip().lower()
                            except Exception:
                                cur_after_assist = ""
                            if "/person_search" in cur_after_assist:
                                assisted_return_to_search = True
                                _append_agent_log(
                                    f"Search/Add human assist confirmed search context before reset row={row_id}"
                                )
                            else:
                                _append_agent_log(
                                    f"Search/Add human assist completed but page not yet search context row={row_id} url='{cur_after_assist[:180]}'"
                                )
                        else:
                            _append_agent_log(f"Search/Add human assist canceled before reset row={row_id}")

                # Deterministic reset: always return to the exact search context before next row.
                try:
                    if assisted_return_to_search:
                        context_ready = _ensure_healthsherpa_marketplace_search_context(timeout_sec=8.0)
                        _append_agent_log(
                            f"Reset skipped auto-navigation; using assisted search context row={row_id} ready={context_ready}"
                        )
                    elif preferred_search_url:
                        driver.get(preferred_search_url)
                        wait_for_page_load(timeout_sec=8)
                        time.sleep(0.5)
                        context_ready = _ensure_healthsherpa_marketplace_search_context(timeout_sec=8.0)
                        _append_agent_log(
                            f"Reset to search context for next row={row_id} ready={context_ready} url='{preferred_search_url[:180]}'"
                        )
                    elif clients_list_url:
                        driver.get(clients_list_url)
                        wait_for_page_load(timeout_sec=8)
                        time.sleep(0.5)
                        context_ready = _ensure_healthsherpa_marketplace_search_context(timeout_sec=8.0)
                        _append_agent_log(
                            f"Reset via clients list for next row={row_id} ready={context_ready}"
                        )
                    else:
                        cur = (driver.current_url or "").strip()
                        if ("/person_search" not in cur) and ("/clients" not in cur):
                            _append_agent_log("Warning: no reset URL available; current page may not be search context")
                except Exception:
                    pass
        else:
            failed += 1
            reason = str(click_result.get("reason", "no_matching_add_button"))
            seen_ids = click_result.get("seenIds", []) if isinstance(click_result, dict) else []
            failures.append({
                "row": row_id,
                "name": display_name,
                "reason": reason,
                "query": query,
                "tokens": tokens,
                "ffm_id": ffm_id,
                "seen_ids": seen_ids,
            })
            _append_agent_log(
                f"Search/Add failed row={row_id} query='{query}' ffm_id='{ffm_id}': {reason} seen_ids={seen_ids}"
            )

            # Failure-path reset: refresh and return to known search context before next row.
            try:
                try:
                    driver.refresh()
                    wait_for_page_load(timeout_sec=8)
                    time.sleep(0.4)
                    _append_agent_log(f"Search/Add failure reset refreshed page row={row_id}")
                except Exception as refresh_e:
                    _append_agent_log(f"Search/Add failure reset refresh skipped row={row_id}: {refresh_e}")

                if preferred_search_url:
                    try:
                        driver.get(preferred_search_url)
                        wait_for_page_load(timeout_sec=8)
                        time.sleep(0.4)
                        context_ready = _ensure_healthsherpa_marketplace_search_context(timeout_sec=8.0)
                        _append_agent_log(
                            f"Search/Add failure reset to search context row={row_id} ready={context_ready}"
                        )
                    except Exception as nav_e:
                        _append_agent_log(
                            f"Search/Add failure reset navigation failed row={row_id}: {nav_e}"
                        )
                elif clients_list_url:
                    try:
                        driver.get(clients_list_url)
                        wait_for_page_load(timeout_sec=8)
                        time.sleep(0.4)
                        context_ready = _ensure_healthsherpa_marketplace_search_context(timeout_sec=8.0)
                        _append_agent_log(
                            f"Search/Add failure reset via clients list row={row_id} ready={context_ready}"
                        )
                    except Exception as nav_e:
                        _append_agent_log(
                            f"Search/Add failure reset via clients list failed row={row_id}: {nav_e}"
                        )
            except Exception:
                pass

            if strict_match:
                return {
                    "success": False,
                    "processed": processed,
                    "added": added,
                    "failed": failed,
                    "error": f"Strict stop: {reason} on row {row_id}",
                    "failures": failures,
                }

    return {
        "success": failed == 0,
        "processed": processed,
        "added": added,
        "failed": failed,
        "error": "" if failed == 0 else f"Completed with {failed} failures",
        "failures": failures,
    }
