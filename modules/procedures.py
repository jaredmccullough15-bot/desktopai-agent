import json
import os
import threading
import time
from dataclasses import dataclass
from typing import List, Optional, Tuple
import re

import mss
from PIL import Image

try:
    import numpy as np
except Exception:
    np = None

try:
    import imageio.v2 as imageio
except Exception:
    imageio = None

try:
    import pytesseract
except Exception:
    pytesseract = None

try:
    from pynput import mouse, keyboard
except Exception:
    mouse = None
    keyboard = None

import pyautogui

from modules.actions import (
    focus_window_by_title,
    selenium_get_current_url,
    selenium_get_active_input_info,
    selenium_set_input_value,
)
from modules.vision import get_active_window_info

PROCEDURES_DIR = os.path.join("data", "procedures")


def get_monitor_choices() -> List[Tuple[int, str]]:
    choices: List[Tuple[int, str]] = []
    with mss.mss() as sct:
        for idx, m in enumerate(sct.monitors[1:], start=1):
            label = f"{idx}: {m['width']}x{m['height']} @ ({m['left']},{m['top']})"
            choices.append((idx, label))
    if not choices:
        choices.append((1, "1: Primary"))
    return choices


def list_procedures() -> List[str]:
    if not os.path.isdir(PROCEDURES_DIR):
        return []
    names = []
    for name in os.listdir(PROCEDURES_DIR):
        manifest = os.path.join(PROCEDURES_DIR, name, "manifest.json")
        if os.path.isfile(manifest):
            names.append(name)
    return sorted(names)


def delete_procedure(name: str) -> bool:
    name = (name or "").strip()
    if not name:
        return False
    target = os.path.join(PROCEDURES_DIR, name)
    if not os.path.isdir(target):
        return False
    for root, dirs, files in os.walk(target, topdown=False):
        for f in files:
            try:
                os.remove(os.path.join(root, f))
            except Exception:
                pass
        for d in dirs:
            try:
                os.rmdir(os.path.join(root, d))
            except Exception:
                pass
    try:
        os.rmdir(target)
    except Exception:
        pass
    return True


@dataclass
class ProcedureInfo:
    name: str
    monitor_index: int
    fps: int
    video_path: Optional[str]
    frames_dir: str
    events: list
    created_at: str


def _load_procedure(name: str) -> Optional[ProcedureInfo]:
    name = (name or "").strip()
    if not name:
        return None
    manifest = os.path.join(PROCEDURES_DIR, name, "manifest.json")
    if not os.path.isfile(manifest):
        return None
    try:
        with open(manifest, "r", encoding="utf-8") as f:
            data = json.load(f)
        return ProcedureInfo(
            name=data.get("name", name),
            monitor_index=int(data.get("monitor_index", 1)),
            fps=int(data.get("fps", 2)),
            video_path=data.get("video_path"),
            frames_dir=data.get("frames_dir", ""),
            events=data.get("events", []),
            created_at=data.get("created_at", ""),
        )
    except Exception:
        return None


class ProcedureRecorder:
    def __init__(self, name: str, monitor_index: int, fps: int = 2) -> None:
        self.name = name
        self.monitor_index = monitor_index
        self.fps = max(1, int(fps))
        self._events = []
        self._recording = False
        self._start_time = 0.0
        self._frame_thread: Optional[threading.Thread] = None
        self._observe_thread: Optional[threading.Thread] = None
        self._mouse_listener = None
        self._keyboard_listener = None
        self._writer = None
        self._frame_count = 0
        self._frame_buffer = []
        self._monitor_rect = None
        self._last_error: Optional[str] = None

        self.base_dir = os.path.join(PROCEDURES_DIR, self.name)
        self.frames_dir = os.path.join(self.base_dir, "frames")
        self.video_path = os.path.join(self.base_dir, f"{self.name}.mp4")

    def start(self) -> None:
        if self._recording:
            return
        os.makedirs(self.frames_dir, exist_ok=True)
        self._events = []
        self._last_error = None
        self._recording = True
        self._start_time = time.time()

        self._writer = None
        self._frame_count = 0
        self._frame_buffer = []

        self._monitor_rect = self._get_monitor_rect(self.monitor_index)

        self._frame_thread = threading.Thread(target=self._record_frames, daemon=True)
        self._frame_thread.start()

        # Fail fast if frame capture thread dies immediately (common when monitor/screen capture fails)
        time.sleep(0.25)
        if self._frame_thread is not None and not self._frame_thread.is_alive():
            self._recording = False
            raise RuntimeError(self._last_error or "Frame recorder failed to start.")

        self._observe_thread = threading.Thread(target=self._observe_active_window, daemon=True)
        self._observe_thread.start()

        if mouse is not None:
            self._mouse_listener = mouse.Listener(on_click=self._on_click, on_scroll=self._on_scroll)
            self._mouse_listener.start()
        if keyboard is not None:
            self._keyboard_listener = keyboard.Listener(on_press=self._on_press)
            self._keyboard_listener.start()

    def stop(self) -> ProcedureInfo:
        self._recording = False
        if self._mouse_listener:
            try:
                self._mouse_listener.stop()
            except Exception:
                pass
        if self._keyboard_listener:
            try:
                self._keyboard_listener.stop()
            except Exception:
                pass
        if self._frame_thread:
            self._frame_thread.join(timeout=2.0)
        if self._observe_thread:
            self._observe_thread.join(timeout=2.0)
        if self._writer is not None:
            try:
                self._writer.close()
            except Exception:
                pass

        if self._frame_count < 2 and os.path.isfile(self.video_path):
            try:
                os.remove(self.video_path)
            except Exception:
                pass

        optimized_events = _intelligently_optimize_recorded_events(self.name, self._events)

        info = ProcedureInfo(
            name=self.name,
            monitor_index=self.monitor_index,
            fps=self.fps,
            video_path=self.video_path if self._writer is not None and self._frame_count >= 2 else None,
            frames_dir=self.frames_dir,
            events=optimized_events,
            created_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        )
        self._write_manifest(info)
        return info

    def add_checkpoint(self, note: str) -> None:
        if not self._recording:
            return
        note = (note or "").strip()
        if not note:
            return
        self._events.append({
            "t": round(self._elapsed(), 3),
            "type": "checkpoint",
            "note": note,
        })

    def _write_manifest(self, info: ProcedureInfo) -> None:
        os.makedirs(self.base_dir, exist_ok=True)
        manifest = {
            "name": info.name,
            "monitor_index": info.monitor_index,
            "fps": info.fps,
            "video_path": info.video_path,
            "frames_dir": info.frames_dir,
            "events": info.events,
            "created_at": info.created_at,
        }
        with open(os.path.join(self.base_dir, "manifest.json"), "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)

    def _elapsed(self) -> float:
        return time.time() - self._start_time

    def _on_click(self, x, y, button, pressed):
        if not self._recording or not pressed:
            return
        hint_text = self._capture_hint_text(int(x), int(y))
        self._events.append({
            "t": round(self._elapsed(), 3),
            "type": "click",
            "x": int(x),
            "y": int(y),
            "button": str(button).split(".")[-1],
            "hint_text": hint_text,
        })

    def _on_press(self, key):
        if not self._recording:
            return
        try:
            if hasattr(key, "char") and key.char:
                k = key.char
            else:
                k = str(key).split(".")[-1]
        except Exception:
            k = ""
        if not k:
            return
        self._events.append({
            "t": round(self._elapsed(), 3),
            "type": "key",
            "key": k,
        })

    def _on_scroll(self, x, y, dx, dy):
        if not self._recording:
            return
        self._events.append({
            "t": round(self._elapsed(), 3),
            "type": "scroll",
            "x": int(x),
            "y": int(y),
            "dx": int(dx),
            "dy": int(dy),
        })

    def _record_frames(self) -> None:
        frame_idx = 0
        try:
            with mss.mss() as sct:
                monitors = sct.monitors
                if self.monitor_index < 1 or self.monitor_index >= len(monitors):
                    monitor = monitors[1]
                else:
                    monitor = monitors[self.monitor_index]

                interval = 1.0 / float(self.fps)
                while self._recording:
                    sct_img = sct.grab(monitor)
                    img = Image.frombytes("RGB", sct_img.size, sct_img.bgra, "raw", "BGRX")
                    frame_path = os.path.join(self.frames_dir, f"frame_{frame_idx:06d}.png")
                    try:
                        img.save(frame_path)
                    except Exception:
                        pass

                    if np is not None and imageio is not None:
                        try:
                            frame_array = np.array(img)
                            if self._writer is None:
                                self._frame_buffer.append(frame_array)
                                if len(self._frame_buffer) >= 2:
                                    self._writer = imageio.get_writer(self.video_path, fps=self.fps)
                                    for buffered in self._frame_buffer:
                                        self._writer.append_data(buffered)
                                    self._frame_buffer = []
                            else:
                                self._writer.append_data(frame_array)
                        except Exception:
                            pass

                    frame_idx += 1
                    self._frame_count += 1
                    time.sleep(interval)
        except Exception as e:
            self._last_error = f"Frame capture failed: {e}"
            self._recording = False

    def _observe_active_window(self) -> None:
        last_title = ""
        last_url = ""
        last_field_key = ""
        last_field_value = ""
        stable_count = 0
        last_recorded = {}
        while self._recording:
            info = get_active_window_info()
            title = (info.get("title") or "").strip()
            if title and title != last_title:
                self._events.append({
                    "t": round(self._elapsed(), 3),
                    "type": "focus_window",
                    "title": title,
                    "app_hint": self._infer_app_hint(title),
                })
                last_title = title

            if title and "chrome" in title.lower():
                url = selenium_get_current_url()
                if url and url != last_url:
                    self._events.append({
                        "t": round(self._elapsed(), 3),
                        "type": "open_url",
                        "url": url,
                    })
                    last_url = url

                field = selenium_get_active_input_info()
                if field:
                    field_id = str(field.get("id") or "").strip()
                    field_name = str(field.get("name") or "").strip()
                    field_label = str(field.get("label") or "").strip()
                    field_value = str(field.get("value") or "")
                    field_key = field_id or field_name or field_label

                    if field_key:
                        if field_key == last_field_key and field_value == last_field_value:
                            stable_count += 1
                        else:
                            stable_count = 1
                            last_field_key = field_key
                            last_field_value = field_value

                        if stable_count >= 2:
                            last_value = last_recorded.get(field_key)
                            if field_value != last_value:
                                self._events.append({
                                    "t": round(self._elapsed(), 3),
                                    "type": "set_field",
                                    "label": field_label,
                                    "name": field_name,
                                    "id": field_id,
                                    "value": field_value,
                                })
                                last_recorded[field_key] = field_value
                                stable_count = 0

            time.sleep(0.5)

    def _infer_app_hint(self, title: str) -> str:
        lowered = (title or "").lower()
        if "chrome" in lowered:
            return "chrome"
        if "edge" in lowered:
            return "edge"
        if "firefox" in lowered:
            return "firefox"
        if "brave" in lowered:
            return "brave"
        if "opera" in lowered:
            return "opera"
        return (title or "")[:60]

    def _get_monitor_rect(self, index: int) -> dict:
        with mss.mss() as sct:
            monitors = sct.monitors
            if index < 1 or index >= len(monitors):
                return monitors[1]
            return monitors[index]

    def _capture_hint_text(self, x: int, y: int) -> str:
        if pytesseract is None:
            return ""
        rect = self._monitor_rect
        if not rect:
            return ""
        left = max(rect["left"], x - 150)
        top = max(rect["top"], y - 100)
        right = min(rect["left"] + rect["width"], x + 150)
        bottom = min(rect["top"] + rect["height"], y + 100)
        if right <= left or bottom <= top:
            return ""
        try:
            if os.getenv("TESSERACT_CMD"):
                pytesseract.pytesseract.tesseract_cmd = os.getenv("TESSERACT_CMD")
            with mss.mss() as sct:
                img = sct.grab({"left": left, "top": top, "width": right - left, "height": bottom - top})
            pil = Image.frombytes("RGB", img.size, img.bgra, "raw", "BGRX")
            text = pytesseract.image_to_string(pil, config="--psm 6")
            text = " ".join(text.split())
            return text[:80]
        except Exception:
            return ""


def _append_agent_log(message: str):
    """Append a timestamped line to data/agent.log for better traceability."""
    try:
        import os, datetime
        log_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'agent.log')
        ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"[{ts}] Procedure: {message}\n")
    except Exception:
        # Non-fatal: logging should not break procedures
        pass


def _is_healthsherpa_clients_url(url: str) -> bool:
    u = (url or "").strip().lower()
    return ("healthsherpa.com" in u) and ("/clients" in u)


def _build_smart_sherpa_events(events: list) -> Optional[list]:
    if not events:
        return None

    first_focus = None
    first_clients_url = None

    for ev in sorted(events, key=lambda e: float(e.get("t", 0) or 0)):
        et = str(ev.get("type", "")).strip()
        if first_focus is None and et == "focus_window":
            first_focus = {
                "t": 0.0,
                "type": "focus_window",
                "title": str(ev.get("title", "")).strip(),
                "app_hint": str(ev.get("app_hint", "chrome")).strip() or "chrome",
            }

        if et == "open_url":
            url = str(ev.get("url", "")).strip()
            if _is_healthsherpa_clients_url(url):
                first_clients_url = url
                break

    if not first_clients_url:
        return None

    smart_events = []
    if first_focus:
        smart_events.append(first_focus)

    smart_events.extend([
        {
            "t": 1.0,
            "type": "open_url",
            "url": first_clients_url,
        },
        {
            "t": 3.0,
            "type": "wait_for_page_load",
            "timeout_sec": 10,
        },
        {
            "t": 14.0,
            "type": "smart_process_all_clients",
            "wait_text": "Sync Complete",
            "wait_timeout": 20.0,
            "max_clients": 10000,
            "close_after_sync": True,
        },
    ])

    return smart_events


def _intelligently_optimize_recorded_events(name: str, events: list) -> list:
    if not events:
        return []

    sorted_events = sorted(events, key=lambda e: float(e.get("t", 0) or 0))
    # Auto-compile any recording that matches the clients-list processing pattern
    # into robust smart processing flow (not tied to procedure name).
    has_clients_navigation = any(
        str(ev.get("type", "")).strip() == "open_url" and _is_healthsherpa_clients_url(str(ev.get("url", "")).strip())
        for ev in sorted_events
    )

    if has_clients_navigation:
        smart = _build_smart_sherpa_events(sorted_events)
        if smart:
            _append_agent_log(
                f"Intelligent compile applied for procedure '{name}': detected clients-list flow and converted {len(sorted_events)} events to {len(smart)} smart events"
            )
            return smart

    # Generic cleanup for all other procedures: remove rapid duplicate focus events.
    cleaned = []
    last_focus_key = None
    last_focus_t = -999.0
    for ev in sorted_events:
        et = str(ev.get("type", "")).strip()
        if et == "focus_window":
            focus_key = (
                str(ev.get("title", "")).strip().lower(),
                str(ev.get("app_hint", "")).strip().lower(),
            )
            t = float(ev.get("t", 0) or 0)
            if focus_key == last_focus_key and (t - last_focus_t) < 0.75:
                continue
            last_focus_key = focus_key
            last_focus_t = t
        cleaned.append(ev)

    return cleaned


def _parse_required_flag(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return True
    text = str(value).strip().lower()
    if text in {"", "1", "true", "yes", "y", "required"}:
        return True
    if text in {"0", "false", "no", "n", "optional"}:
        return False
    return True


def _load_verification_checks_from_excel(mapping_excel_path: str, mapping_sheet: str = "") -> list:
    path = (mapping_excel_path or "").strip()
    if not path:
        return []

    resolved = path
    if not os.path.isabs(resolved):
        resolved = os.path.join(os.getcwd(), resolved)

    if not os.path.isfile(resolved):
        _append_agent_log(f"Verification mapping Excel not found: {resolved}")
        return []

    try:
        from openpyxl import load_workbook
    except Exception as e:
        _append_agent_log(f"openpyxl unavailable for verification mapping: {e}")
        return []

    try:
        wb = load_workbook(resolved, data_only=True)
        if mapping_sheet and mapping_sheet in wb.sheetnames:
            ws = wb[mapping_sheet]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header = [str(c or "").strip().lower() for c in rows[0]]
        idx = {name: i for i, name in enumerate(header) if name}

        def _pick(row_vals, names: list[str]) -> str:
            for name in names:
                i = idx.get(name)
                if i is None or i >= len(row_vals):
                    continue
                value = row_vals[i]
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    return text
            return ""

        checks = []
        for r in rows[1:]:
            marketplace_label = _pick(r, [
                "marketplace_label",
                "healthcaregov_label",
                "healthcare_label",
                "hcgov_label",
            ])
            profile_label = _pick(r, [
                "profile_label",
                "infusionsoft_label",
                "keap_label",
            ])
            marketplace_id = _pick(r, ["marketplace_id", "healthcaregov_id", "healthcare_id", "hcgov_id"])
            profile_id = _pick(r, ["profile_id", "infusionsoft_id", "keap_id"])
            marketplace_name = _pick(r, ["marketplace_name", "healthcaregov_name", "healthcare_name", "hcgov_name"])
            profile_name = _pick(r, ["profile_name", "infusionsoft_name", "keap_name"])

            has_marketplace_selector = bool(marketplace_label or marketplace_id or marketplace_name)
            has_profile_selector = bool(profile_label or profile_id or profile_name)
            if not has_marketplace_selector or not has_profile_selector:
                continue

            mode = _pick(r, ["mode", "compare_mode", "type"]).lower()
            required_raw = _pick(r, ["required", "is_required", "mandatory"]) 

            check = {
                "required": _parse_required_flag(required_raw),
            }
            if marketplace_label:
                check["marketplace_label"] = marketplace_label
            if profile_label:
                check["profile_label"] = profile_label
            if marketplace_id:
                check["marketplace_id"] = marketplace_id
            if profile_id:
                check["profile_id"] = profile_id
            if marketplace_name:
                check["marketplace_name"] = marketplace_name
            if profile_name:
                check["profile_name"] = profile_name
            if mode in {"text", "email", "phone", "date"}:
                check["mode"] = mode
            checks.append(check)

        _append_agent_log(
            f"Loaded {len(checks)} verification checks from Excel: {resolved} (sheet={ws.title})"
        )
        return checks
    except Exception as e:
        _append_agent_log(f"Failed loading verification checks from Excel '{resolved}': {e}")
        return []

def run_procedure(name: str, speed: float = 1.0, checkpoint_handler=None, runtime_overrides: Optional[dict] = None) -> bool:
    info = _load_procedure(name)
    if info is None:
        return False
    events = sorted(info.events, key=lambda e: e.get("t", 0))
    if not events:
        return False

    actionable_types = {
        "open_url",
        "set_field",
        "key",
        "click",
        "scroll",
        "reset_view_button_counter",
        "click_next_view_button",
        "wait_for_page_load",
        "verify_marketplace_profile_match",
        "smart_process_all_clients",
        "smart_search_and_add_clients",
    }
    has_actionable = any(
        str((ev or {}).get("type", "")).strip().lower() in actionable_types
        for ev in events
        if isinstance(ev, dict)
    )
    if not has_actionable:
        msg = (
            f"Procedure '{name}' has no actionable events (only focus/checkpoints/unknown). "
            "Re-record it or choose a different procedure."
        )
        print(f"[Procedure] {msg}")
        _append_agent_log(msg)
        return False

    pyautogui.FAILSAFE = True

    # Determine monitor rect once
    with mss.mss() as sct:
        monitors = sct.monitors
        if info.monitor_index < 1 or info.monitor_index >= len(monitors):
            monitor_rect = monitors[1]
        else:
            monitor_rect = monitors[info.monitor_index]

    # Ensure Selenium attaches to Chrome debug port
    os.environ["CHROME_DEBUG_PORT"] = os.getenv("CHROME_DEBUG_PORT", "9222") or "9222"
    os.environ["CURRENT_PROCEDURE_NAME"] = str(name or "").strip()

    # Track one-time Healthcare.gov application-link fallback click
    healthcare_app_link_clicked = False

    # Replay-time key combo state (for Ctrl/Shift/Alt + key hotkeys)
    pending_modifiers = set()
    pending_mod_ts = 0.0

    healthcare_page_signatures = {
        "login": {
            "title": ["log in", "sign in"],
            "url": ["/marketplace/auth/login", "/login"],
            "heading": ["log in", "sign in"],
        },
        "account_home": {
            "title": ["account homepage", "my account"],
            "url": ["/marketplace/my-account", "/my-account"],
            "heading": ["my account", "your applications"],
        },
        "application_status": {
            "title": ["application status"],
            "url": ["application", "status"],
            "heading": ["application status"],
        },
        "life_change": {
            "title": ["report a life change", "update application"],
            "url": ["life-change", "update"],
            "heading": ["report a life change", "update application"],
        },
        "privacy_notice": {
            "title": ["protecting your personal information"],
            "url": ["privacy", "protect"],
            "heading": ["protecting your personal information"],
        },
        "marketplace_application": {
            "title": ["your marketplace application", "application setup", "savings setup"],
            "url": ["marketplace", "application"],
            "heading": ["your marketplace application", "application setup", "savings setup"],
        },
    }

    def _normalize_modifier(raw_key: str) -> Optional[str]:
        key = (raw_key or "").strip().lower()
        if key in {"ctrl", "ctrl_l", "ctrl_r", "control", "control_l", "control_r"}:
            return "ctrl"
        if key in {"shift", "shift_l", "shift_r"}:
            return "shift"
        if key in {"alt", "alt_l", "alt_r", "alt_gr"}:
            return "alt"
        if key in {"cmd", "cmd_l", "cmd_r", "win", "windows", "super"}:
            return "win"
        return None

    def _normalize_key_for_pyautogui(raw_key: str) -> str:
        key = (raw_key or "").strip()
        control_char_map = {
            "\u0001": "a",  # Ctrl+A
            "\u0003": "c",  # Ctrl+C
            "\u0016": "v",  # Ctrl+V
            "\u0018": "x",  # Ctrl+X
            "\u001A": "z",  # Ctrl+Z
            "\u0019": "y",  # Ctrl+Y
        }
        if key in control_char_map:
            return control_char_map[key]

        lowered = key.lower()
        key_name_map = {
            "esc": "escape",
            "return": "enter",
            "page_up": "pageup",
            "page_down": "pagedown",
            "caps_lock": "capslock",
            "num_lock": "numlock",
            "scroll_lock": "scrolllock",
            "print_screen": "printscreen",
            "left": "left",
            "right": "right",
            "up": "up",
            "down": "down",
            "space": "space",
            "tab": "tab",
            "enter": "enter",
            "backspace": "backspace",
            "delete": "delete",
            "home": "home",
            "end": "end",
            "insert": "insert",
        }
        if lowered in key_name_map:
            return key_name_map[lowered]

        if len(key) == 1:
            return key.lower()
        return lowered

    def _get_browser_state() -> tuple[str, str, str]:
        try:
            from modules.actions import _get_selenium_driver
            driver = _get_selenium_driver()
            if driver is None:
                return "", "", ""
            title = (driver.title or "").strip()
            url = (driver.current_url or "").strip()
            heading = ""
            try:
                heading = (driver.execute_script(
                    """
                    const h = document.querySelector('h1, h2, [role=\"heading\"], main h1, main h2');
                    return h ? (h.textContent || '').trim() : '';
                    """
                ) or "").strip()
            except Exception:
                heading = ""
            return title, url, heading
        except Exception:
            return "", "", ""

    def _classify_healthcare_page(title: str, url: str, heading: str) -> set[str]:
        title_l = (title or "").lower()
        url_l = (url or "").lower()
        heading_l = (heading or "").lower()
        found = set()
        for page_name, sig in healthcare_page_signatures.items():
            title_hit = any(k in title_l for k in sig.get("title", []))
            url_hit = any(k in url_l for k in sig.get("url", []))
            heading_hit = any(k in heading_l for k in sig.get("heading", []))
            if title_hit or (url_hit and heading_hit) or heading_hit:
                found.add(page_name)
        return found

    def _is_healthcare_context() -> bool:
        title, url, heading = _get_browser_state()
        hay = f"{title} {url} {heading}".lower()
        return "healthcare.gov" in hay

    def _normalize_title_for_match(title: str) -> str:
        t = (title or "").strip().lower()
        t = t.replace(" - google chrome", "").strip()
        return t

    def _wait_for_expected_focus(expected_title: str, timeout_sec: float) -> bool:
        expected = _normalize_title_for_match(expected_title)
        if not expected:
            return True

        # Build expected Healthcare.gov page classes from recorded expected title
        expected_classes = _classify_healthcare_page(expected_title, "", expected_title)

        end_time = time.time() + max(0.5, timeout_sec)
        while time.time() < end_time:
            cur_title, cur_url, cur_heading = _get_browser_state()
            cur = _normalize_title_for_match(cur_title)
            if cur and (expected in cur or cur in expected):
                return True

            if expected_classes:
                cur_classes = _classify_healthcare_page(cur_title, cur_url, cur_heading)
                if expected_classes.intersection(cur_classes):
                    return True

            time.sleep(0.15)
        return False

    def _infer_click_targets_from_expected_focus(expected_title: str) -> list[str]:
        expected_clean = (expected_title or "").strip()
        expected = expected_clean.lower()
        expected = expected.replace(" - google chrome", "").strip()
        targets = []

        # Prefer exact-ish next-step labels when available.
        if expected:
            targets.append(expected.title())

        if "application status" in expected:
            targets.extend(["Application Status", "View status"])
        if "report a life change" in expected or "update application" in expected:
            targets.extend(["Report a life change", "Update application"])
        if "my account" in expected or "account homepage" in expected:
            targets.extend(["2026 Michigan application", "Michigan application", "Your applications"])
        if "protecting your personal information" in expected:
            targets.extend(["Continue", "Next"])
        if "review, sign, & submit" in expected:
            targets.extend(["Review, sign, & submit", "Continue"])

        # Login/account transitions
        if "log in" in expected or "sign in" in expected:
            targets.extend(["Log In", "Sign in", "Continue"])
        if "application setup" in expected or "savings setup" in expected:
            targets.extend(["Start application", "Continue", "Get started"])

        # Application step transitions (Healthcare.gov wizard-like pages)
        step_keywords = [
            "tell us about yourself",
            "home address",
            "mailing address",
            "contact information",
            "preferred language",
            "contact preferences",
            "application help",
            "who needs health coverage",
            "medicare enrollment",
            "marital status",
            "household tax returns",
            "parents & caretaker relatives",
            "household information",
            "race & ethnicity",
            "disabilities & help with activities",
            "medicaid or chip coverage ending",
            "recent household or income changes",
            "household income",
            "income for this month",
            "estimated income for this year",
            "current coverage & life changes",
            "current coverage",
            "job-based health coverage",
            "health reimbursement arrangements",
            "hra",
            "offers",
            "special enrollment period eligibility",
            "upcoming coverage changes",
            "life changes",
            "citizenship & immigration status",
            "personal & household information",
            "your marketplace application",
            "review your application",
            "read & agree to these statements",
            "sign & submit",
        ]
        if any(k in expected for k in step_keywords):
            targets.extend([
                "Continue",
                "Save & continue",
                "Next",
                "Confirm",
                "Submit",
            ])

        # Finalization-oriented transitions
        if "review your application" in expected:
            targets.extend(["Review, sign, & submit", "Continue"])
        if "read & agree" in expected or "sign & submit" in expected:
            targets.extend(["I agree", "Agree", "Sign & submit", "Submit"])

        seen = set()
        deduped = []
        for t in targets:
            cleaned = (t or "").strip()
            if not cleaned:
                continue
            key = cleaned.lower()
            if key not in seen:
                seen.add(key)
                deduped.append(cleaned)
        return deduped

    start = time.time()
    strict_transition = os.getenv("PROCEDURE_STRICT_TRANSITIONS", "1").strip().lower() not in {"0", "false", "no"}

    for idx, ev in enumerate(events):
        t = float(ev.get("t", 0)) / max(speed, 0.1)
        while time.time() - start < t:
            time.sleep(0.01)

        etype = str(ev.get("type", "")).strip().lower()

        if etype == "checkpoint":
            note = str(ev.get("note", "")).strip()
            if callable(checkpoint_handler):
                pause_start = time.time()
                should_continue = checkpoint_handler(note, monitor_rect)
                pause_end = time.time()
                start += max(0.0, pause_end - pause_start)
                if should_continue is False:
                    print(f"[Procedure] Checkpoint handler stopped procedure: {name}")
                    return False
            continue

        if etype == "focus_window":
            app_hint = str(ev.get("app_hint", "")).strip()
            title = str(ev.get("title", "")).strip()
            focus_window_by_title(app_hint or title)
            time.sleep(0.2)
            _append_agent_log(f"Focused window title='{title}' app_hint='{app_hint}'")
            continue

        if etype == "open_url":
            url = str(ev.get("url", "")).strip()
            if url:
                try:
                    from modules.actions import _get_selenium_driver
                    driver = _get_selenium_driver()
                    if driver is None:
                        print(f"[Procedure] Selenium driver not available for open_url: {url}")
                        _append_agent_log("Selenium driver not available for open_url")
                        continue
                    # If already on HealthSherpa, respect the current page and don't force navigation
                    try:
                        current = driver.current_url
                    except Exception:
                        current = ""
                    # Only skip navigation if we're already on the intended clients list path
                    if current and ("healthsherpa.com" in current) and ("healthsherpa.com" in url) and ("/clients" in current):
                        print(f"[Procedure] Already on HealthSherpa clients; skipping navigation to preset URL.")
                        _append_agent_log("On HealthSherpa clients; skipped navigation to preset URL")
                        time.sleep(0.5)
                    else:
                        driver.get(url)
                        time.sleep(1.0)
                        _append_agent_log(f"Navigated to URL: {url}")
                except Exception as e:
                    print(f"[Procedure] Selenium failed to open URL {url}: {e}")
                    _append_agent_log(f"Failed to open URL: {e}")
            continue

        if etype == "set_field":
            label = str(ev.get("label", "")).strip()
            value = str(ev.get("value", "")).strip()
            try:
                from modules.actions import selenium_set_input_value
                selenium_set_input_value(label, value)
            except Exception as e:
                print(f"[Procedure] Selenium failed to set field {label}: {e}")
            continue

        if etype == "key":
            raw_key = str(ev.get("key", "")).strip()
            if not raw_key:
                continue
            try:
                now = time.time()
                if pending_modifiers and (now - pending_mod_ts) > 1.0:
                    pending_modifiers.clear()

                mod = _normalize_modifier(raw_key)
                if mod:
                    pending_modifiers.add(mod)
                    pending_mod_ts = now
                    continue

                key_norm = _normalize_key_for_pyautogui(raw_key)

                if raw_key in {"\u0001", "\u0003", "\u0016", "\u0018", "\u001A", "\u0019"}:
                    pending_modifiers.add("ctrl")

                if pending_modifiers:
                    ordered_mods = [m for m in ("ctrl", "shift", "alt", "win") if m in pending_modifiers]
                    before_input = None
                    if "ctrl" in ordered_mods and key_norm == "v":
                        try:
                            before_input = selenium_get_active_input_info()
                        except Exception:
                            before_input = None

                    pyautogui.hotkey(*ordered_mods, key_norm)

                    if "ctrl" in ordered_mods and key_norm == "v":
                        # Retry paste once if focused field value appears unchanged
                        try:
                            time.sleep(0.08)
                            after_input = selenium_get_active_input_info()
                            if before_input and after_input:
                                before_key = (before_input.get("id") or before_input.get("name") or before_input.get("label") or "").strip().lower()
                                after_key = (after_input.get("id") or after_input.get("name") or after_input.get("label") or "").strip().lower()
                                before_val = str(before_input.get("value") or "")
                                after_val = str(after_input.get("value") or "")
                                if before_key and before_key == after_key and before_val == after_val:
                                    pyautogui.hotkey(*ordered_mods, key_norm)
                                    _append_agent_log("Retried Ctrl+V paste because field value did not change")
                        except Exception:
                            pass

                    pending_modifiers.clear()
                else:
                    if len(key_norm) == 1 and key_norm.isprintable():
                        pyautogui.write(key_norm)
                    else:
                        pyautogui.press(key_norm)
            except Exception as e:
                print(f"[Procedure] PyAutoGUI failed to press key {raw_key}: {e}")
            continue

        if etype == "click":
            x = int(ev.get("x", 0))
            y = int(ev.get("y", 0))
            button = str(ev.get("button", "left")).strip()
            hint_text = str(ev.get("hint_text", "")).strip()

            expected_focus = None
            expected_delta = None
            cur_t = float(ev.get("t", 0) or 0)
            for j in range(idx + 1, len(events)):
                next_ev = events[j]
                if str(next_ev.get("type", "")).strip() == "focus_window":
                    next_t = float(next_ev.get("t", 0) or 0)
                    delta = max(0.0, next_t - cur_t)
                    if delta <= 8.0:
                        expected_focus = str(next_ev.get("title", "")).strip()
                        expected_delta = delta
                    break

            # Smart pre-click strategy: use expected next page to infer robust text-based click targets.
            if _is_healthcare_context() and expected_focus:
                try:
                    from modules.actions import click_element_by_text
                    smart_clicked = False
                    for target_text in _infer_click_targets_from_expected_focus(expected_focus):
                        if click_element_by_text(target_text, element_type="any"):
                            _append_agent_log(f"Expected-focus smart click succeeded: '{target_text}' -> '{expected_focus}'")
                            smart_clicked = True
                            if strict_transition:
                                timeout = min(8.0, max(1.5, (expected_delta or 2.0) + 1.5))
                                if not _wait_for_expected_focus(expected_focus, timeout):
                                    msg = (
                                        f"Stopped: expected next page '{expected_focus}' after smart click '{target_text}' "
                                        f"but it did not appear within {timeout:.1f}s."
                                    )
                                    print(f"[Procedure] {msg}")
                                    _append_agent_log(msg)
                                    return False
                            break
                    if smart_clicked:
                        continue
                except Exception:
                    pass

            # Resilient fallback for Healthcare.gov app-link click on account pages
            if not healthcare_app_link_clicked:
                try:
                    from modules.actions import _get_selenium_driver, click_element_by_text
                    driver = _get_selenium_driver()
                    if driver is not None:
                        title = (driver.title or "").lower()
                        current_url = (driver.current_url or "").lower()
                        if "healthcare.gov" in (title + " " + current_url) and (
                            "account homepage" in title or "my account" in title
                        ):
                            for target_text in ("2026 Michigan application", "Michigan application", "2026 application"):
                                if click_element_by_text(target_text, element_type="any"):
                                    healthcare_app_link_clicked = True
                                    _append_agent_log(f"Healthcare.gov fallback clicked: {target_text}")
                                    time.sleep(0.5)
                                    break
                            if healthcare_app_link_clicked:
                                continue
                except Exception:
                    pass

            # If we have hint text and OCR available, try locating by text first
            if hint_text and pytesseract is not None and monitor_rect is not None:
                pos = _find_text_position(hint_text, monitor_rect)
                if pos is None:
                    for _ in range(6):
                        pyautogui.scroll(-500)
                        time.sleep(0.25)
                        pos = _find_text_position(hint_text, monitor_rect)
                        if pos is not None:
                            break
                if pos is not None:
                    try:
                        pyautogui.click(pos[0], pos[1])
                        continue
                    except Exception:
                        pass
            try:
                pyautogui.click(x, y, button=button)
            except Exception as e:
                print(f"[Procedure] PyAutoGUI failed to click at ({x},{y}): {e}")

            # Fail-fast validation for Healthcare.gov: after click, expected next focus page must appear.
            if strict_transition and _is_healthcare_context():
                if expected_focus:
                    timeout = min(8.0, max(1.5, (expected_delta or 2.0) + 1.5))
                    if not _wait_for_expected_focus(expected_focus, timeout):
                        msg = (
                            f"Stopped: expected next page '{expected_focus}' after click at ({x},{y}) "
                            f"but it did not appear within {timeout:.1f}s."
                        )
                        print(f"[Procedure] {msg}")
                        _append_agent_log(msg)
                        return False
            continue

        if etype == "scroll":
            dx = int(ev.get("dx", 0))
            dy = int(ev.get("dy", 0))
            x = ev.get("x")
            y = ev.get("y")
            # Move mouse to scroll position first (critical for proper scrolling)
            if x is not None and y is not None:
                try:
                    pyautogui.moveTo(int(x), int(y))
                except Exception:
                    pass
            try:
                if dy:
                    pyautogui.scroll(dy)
                if dx:
                    pyautogui.hscroll(dx)
            except Exception as e:
                print(f"[Procedure] PyAutoGUI failed to scroll: {e}")
            continue

        if etype == "reset_view_button_counter":
            try:
                from modules.actions import reset_view_button_counter
                reset_view_button_counter()
            except Exception as e:
                print(f"[Procedure] Failed to reset view button counter: {e}")
            continue

        if etype == "click_next_view_button":
            try:
                from modules.actions import click_next_view_button
                click_next_view_button()
            except Exception as e:
                print(f"[Procedure] Failed to click next view button: {e}")
            continue

        if etype == "wait_for_page_load":
            timeout_sec = float(ev.get("timeout_sec", 10.0))
            try:
                from modules.actions import wait_for_page_load
                ok = wait_for_page_load(timeout_sec=timeout_sec)
                _append_agent_log(f"wait_for_page_load(timeout={timeout_sec}) -> {bool(ok)}")
            except Exception as e:
                print(f"[Procedure] Failed to wait for page load: {e}")
                _append_agent_log(f"wait_for_page_load error: {e}")
            continue

        if etype == "verify_marketplace_profile_match":
            checks = ev.get("checks", [])
            strict = bool(ev.get("strict", True))
            marketplace_tokens = ev.get("marketplace_match_tokens", ["healthcare.gov", "marketplace"])
            profile_tokens = ev.get("profile_match_tokens", ["infusionsoft", "keap"])
            mapping_excel_path = str(ev.get("mapping_excel_path", "")).strip()
            mapping_sheet = str(ev.get("mapping_sheet", "")).strip()

            if (not checks) and mapping_excel_path:
                checks = _load_verification_checks_from_excel(mapping_excel_path, mapping_sheet)

            if not checks:
                msg = "Verification skipped: no checks provided (inline or Excel mapping)."
                print(f"[Procedure] {msg}")
                _append_agent_log(msg)
                if strict:
                    return False
                continue

            try:
                from modules.actions import verify_marketplace_profile_match
                result = verify_marketplace_profile_match(
                    checks=checks,
                    marketplace_match_tokens=marketplace_tokens,
                    profile_match_tokens=profile_tokens,
                )

                success = bool(result.get("success"))
                checked = int(result.get("checked", 0) or 0)
                mismatch_count = len(result.get("mismatches", []) or [])
                _append_agent_log(
                    f"verify_marketplace_profile_match success={success} checked={checked} mismatches={mismatch_count}"
                )

                if success:
                    print(f"[Procedure] Marketplace profile verification passed ({checked} checks).")
                else:
                    err = str(result.get("error", "")).strip()
                    mismatches = result.get("mismatches", []) or []
                    if err:
                        print(f"[Procedure] Verification failed: {err}")
                        _append_agent_log(f"Verification failed: {err}")
                    for mm in mismatches[:10]:
                        print(
                            "[Procedure] Mismatch: "
                            f"{mm.get('marketplace_label', '')}='{mm.get('marketplace_value', '')}' vs "
                            f"{mm.get('profile_label', '')}='{mm.get('profile_value', '')}' "
                            f"reason={mm.get('reason', 'value_mismatch')}"
                        )

                    if strict:
                        stop_msg = "Stopped: marketplace/profile verification mismatch detected."
                        print(f"[Procedure] {stop_msg}")
                        _append_agent_log(stop_msg)
                        return False
            except Exception as e:
                print(f"[Procedure] Verification step error: {e}")
                _append_agent_log(f"verify_marketplace_profile_match error: {e}")
                if strict:
                    return False
            continue

        if etype == "smart_process_all_clients":
            try:
                from modules.actions import smart_process_all_clients
            except Exception:
                smart_process_all_clients = None
            wait_text = str(ev.get("wait_text", "Sync Complete")).strip()
            wait_timeout = float(ev.get("wait_timeout", 20.0))
            max_clients = int(ev.get("max_clients", 10000))
            close_after_sync = bool(ev.get("close_after_sync", True))
            pagination_mode = str(ev.get("pagination_mode", "auto")).strip() or "auto"

            print("\n🤖 Smart Client Processing Mode Activated")
            print("   This will automatically:")
            print("   • Process all clients on current page")
            print("   • Detect when page is exhausted")
            print("   • Navigate to next page automatically")
            print("   • Stop when all pages are complete\n")
            _append_agent_log(
                f"smart_process_all_clients start wait_text='{wait_text}' timeout={wait_timeout} max={max_clients} close={close_after_sync} pagination_mode={pagination_mode}"
            )

            if callable(smart_process_all_clients):
                prev_proc_name = os.getenv("CURRENT_PROCEDURE_NAME", "")
                os.environ["CURRENT_PROCEDURE_NAME"] = name
                try:
                    result = smart_process_all_clients(
                        wait_text=wait_text,
                        wait_timeout=wait_timeout,
                        max_clients=max_clients,
                        close_after_sync=close_after_sync,
                        pagination_mode=pagination_mode,
                    )
                finally:
                    if prev_proc_name:
                        os.environ["CURRENT_PROCEDURE_NAME"] = prev_proc_name
                    else:
                        os.environ.pop("CURRENT_PROCEDURE_NAME", None)
                if result.get("success"):
                    print("\n✅ Smart processing complete!")
                    print(f"   Clients processed: {result.get('clients_processed', 0)}")
                    print(f"   Pages processed: {result.get('pages_processed', 0)}")
                    _append_agent_log(f"smart_process_all_clients complete clients={result.get('clients_processed', 0)} pages={result.get('pages_processed', 0)}")
                else:
                    print("\n⚠️ Smart processing ended with errors")
                    print(f"   Error: {result.get('error', 'Unknown error')}")
                    print(f"   Clients processed before error: {result.get('clients_processed', 0)}")
                    _append_agent_log(f"smart_process_all_clients error: {result.get('error', 'Unknown')} processed={result.get('clients_processed', 0)}")
            else:
                print("[Procedure] smart_process_all_clients not available.")
                _append_agent_log("smart_process_all_clients not available (import failed)")
            continue

        if etype == "smart_search_and_add_clients":
            mapping_excel_path = str(ev.get("mapping_excel_path", "")).strip()
            mapping_sheet = str(ev.get("mapping_sheet", "")).strip()
            strict_match = bool(ev.get("strict_match", True))
            max_clients = int(ev.get("max_clients", 500) or 500)
            add_button_text = str(ev.get("add_button_text", "Add client")).strip() or "Add client"
            search_wait_sec = float(ev.get("search_wait_sec", 1.0) or 1.0)
            wait_text = str(ev.get("wait_text", "Sync Complete")).strip() or "Sync Complete"
            wait_timeout = float(ev.get("wait_timeout", 20.0) or 20.0)
            close_after_sync = bool(ev.get("close_after_sync", True))
            search_context_url = str(ev.get("search_context_url", "")).strip()
            virtual_grid_mode = bool(ev.get("virtual_grid_mode", True))

            smart_overrides = {}
            if isinstance(runtime_overrides, dict):
                smart_overrides = runtime_overrides.get("smart_search_and_add_clients", {}) or {}
            if isinstance(smart_overrides, dict):
                mapping_excel_path = str(smart_overrides.get("mapping_excel_path", mapping_excel_path) or "").strip()
                mapping_sheet = str(smart_overrides.get("mapping_sheet", mapping_sheet) or "").strip()
                search_context_url = str(smart_overrides.get("search_context_url", search_context_url) or "").strip()
                virtual_grid_mode = bool(smart_overrides.get("virtual_grid_mode", virtual_grid_mode))

            try:
                from modules.actions import smart_search_and_add_clients
            except Exception:
                smart_search_and_add_clients = None

            if not mapping_excel_path:
                msg = "smart_search_and_add_clients skipped: mapping_excel_path is required"
                print(f"[Procedure] {msg}")
                _append_agent_log(msg)
                return False

            # Explicit Selenium preflight: require attached debug Chrome before smart click workflow
            try:
                from modules.actions import _get_selenium_driver
                driver = _get_selenium_driver()
                if driver is None:
                    msg = (
                        "smart_search_and_add_clients requires Selenium attached to Chrome debug port. "
                        "Start Chrome with --remote-debugging-port=9222 and try again."
                    )
                    print(f"[Procedure] {msg}")
                    _append_agent_log(msg)
                    return False
                try:
                    _ = driver.current_url
                except Exception:
                    msg = "smart_search_and_add_clients Selenium session is not ready."
                    print(f"[Procedure] {msg}")
                    _append_agent_log(msg)
                    return False
            except Exception as e:
                msg = f"smart_search_and_add_clients Selenium preflight error: {e}"
                print(f"[Procedure] {msg}")
                _append_agent_log(msg)
                return False

            if callable(smart_search_and_add_clients):
                print("\n🤖 Smart Search + Add Clients Mode Activated")
                print(f"   Excel: {mapping_excel_path}")
                if mapping_sheet:
                    print(f"   Sheet: {mapping_sheet}")
                result = smart_search_and_add_clients(
                    mapping_excel_path=mapping_excel_path,
                    mapping_sheet=mapping_sheet,
                    strict_match=strict_match,
                    max_clients=max_clients,
                    add_button_text=add_button_text,
                    search_wait_sec=search_wait_sec,
                    wait_text=wait_text,
                    wait_timeout=wait_timeout,
                    close_after_sync=close_after_sync,
                    search_context_url=search_context_url,
                    virtual_grid_mode=virtual_grid_mode,
                )
                success = bool(result.get("success"))
                processed = int(result.get("processed", 0) or 0)
                added = int(result.get("added", 0) or 0)
                failed = int(result.get("failed", 0) or 0)

                _append_agent_log(
                    f"smart_search_and_add_clients success={success} processed={processed} added={added} failed={failed}"
                )
                print(f"[Procedure] Search/Add summary: processed={processed}, added={added}, failed={failed}")

                if strict_match and not success:
                    err = str(result.get("error", "Search/Add failed in strict mode"))
                    print(f"[Procedure] Stopped: {err}")
                    return False
            else:
                print("[Procedure] smart_search_and_add_clients not available.")
                _append_agent_log("smart_search_and_add_clients not available (import failed)")
                return False
            continue

        _append_agent_log(f"Procedure '{name}' skipped unknown event type: '{etype}'")
        # Unknown event type; skip
        continue

    print(f"[Procedure] Finished running procedure: {name}")
    _append_agent_log(f"Finished procedure: {name}")
    return True


def run_procedure_loop(name: str, repeat_count: int = 0, delay_sec: float = 1.0, stop_event: Optional[threading.Event] = None, checkpoint_handler=None, runtime_overrides: Optional[dict] = None) -> bool:
    """Run a procedure repeatedly. repeat_count=0 means run until stop_event is set."""
    count = 0
    while True:
        if stop_event is not None and stop_event.is_set():
            return True
        ok = run_procedure(name, checkpoint_handler=checkpoint_handler, runtime_overrides=runtime_overrides)
        if not ok:
            return False
        count += 1
        if repeat_count > 0 and count >= repeat_count:
            return True
        time.sleep(max(0.1, delay_sec))


def _find_text_position(target_text: str, monitor_rect: dict) -> Optional[Tuple[int, int]]:
    if pytesseract is None:
        return None
    try:
        if os.getenv("TESSERACT_CMD"):
            pytesseract.pytesseract.tesseract_cmd = os.getenv("TESSERACT_CMD")
        with mss.mss() as sct:
            img = sct.grab(monitor_rect)
        pil = Image.frombytes("RGB", img.size, img.bgra, "raw", "BGRX")
        data = pytesseract.image_to_data(pil, output_type=pytesseract.Output.DICT, config="--psm 6")
        words = [w for w in re.split(r"\s+", target_text) if w]
        if not words:
            return None
        for i, text in enumerate(data.get("text", [])):
            if not text:
                continue
            if words[0].lower() in text.lower():
                x = data.get("left", [0])[i]
                y = data.get("top", [0])[i]
                w = data.get("width", [0])[i]
                h = data.get("height", [0])[i]
                return (monitor_rect["left"] + x + w // 2, monitor_rect["top"] + y + h // 2)
    except Exception:
        return None
    return None


def find_text_position(target_text: str, monitor_rect: dict) -> Optional[Tuple[int, int]]:
    """Public wrapper for OCR-based text lookup."""
    return _find_text_position(target_text, monitor_rect)
