from __future__ import annotations

import csv
import os
import re
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text or "").strip().lower())


def _split_list(text: str) -> List[str]:
    raw = [p.strip(" .'\"") for p in re.split(r"\s*(?:,| and | & )\s*", text or "", flags=re.IGNORECASE)]
    return [x for x in raw if x]


def _to_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _to_datetime(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m-%d-%Y",
        "%m-%d-%y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    return None


def _candidate_dirs() -> List[str]:
    dirs: List[str] = []
    user_profile = os.environ.get("USERPROFILE", "").strip()
    if user_profile:
        dirs.append(os.path.join(user_profile, "Downloads"))
    dirs.append(os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "exports"))
    return [d for d in dirs if d and os.path.isdir(d)]


def _find_latest_data_file() -> Optional[str]:
    exts = {".xlsx", ".xlsm", ".csv", ".zip"}
    found: List[Tuple[float, str]] = []
    for base in _candidate_dirs():
        try:
            for name in os.listdir(base):
                path = os.path.join(base, name)
                if not os.path.isfile(path):
                    continue
                ext = os.path.splitext(name)[1].lower()
                if ext in exts:
                    found.append((os.path.getmtime(path), path))
        except Exception:
            continue
    if not found:
        return None
    found.sort(key=lambda x: x[0], reverse=True)
    return found[0][1]


def _find_header_index(headers: List[str], requested: str) -> int:
    if not headers or not requested:
        return -1
    raw_requested = str(requested or "").strip()
    m_col = re.match(r"^(?:column\s+)?([A-Za-z]{1,3})$", raw_requested, re.IGNORECASE)
    if m_col:
        letters = m_col.group(1).upper()
        col_index = 0
        for ch in letters:
            col_index = (col_index * 26) + (ord(ch) - ord("A") + 1)
        idx = col_index - 1
        if 0 <= idx < len(headers):
            return idx

    requested_norm = _norm(requested)
    for idx, col in enumerate(headers):
        if _norm(col) == requested_norm:
            return idx
    for idx, col in enumerate(headers):
        if requested_norm and requested_norm in _norm(col):
            return idx
    return -1


def _parse_sort_specs(text: str) -> List[Dict[str, Any]]:
    match = re.search(r"\bsort\s+by\s+(.+?)(?:\b(?:save|then|and\s+save|and\s+filter|and\s+where)\b|$)", text, re.IGNORECASE)
    if not match:
        return []
    segment = match.group(1).strip(" .")
    parts = [p.strip(" .") for p in re.split(r"\s*,\s*|\s+then\s+|\s+and\s+", segment, flags=re.IGNORECASE) if p.strip()]
    specs: List[Dict[str, Any]] = []
    for part in parts:
        m = re.match(r"(.+?)\s+(ascending|asc|descending|desc)$", part, re.IGNORECASE)
        if m:
            column = m.group(1).strip(" .'\"")
            desc = m.group(2).strip().lower() in {"descending", "desc"}
        else:
            column = part.strip(" .'\"")
            desc = False
        if column:
            specs.append({"column": column, "desc": desc})
    return specs


def _parse_filter_specs(text: str) -> List[Dict[str, Any]]:
    specs: List[Dict[str, Any]] = []

    m_col_hint = re.search(r"\bcolumn\s+([A-Za-z]{1,3})\b", text, re.IGNORECASE)
    col_hint = f"column {m_col_hint.group(1).upper()}" if m_col_hint else ""

    for m in re.finditer(r"(?:where|filter|keep\s+only)\s+(.+?)\s+between\s+([^\s]+)\s+and\s+([^\s\.,;]+)", text, re.IGNORECASE):
        specs.append({"column": m.group(1).strip(" .'\""), "op": "between", "value": m.group(2).strip(), "value2": m.group(3).strip()})

    for m in re.finditer(r"(?:where|filter|keep\s+only)\s+(.+?)\s*(>=|<=|>|<)\s*([^\s\.,;]+)", text, re.IGNORECASE):
        specs.append({"column": m.group(1).strip(" .'\""), "op": m.group(2), "value": m.group(3).strip()})

    for m in re.finditer(r"(?:where|filter|keep\s+only)\s+(.+?)\s+(contains|starts with|ends with)\s+['\"]?([^'\"\n]+?)['\"]?(?:$|\s+and\s+|\s+then\s+)", text, re.IGNORECASE):
        op_raw = m.group(2).strip().lower()
        op = "contains" if op_raw == "contains" else "startswith" if op_raw == "starts with" else "endswith"
        specs.append({"column": m.group(1).strip(" .'\""), "op": op, "value": m.group(3).strip()})

    for m in re.finditer(r"(?:where|filter|keep\s+only)\s+(.+?)\s*(=|equals|is|!=|not equal to)\s*['\"]?([^'\"\n]+?)['\"]?(?:$|\s+and\s+|\s+then\s+)", text, re.IGNORECASE):
        op_raw = m.group(2).strip().lower()
        op = "neq" if op_raw in {"!=", "not equal to"} else "eq"
        specs.append({"column": m.group(1).strip(" .'\""), "op": op, "value": m.group(3).strip()})

    m_before = re.search(r"\bbefore\s+([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})", text, re.IGNORECASE)
    if m_before and col_hint:
        specs.append({"column": col_hint, "op": "<", "value": m_before.group(1).strip()})

    m_after = re.search(r"\bafter\s+([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})", text, re.IGNORECASE)
    if m_after and col_hint:
        specs.append({"column": col_hint, "op": ">", "value": m_after.group(1).strip()})

    return specs


def _parse_instruction(instruction: str) -> Dict[str, Any]:
    text = (instruction or "").strip()
    lowered = text.lower()

    m_sheet = re.search(r"\bsheet\s+['\"]?([^'\"]+)['\"]?", text, re.IGNORECASE)
    sheet_name = m_sheet.group(1).strip() if m_sheet else ""

    m_save = re.search(r"\bsave\s+(?:as|to)\s+['\"]?([^'\"\n]+?)['\"]?$", text, re.IGNORECASE)
    output_name = m_save.group(1).strip() if m_save else ""

    dedupe = bool(re.search(r"\b(remove|drop)\s+duplicates?\b", lowered))
    dedupe_cols: List[str] = []
    m_dedupe_by = re.search(r"\b(?:remove|drop)\s+duplicates?\s+by\s+(.+?)(?:$|\s+and\s+|\s+then\s+)", text, re.IGNORECASE)
    if m_dedupe_by:
        dedupe_cols = _split_list(m_dedupe_by.group(1))

    keep_columns: List[str] = []
    drop_columns: List[str] = []

    m_keep = re.search(r"\bkeep\s+columns?\s+(.+?)(?:$|\s+and\s+|\s+then\s+)", text, re.IGNORECASE)
    if m_keep:
        keep_columns = _split_list(m_keep.group(1))

    m_drop = re.search(r"\b(?:drop|remove|delete)\s+columns?\s+(.+?)(?:$|\s+and\s+|\s+then\s+)", text, re.IGNORECASE)
    if m_drop:
        drop_columns = _split_list(m_drop.group(1))

    rename_map: Dict[str, str] = {}
    for m in re.finditer(r"\brename\s+column\s+['\"]?([^'\"]+?)['\"]?\s+to\s+['\"]?([^'\"]+?)['\"]?(?:$|\s+and\s+|\s+then\s+)", text, re.IGNORECASE):
        old = m.group(1).strip()
        new = m.group(2).strip()
        if old and new:
            rename_map[old] = new

    fill_blanks: List[Dict[str, str]] = []
    for m in re.finditer(r"\bfill\s+blanks?\s+in\s+(.+?)\s+with\s+['\"]?([^'\"\n]+?)['\"]?(?:$|\s+and\s+|\s+then\s+)", text, re.IGNORECASE):
        fill_blanks.append({"column": m.group(1).strip(" .'\""), "value": m.group(2).strip()})

    return {
        "sheet_name": sheet_name,
        "output_name": output_name,
        "sort_specs": _parse_sort_specs(text),
        "filter_specs": _parse_filter_specs(text),
        "dedupe": dedupe,
        "dedupe_cols": dedupe_cols,
        "keep_columns": keep_columns,
        "drop_columns": drop_columns,
        "rename_map": rename_map,
        "fill_blanks": fill_blanks,
    }


def _coerce_sort_value(value: Any):
    if value is None:
        return (3, "")
    text = str(value).strip()
    if not text:
        return (3, "")
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(text, fmt)
            return (0, dt)
        except Exception:
            pass
    try:
        num = float(text.replace(",", ""))
        return (1, num)
    except Exception:
        pass
    return (2, text.lower())


def _coerce_filter_compare(value: Any):
    dt = _to_datetime(value)
    if dt is not None:
        return ("date", dt)
    num = _to_number(value)
    if num is not None:
        return ("number", num)
    return ("text", str(value or "").strip().lower())


def _match_filter(cell_value: Any, spec: Dict[str, Any]) -> bool:
    op = str(spec.get("op") or "").lower()
    target = spec.get("value")
    target2 = spec.get("value2")

    cell_text = str(cell_value or "").strip()
    cell_lower = cell_text.lower()
    target_text = str(target or "").strip()
    target_lower = target_text.lower()

    if op == "eq":
        return cell_lower == target_lower
    if op == "neq":
        return cell_lower != target_lower
    if op == "contains":
        return target_lower in cell_lower
    if op == "startswith":
        return cell_lower.startswith(target_lower)
    if op == "endswith":
        return cell_lower.endswith(target_lower)

    if op in {">", ">=", "<", "<=", "between"}:
        c_kind, c_val = _coerce_filter_compare(cell_value)
        t_kind, t_val = _coerce_filter_compare(target_text)
        if c_kind != t_kind:
            return False
        if op == ">":
            return c_val > t_val
        if op == ">=":
            return c_val >= t_val
        if op == "<":
            return c_val < t_val
        if op == "<=":
            return c_val <= t_val
        if op == "between":
            if target2 is None:
                return False
            t2_kind, t2_val = _coerce_filter_compare(target2)
            if t2_kind != c_kind:
                return False
            lo, hi = (t_val, t2_val) if t_val <= t2_val else (t2_val, t_val)
            return lo <= c_val <= hi

    return True


def _build_sort_key(row: List[Any], idx: int):
    return _coerce_sort_value(row[idx] if idx < len(row) else "")


def _reindex_rows(rows: List[List[Any]], selected_indexes: List[int]) -> List[List[Any]]:
    return [[row[i] if i < len(row) else "" for i in selected_indexes] for row in rows]


def _apply_ops(headers: List[str], rows: List[List[Any]], parsed: Dict[str, Any]) -> Tuple[List[List[Any]], List[str]]:
    notes: List[str] = []
    out_headers = list(headers)
    out_rows = [list(r) for r in rows]

    rename_map = parsed.get("rename_map") if isinstance(parsed.get("rename_map"), dict) else {}
    for old_name, new_name in rename_map.items():
        idx = _find_header_index(out_headers, old_name)
        if idx >= 0:
            original = out_headers[idx]
            out_headers[idx] = str(new_name)
            notes.append(f"Renamed column '{original}' to '{new_name}'.")
        else:
            notes.append(f"Rename skipped; column not found: {old_name}")

    keep_columns = parsed.get("keep_columns") if isinstance(parsed.get("keep_columns"), list) else []
    if keep_columns:
        keep_indexes = []
        for col in keep_columns:
            idx = _find_header_index(out_headers, str(col))
            if idx >= 0 and idx not in keep_indexes:
                keep_indexes.append(idx)
        if keep_indexes:
            out_headers = [out_headers[i] for i in keep_indexes]
            out_rows = _reindex_rows(out_rows, keep_indexes)
            notes.append("Kept columns: " + ", ".join(out_headers))
        else:
            notes.append("Keep-columns skipped; no matching columns found.")

    drop_columns = parsed.get("drop_columns") if isinstance(parsed.get("drop_columns"), list) else []
    if drop_columns:
        drop_indexes = set()
        for col in drop_columns:
            idx = _find_header_index(out_headers, str(col))
            if idx >= 0:
                drop_indexes.add(idx)
        if drop_indexes:
            selected_indexes = [i for i in range(len(out_headers)) if i not in drop_indexes]
            dropped = [out_headers[i] for i in sorted(drop_indexes)]
            out_headers = [out_headers[i] for i in selected_indexes]
            out_rows = _reindex_rows(out_rows, selected_indexes)
            notes.append("Dropped columns: " + ", ".join(dropped))
        else:
            notes.append("Drop-columns skipped; no matching columns found.")

    fill_blanks = parsed.get("fill_blanks") if isinstance(parsed.get("fill_blanks"), list) else []
    for rule in fill_blanks:
        col_name = str(rule.get("column") or "").strip()
        fill_value = str(rule.get("value") or "")
        idx = _find_header_index(out_headers, col_name)
        if idx < 0:
            notes.append(f"Fill-blanks skipped; column not found: {col_name}")
            continue
        count = 0
        for row in out_rows:
            current = row[idx] if idx < len(row) else ""
            if str(current or "").strip() == "":
                while len(row) <= idx:
                    row.append("")
                row[idx] = fill_value
                count += 1
        notes.append(f"Filled {count} blank values in '{out_headers[idx]}'.")

    filter_specs = parsed.get("filter_specs") if isinstance(parsed.get("filter_specs"), list) else []
    for spec in filter_specs:
        idx = _find_header_index(out_headers, str(spec.get("column") or ""))
        if idx < 0:
            notes.append(f"Filter skipped; column not found: {spec.get('column')}")
            continue
        before = len(out_rows)
        out_rows = [r for r in out_rows if _match_filter(r[idx] if idx < len(r) else "", spec)]
        after = len(out_rows)
        notes.append(f"Filtered '{out_headers[idx]}' with {spec.get('op')} ({before} -> {after}).")

    if parsed.get("dedupe"):
        dedupe_cols = parsed.get("dedupe_cols") if isinstance(parsed.get("dedupe_cols"), list) else []
        dedupe_indexes = []
        for col in dedupe_cols:
            idx = _find_header_index(out_headers, str(col))
            if idx >= 0:
                dedupe_indexes.append(idx)
        seen = set()
        deduped = []
        for row in out_rows:
            if dedupe_indexes:
                key = tuple(str(row[i] if i < len(row) else "").strip().lower() for i in dedupe_indexes)
            else:
                key = tuple(str(v or "").strip().lower() for v in row)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(row)
        removed = len(out_rows) - len(deduped)
        out_rows = deduped
        if dedupe_indexes:
            cols = ", ".join(out_headers[i] for i in dedupe_indexes)
            notes.append(f"Removed {removed} duplicate rows by columns: {cols}.")
        else:
            notes.append(f"Removed {removed} duplicate rows.")

    sort_specs = parsed.get("sort_specs") if isinstance(parsed.get("sort_specs"), list) else []
    for spec in reversed(sort_specs):
        idx = _find_header_index(out_headers, str(spec.get("column") or ""))
        if idx < 0:
            notes.append(f"Sort skipped; column not found: {spec.get('column')}")
            continue
        desc = bool(spec.get("desc"))
        out_rows = sorted(out_rows, key=lambda r: _build_sort_key(r, idx), reverse=desc)
        direction = "descending" if desc else "ascending"
        notes.append(f"Sorted by '{out_headers[idx]}' ({direction}).")

    return out_headers, out_rows, notes


def _read_csv(path: str) -> Tuple[List[str], List[List[Any]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return [], []
    headers = [str(x or "").strip() for x in rows[0]]
    body = [list(r) for r in rows[1:]]
    return headers, body


def _write_csv(path: str, headers: List[str], rows: List[List[Any]]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


def _resolve_output_path(input_path: str, output_name: str, preferred_ext: str) -> str:
    base_dir = os.path.dirname(input_path)
    if output_name:
        name = output_name
        if not os.path.splitext(name)[1]:
            name = f"{name}{preferred_ext}"
        return os.path.join(base_dir, name)

    stem = os.path.splitext(os.path.basename(input_path))[0]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(base_dir, f"{stem}_processed_{ts}{preferred_ext}")


def _write_xlsx(path: str, headers: List[str], rows: List[List[Any]], sheet_name: str = "Sheet1") -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Sheet1"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _write_table(path: str, headers: List[str], rows: List[List[Any]], sheet_name: str = "Sheet1") -> None:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        _write_csv(path, headers, rows)
        return
    if ext in {".xlsx", ".xlsm"}:
        _write_xlsx(path, headers, rows, sheet_name=sheet_name)
        return
    raise ValueError(f"Unsupported output extension: {ext}")


def _pick_output_ext(input_path: str, output_name: str, default_ext: str) -> str:
    if output_name:
        ext = os.path.splitext(output_name)[1].lower()
        if ext in {".csv", ".xlsx", ".xlsm"}:
            return ext
    in_ext = os.path.splitext(input_path)[1].lower()
    if in_ext in {".csv", ".xlsx", ".xlsm"}:
        return in_ext
    return default_ext


def _process_csv(path: str, instruction: str, output_name: str = "") -> Dict[str, Any]:
    headers, rows = _read_csv(path)
    if not headers:
        return {"success": False, "error": "CSV file has no rows.", "file_path": path}

    parsed = _parse_instruction(instruction)
    new_headers, new_rows, notes = _apply_ops(headers, rows, parsed)
    chosen_output_name = output_name or parsed.get("output_name", "")
    out_ext = _pick_output_ext(path, chosen_output_name, ".csv")
    output_path = _resolve_output_path(path, chosen_output_name, out_ext)
    _write_table(output_path, new_headers, new_rows, sheet_name="Processed")

    return {
        "success": True,
        "file_path": output_path,
        "source_file_path": path,
        "rows_before": len(rows),
        "rows_after": len(new_rows),
        "notes": notes,
        "columns": new_headers,
    }


def _process_xlsx(path: str, instruction: str, sheet_name: str = "", output_name: str = "") -> Dict[str, Any]:
    wb = load_workbook(path)
    parsed = _parse_instruction(instruction)
    requested_sheet = sheet_name or parsed.get("sheet_name", "")

    ws = None
    if requested_sheet:
        for s in wb.sheetnames:
            if s.strip().lower() == requested_sheet.strip().lower():
                ws = wb[s]
                break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    values = list(ws.iter_rows(values_only=True))
    if not values:
        return {"success": False, "error": "Excel sheet is empty.", "file_path": path, "sheet_name": ws.title}

    headers = [str(x or "").strip() for x in values[0]]
    rows = [list(r) for r in values[1:]]

    new_headers, new_rows, notes = _apply_ops(headers, rows, parsed)

    ws.delete_rows(1, ws.max_row)
    ws.append(new_headers)
    for row in new_rows:
        ws.append(row)

    chosen_output_name = output_name or parsed.get("output_name", "")
    out_ext = _pick_output_ext(path, chosen_output_name, ".xlsx")
    output_path = _resolve_output_path(path, chosen_output_name, out_ext)
    wb.save(output_path)
    wb.close()

    return {
        "success": True,
        "file_path": output_path,
        "source_file_path": path,
        "sheet_name": ws.title,
        "rows_before": len(rows),
        "rows_after": len(new_rows),
        "notes": notes,
        "columns": new_headers,
    }


def _process_zip(path: str, instruction: str, output_name: str = "") -> Dict[str, Any]:
    with zipfile.ZipFile(path, "r") as zf:
        members = [m for m in zf.namelist() if not m.endswith("/")]
        target = None
        for name in members:
            if name.lower().endswith(".csv"):
                target = name
                break
        if target is None:
            return {"success": False, "error": "ZIP file does not contain a CSV file.", "file_path": path}

        raw = zf.read(target)
        temp_csv = os.path.join(os.path.dirname(path), f"_tmp_{os.path.basename(target)}")
        with open(temp_csv, "wb") as handle:
            handle.write(raw)

    try:
        result = _process_csv(temp_csv, instruction=instruction, output_name=output_name)
        result["source_archive"] = path
        result["source_member"] = target
        return result
    finally:
        try:
            if os.path.isfile(temp_csv):
                os.remove(temp_csv)
        except Exception:
            pass


def run_excel_sheet_task(
    instruction: str,
    file_path: str = "",
    sheet_name: str = "",
    output_filename: str = "",
) -> Dict[str, Any]:
    try:
        selected = (file_path or "").strip()
        if selected and not os.path.isfile(selected):
            return {"success": False, "error": f"Excel file not found: {selected}", "file_path": selected}

        if not selected:
            selected = _find_latest_data_file() or ""
            if not selected:
                return {
                    "success": False,
                    "error": "No Excel/CSV export file found. Provide a file_path or place a file in Downloads.",
                }

        ext = os.path.splitext(selected)[1].lower()
        if ext == ".csv":
            return _process_csv(selected, instruction=instruction, output_name=output_filename)
        if ext in {".xlsx", ".xlsm"}:
            return _process_xlsx(selected, instruction=instruction, sheet_name=sheet_name, output_name=output_filename)
        if ext == ".zip":
            return _process_zip(selected, instruction=instruction, output_name=output_filename)

        return {"success": False, "error": f"Unsupported file extension: {ext}", "file_path": selected}
    except Exception as e:
        return {"success": False, "error": f"Excel task failed: {type(e).__name__}: {e}", "file_path": file_path}
